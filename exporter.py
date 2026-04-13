"""
exporter.py — Excel (.xlsx) and PDF export for CamMonitor
"""
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm

GREEN = "FF27AE60"
RED   = "FFC0392B"
AMBER = "FFE67E22"
GREY  = "FFF5F5F5"
HEADER= "FF2C3E50"

def export_cameras_excel(cameras):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Cameras"
    headers = ["Name","IP","Location","Zone","NVR","NVR IP","Channel","Brand","Status","Health %","Maintenance","Offline Since","Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.alignment = Alignment(horizontal="center")
    for row, c in enumerate(cameras, 2):
        status = "Maintenance" if c.get("maintenance") else ("Offline" if not c.get("online") else "Online")
        vals = [c.get("name",""), c.get("ip",""), c.get("location",""), c.get("zone",""),
                c.get("nvr_name",""), c.get("nvr_ip",""), c.get("nvr_channel",""), c.get("brand","").capitalize(),
                status, f"{c.get('health_7d',100):.0f}%",
                "Yes" if c.get("maintenance") else "No",
                str(c.get("offline_since",""))[:16] or "—", c.get("notes","")]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            if status == "Offline":
                cell.fill = PatternFill("solid", fgColor="FFFDECEA")
            elif status == "Maintenance":
                cell.fill = PatternFill("solid", fgColor="FFFFF9F0")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def export_offline_excel(cameras):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offline Cameras"
    ws["A1"] = f"Offline Camera Report — {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws["A1"].font = Font(bold=True, size=13)
    ws.merge_cells("A1:G1")
    headers = ["Name","IP","Location","Zone","NVR","Brand","Offline Since"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="FFC0392B")
    for row, c in enumerate(cameras, 4):
        vals = [c.get("name",""), c.get("ip",""), c.get("location",""), c.get("zone",""),
                c.get("nvr_name",""), c.get("brand","").capitalize(),
                str(c.get("offline_since",""))[:16] or "Unknown"]
        for col, val in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=val).fill = PatternFill("solid", fgColor="FFFDECEA")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or "")) for c in col) + 4
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def export_offline_pdf(cameras):
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph(f"Offline Camera Report — {datetime.now().strftime('%d-%b-%Y %H:%M')}", styles["Title"]))
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(f"Total offline: {len(cameras)}", styles["Normal"]))
    elements.append(Spacer(1, 0.5*cm))
    headers = ["Name","IP","Location","Zone","NVR","Brand","Offline Since"]
    data = [headers] + [
        [c.get("name",""), c.get("ip",""), c.get("location",""), c.get("zone",""),
         c.get("nvr_name",""), c.get("brand","").capitalize(),
         str(c.get("offline_since",""))[:16] or "Unknown"]
        for c in cameras
    ]
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2C3E50")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("BACKGROUND", (0,1), (-1,-1), colors.HexColor("#FDECEA")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.HexColor("#FDECEA"), colors.HexColor("#FAFAFA")]),
        ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))
    elements.append(t)
    doc.build(elements)
    out.seek(0)
    return out

def export_template_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Camera Import Template"
    headers = ["ip","name","location","zone","nvr_name","nvr_ip","nvr_channel","brand","username","password","notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER)
    example = ["192.168.1.101","CAM-01","Main Gate","Entry","NVR-01","192.168.1.200","1","hikvision","admin","admin123","Fixed dome"]
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val)
    brand_validation = DataValidation(
        type="list",
        formula1='"hikvision,dahua,prama,cpplus,other"',
        allow_blank=True
    )
    brand_validation.prompt = "Choose a supported camera brand."
    brand_validation.error = "Use one of: hikvision, dahua, prama, cpplus, other."
    ws.add_data_validation(brand_validation)
    brand_validation.add("H2:H5000")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
