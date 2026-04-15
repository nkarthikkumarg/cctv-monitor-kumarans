# -*- coding: utf-8 -*-
"""
app.py — Flask web application for CamMonitor
Run: python app.py
"""
import io
import os
import logging
import logging.handlers
import configparser
import json
import base64
import hashlib
import hmac
import ipaddress
import secrets as _secrets_mod
import ssl
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from functools import wraps

import certifi
from flask import (Flask, render_template, render_template_string, request, jsonify,
                   Response, stream_with_context,
                   redirect, url_for, session, send_file, flash)
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.middleware.proxy_fix import ProxyFix

import db
import monitor
import exporter
import central_sync
import alerts
from notification_settings import load_settings as load_notification_file, save_settings as save_notification_file
from preview import get_stream_urls, normalize_rtsp_url

# ── Setup ─────────────────────────────────────────────────────────────────────
_LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "monitor.log")
_LOG_FMT = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
_fh = logging.handlers.RotatingFileHandler(
    _LOG_FILE, maxBytes=10 * 1024 * 1024, backupCount=5, encoding="utf-8"
)
_fh.setFormatter(_LOG_FMT)
_sh = logging.StreamHandler()
_sh.setFormatter(_LOG_FMT)
logging.basicConfig(level=logging.INFO, handlers=[_fh, _sh])
log = logging.getLogger(__name__)

# ── Login rate limiting ────────────────────────────────────────────────────────
_login_attempts: dict = defaultdict(list)
_login_lock = threading.Lock()
_LOGIN_WINDOW = 300   # 5-minute sliding window
_LOGIN_MAX = 20       # max attempts per IP in that window

def _check_login_rate_limit(ip: str) -> bool:
    """Return True if this IP is rate-limited (too many recent failures)."""
    now = time.monotonic()
    with _login_lock:
        _login_attempts[ip] = [t for t in _login_attempts[ip] if now - t < _LOGIN_WINDOW]
        if len(_login_attempts[ip]) >= _LOGIN_MAX:
            return True
        _login_attempts[ip].append(now)
        return False

def _clear_login_attempts(ip: str) -> None:
    with _login_lock:
        _login_attempts.pop(ip, None)

cfg = configparser.ConfigParser()
CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config.ini")
cfg.read(CONFIG_PATH)

WEB_HOST = "0.0.0.0"
WEB_PORT = 5000
SECRET = "dev-secret"
SETUP_COMPLETE = False
USERNAME = "admin"
PASSWORD = "admin123"
ADMIN_DISPLAY_NAME = "Administrator"
CENTRAL_API_KEY = "local-dev-key"
SITE_NAME = "Local Site"
SSL_CONTEXT = ssl.create_default_context(cafile=certifi.where())
GO2RTC_ENABLED = True
GO2RTC_BASE_URL = "http://127.0.0.1:1984"
GO2RTC_PROXY_PATH = "/go2rtc"
ALLOWED_CAMERA_BRANDS = ("hikvision", "dahua", "prama", "cpplus", "other")
WEAK_SECRETS = {"", "dev-secret", "change-this-to-a-random-secret-key"}
WEAK_ADMIN_CREDENTIALS = {("admin", "admin123")}


def _b64url_decode(data):
    data = (data or "").encode("ascii")
    data += b"=" * (-len(data) % 4)
    return base64.urlsafe_b64decode(data)


def verify_sso_token(token):
    try:
        payload_b64, signature_b64 = (token or "").split(".", 1)
        expected = hmac.new(CENTRAL_API_KEY.encode("utf-8"), payload_b64.encode("utf-8"), hashlib.sha256).digest()
        provided = _b64url_decode(signature_b64)
        if not hmac.compare_digest(expected, provided):
            log.warning("SSO signature mismatch (local key starts with %s…)", CENTRAL_API_KEY[:8])
            return None
        payload = json.loads(_b64url_decode(payload_b64).decode("utf-8"))
        exp = int(payload.get("exp") or 0)
        now = int(datetime.now(timezone.utc).timestamp())
        if not exp or exp < now:
            log.warning("SSO token expired: exp=%s now=%s (delta=%ss)", exp, now, now - exp)
            return None
        username = (payload.get("username") or "").strip()
        if not username:
            log.warning("SSO token has no username")
            return None
        return payload
    except Exception as exc:
        log.warning("SSO token parse error: %s", exc)
        return None


def reload_runtime_settings():
    global WEB_HOST, WEB_PORT, SECRET, USERNAME, PASSWORD, CENTRAL_API_KEY, SITE_NAME, ADMIN_DISPLAY_NAME, GO2RTC_ENABLED, GO2RTC_BASE_URL, GO2RTC_PROXY_PATH
    cfg.clear()
    cfg.read(CONFIG_PATH)
    db.reload_config()
    db.init_db()
    WEB_HOST = cfg.get("web", "host", fallback="0.0.0.0")
    WEB_PORT = int(os.environ.get("PORT") or cfg.getint("web", "port", fallback=5000))
    SECRET = cfg.get("web", "secret_key", fallback="dev-secret")
    USERNAME = cfg.get("web", "dashboard_username", fallback="admin")
    PASSWORD = cfg.get("web", "dashboard_password", fallback="admin123")
    ADMIN_DISPLAY_NAME = cfg.get("web", "dashboard_admin_name", fallback="Administrator").strip() or "Administrator"
    CENTRAL_API_KEY = cfg.get("central", "api_key", fallback="local-dev-key")
    SITE_NAME = cfg.get("central", "site_name", fallback="Local Site").strip() or "Local Site"
    GO2RTC_ENABLED = cfg.getboolean("go2rtc", "enabled", fallback=True)
    GO2RTC_BASE_URL = cfg.get("go2rtc", "base_url", fallback="http://127.0.0.1:1984").rstrip("/")
    GO2RTC_PROXY_PATH = cfg.get("go2rtc", "proxy_path", fallback="/go2rtc").rstrip("/") or "/go2rtc"
    db.ensure_default_admin(USERNAME, PASSWORD, ADMIN_DISPLAY_NAME)
    load_setup_flag()
    if "app" in globals():
        app.secret_key = SECRET


def get_runtime_warnings():
    warnings = []
    if SECRET in WEAK_SECRETS or len(SECRET) < 32:
        warnings.append("web.secret_key must be replaced with a strong random value")
    if ((USERNAME or "").strip().lower(), PASSWORD or "") in WEAK_ADMIN_CREDENTIALS:
        warnings.append("default admin username/password must be changed")
    if not db.DB_PATH or not os.path.isabs(db.DB_PATH):
        warnings.append("database.db_path must resolve to an absolute filesystem path")
    if GO2RTC_ENABLED and not GO2RTC_BASE_URL.startswith("http://127.0.0.1:"):
        warnings.append("go2rtc.base_url should stay on localhost in production")
    return warnings


def validate_runtime_settings(strict=False):
    warnings = get_runtime_warnings()
    if warnings:
        for item in warnings:
            log.warning("Runtime configuration warning: %s", item)
        if strict:
            raise RuntimeError("Production startup blocked: " + "; ".join(warnings))
    return warnings

def load_setup_flag():
    global SETUP_COMPLETE
    SETUP_COMPLETE = cfg.getboolean("setup", "completed", fallback=False)

def mark_setup_complete(value=True):
    if not cfg.has_section("setup"):
        cfg.add_section("setup")
    cfg.set("setup", "completed", "true" if value else "false")
    with open(CONFIG_PATH, "w", encoding="utf-8") as fh:
        cfg.write(fh)
    load_setup_flag()


def get_go2rtc_source(cam):
    if cam.get("rtsp_url"):
        return normalize_rtsp_url(cam["rtsp_url"])
    urls = get_stream_urls(
        cam.get("ip", ""),
        cam.get("brand", ""),
        cam.get("username", ""),
        cam.get("password", ""),
    )
    return urls.get("rtsp", "")


def get_go2rtc_urls(cam):
    if not GO2RTC_ENABLED:
        return None
    src = get_go2rtc_source(cam)
    if not src:
        return None
    encoded = urllib.parse.quote(src, safe="")
    return {
        "upstream_mjpeg": f"{GO2RTC_BASE_URL}/api/stream.mjpeg?src={encoded}",
        "upstream_snapshot": f"{GO2RTC_BASE_URL}/api/frame.jpeg?src={encoded}",
    }


def get_go2rtc_player_path(cam):
    if not GO2RTC_ENABLED:
        return ""
    src = get_go2rtc_source(cam)
    if not src:
        return ""
    query = urllib.parse.urlencode(
        {
            "src": src,
            "mode": "webrtc,mse",
            "background": "false",
        }
    )
    return f"{GO2RTC_PROXY_PATH}/stream.html?{query}"


def get_go2rtc_local_player_path(cam):
    if not GO2RTC_ENABLED:
        return ""
    src = get_go2rtc_source(cam)
    if not src:
        return ""
    query = urllib.parse.urlencode(
        {
            "src": src,
            "mode": "webrtc,mse",
            "background": "false",
        }
    )
    return f"{GO2RTC_BASE_URL}/stream.html?{query}"


def get_site_settings():
    dashboard_url = cfg.get("central", "dashboard_url", fallback="").strip()
    if not dashboard_url:
        host = cfg.get("web", "host", fallback="127.0.0.1")
        port = cfg.getint("web", "port", fallback=5000)
        host = "127.0.0.1" if host in ("0.0.0.0", "::") else host
        dashboard_url = f"http://{host}:{port}"
    return {
        "enabled": cfg.getboolean("central", "enabled", fallback=False),
        "site_id": cfg.get("central", "site_id", fallback="local-site").strip(),
        "site_name": cfg.get("central", "site_name", fallback="Local Site").strip() or "Local Site",
        "campus": cfg.get("central", "campus", fallback="").strip(),
        "site_address": cfg.get("central", "site_address", fallback="").strip(),
        "contact_name": cfg.get("central", "contact_name", fallback="").strip(),
        "contact_phone": cfg.get("central", "contact_phone", fallback="").strip(),
        "contact_email": cfg.get("central", "contact_email", fallback="").strip(),
        "dashboard_url": dashboard_url,
        "refresh_url": cfg.get("central", "refresh_url", fallback="").strip(),
        "api_url": cfg.get("central", "api_url", fallback="").strip(),
        "api_key": cfg.get("central", "api_key", fallback="local-dev-key"),
    }


def get_registration_url(api_url):
    api_url = (api_url or "").strip()
    if api_url.endswith("/api/site-summary"):
        return api_url[:-len("/api/site-summary")] + "/api/register-site"
    return ""


def save_site_settings(data):
    previous = get_site_settings()
    parser = configparser.ConfigParser()
    parser.read(CONFIG_PATH)
    if not parser.has_section("central"):
        parser.add_section("central")
    values = {
        "enabled": "true" if data.get("enabled") else "false",
        "site_id": (data.get("site_id") or "").strip(),
        "site_name": (data.get("site_name") or "").strip() or "Local Site",
        "campus": (data.get("campus") or "").strip(),
        "site_address": (data.get("site_address") or "").strip(),
        "contact_name": (data.get("contact_name") or "").strip(),
        "contact_phone": (data.get("contact_phone") or "").strip(),
        "contact_email": (data.get("contact_email") or "").strip(),
        "dashboard_url": (data.get("dashboard_url") or "").strip(),
        "refresh_url": (data.get("refresh_url") or "").strip(),
        "api_url": (data.get("api_url") or "").strip(),
        "api_key": data.get("api_key") or "local-dev-key",
    }
    for key, value in values.items():
        parser.set("central", key, value)
    with open(CONFIG_PATH, "w", encoding="utf-8") as fh:
        parser.write(fh)
    reload_runtime_settings()
    central_sync.reload_config()
    if previous["site_id"] and previous["site_id"] != values["site_id"]:
        central_sync.delete_remote_site(previous["site_id"], previous["api_url"], previous["api_key"])
    return get_site_settings()


def get_notification_settings():
    stored = load_notification_file()
    return {
        "poll_interval": cfg.getint("monitor", "poll_interval", fallback=10),
        "status_retries": cfg.getint(
            "monitor",
            "status_retries",
            fallback=cfg.getint("monitor", "ping_retries", fallback=2),
        ),
        "alert_ping_retries": cfg.getint("monitor", "alert_ping_retries", fallback=6),
        "alert_cooldown_minutes": cfg.getint("monitor", "alert_cooldown_minutes", fallback=30),
        "notify_offline": cfg.getboolean("notifications", "notify_offline", fallback=True),
        "notify_recovery": cfg.getboolean("notifications", "notify_recovery", fallback=True),
        "daily_summary_enabled": cfg.getboolean("notifications", "daily_summary_enabled", fallback=True),
        "daily_report_time": cfg.get("email", "daily_report_time", fallback="08:00"),
        "email_enabled": cfg.getboolean("email", "enabled", fallback=False),
        "smtp_host": cfg.get("email", "smtp_host", fallback="smtp.gmail.com"),
        "smtp_port": cfg.getint("email", "smtp_port", fallback=587),
        "smtp_use_tls": cfg.getboolean("email", "smtp_use_tls", fallback=True),
        "sender_email": cfg.get("email", "sender_email", fallback=""),
        "sender_password": cfg.get("email", "sender_password", fallback=""),
        "subject_prefix": cfg.get("email", "subject_prefix", fallback="[CAM ALERT]"),
        "whatsapp_enabled": cfg.getboolean("whatsapp", "enabled", fallback=False),
        "account_sid": cfg.get("whatsapp", "account_sid", fallback=""),
        "auth_token": cfg.get("whatsapp", "auth_token", fallback=""),
        "from_number": cfg.get("whatsapp", "from_number", fallback=""),
        "greeting_template": stored.get("greeting_template", "Dear {name},"),
        "templates": stored.get("templates", {}),
        "recipients": stored.get("recipients", []),
    }


def save_notification_settings(data):
    parser = configparser.ConfigParser()
    parser.read(CONFIG_PATH)
    for section in ("monitor", "email", "whatsapp", "notifications"):
        if not parser.has_section(section):
            parser.add_section(section)

    parser.set("monitor", "alert_cooldown_minutes", str(int(data.get("alert_cooldown_minutes") or 30)))
    parser.set("monitor", "status_retries", str(int(data.get("status_retries") or 1)))
    parser.set("monitor", "alert_ping_retries", str(int(data.get("alert_ping_retries") or 6)))
    parser.set("monitor", "poll_interval", str(int(data.get("poll_interval") or 10)))

    parser.set("notifications", "notify_offline", "true" if data.get("notify_offline") else "false")
    parser.set("notifications", "notify_recovery", "true" if data.get("notify_recovery") else "false")
    parser.set("notifications", "daily_summary_enabled", "true" if data.get("daily_summary_enabled") else "false")

    parser.set("email", "enabled", "true" if data.get("email_enabled") else "false")
    parser.set("email", "smtp_host", (data.get("smtp_host") or "").strip())
    parser.set("email", "smtp_port", str(int(data.get("smtp_port") or 587)))
    parser.set("email", "smtp_use_tls", "true" if data.get("smtp_use_tls") else "false")
    parser.set("email", "sender_email", (data.get("sender_email") or "").strip())
    parser.set("email", "sender_password", data.get("sender_password") or "")
    parser.set("email", "subject_prefix", (data.get("subject_prefix") or "[CAM ALERT]").strip() or "[CAM ALERT]")
    parser.set("email", "daily_report_time", (data.get("daily_report_time") or "08:00").strip() or "08:00")

    parser.set("whatsapp", "enabled", "true" if data.get("whatsapp_enabled") else "false")
    parser.set("whatsapp", "account_sid", (data.get("account_sid") or "").strip())
    parser.set("whatsapp", "auth_token", (data.get("auth_token") or "").strip())
    parser.set("whatsapp", "from_number", (data.get("from_number") or "").strip())

    with open(CONFIG_PATH, "w", encoding="utf-8") as fh:
        parser.write(fh)

    save_notification_file({
        "greeting_template": data.get("greeting_template") or "Dear {name},",
        "templates": data.get("templates") or {},
        "recipients": data.get("recipients") or [],
    })

    reload_runtime_settings()
    central_sync.reload_config()
    alerts.reload_settings()
    monitor.reload_settings()
    monitor.reschedule_jobs()
    return get_notification_settings()


reload_runtime_settings()

app = Flask(__name__)
app.secret_key = SECRET
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    # Only send session cookie over HTTPS.
    # Set https_only = true in config.ini [web] once Cloudflare Tunnel / Caddy TLS is confirmed.
    # Leave false (default) for plain HTTP local installs — required for setup to work over HTTP.
    SESSION_COOKIE_SECURE=cfg.getboolean("web", "https_only", fallback=False),
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
)
login_manager = LoginManager(app)
login_manager.login_view = "login"

# ── CSRF helpers ──────────────────────────────────────────────────────────────
def _get_csrf_token() -> str:
    from flask import session as _sess
    if "_csrf" not in _sess:
        _sess["_csrf"] = _secrets_mod.token_hex(32)
    return _sess["_csrf"]

def _validate_csrf() -> bool:
    from flask import session as _sess
    expected = _sess.get("_csrf")
    if not expected:
        return False
    provided = (request.headers.get("X-CSRF-Token")
                or (request.get_json(silent=True) or {}).get("_csrf")
                or request.form.get("_csrf"))
    return bool(provided and hmac.compare_digest(expected, provided))

# Endpoints exempt from CSRF (use their own auth: API key or no state change)
_CSRF_EXEMPT = frozenset({
    "api_central_refresh", "api_central_sync_users",
    "api_bulk_preview",   # read-only preview
    "sso_login",
    "login", "logout",    # login has its own token in the form
})

@app.before_request
def _enforce_csrf():
    if request.method in ("GET", "HEAD", "OPTIONS"):
        return
    if request.endpoint in _CSRF_EXEMPT:
        return
    # API-key-authenticated endpoints skip CSRF
    if request.headers.get("X-API-Key"):
        return
    if not _validate_csrf():
        if request.path.startswith("/api/"):
            return jsonify({"error": "CSRF token missing or invalid"}), 403
        return redirect(url_for("login"))

# ── Security response headers ─────────────────────────────────────────────────
@app.after_request
def _add_security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    # Allow go2rtc origin in frame-src, img-src, connect-src so the live preview
    # iframe and WebRTC/WebSocket signalling are not blocked by CSP.
    _g2 = GO2RTC_BASE_URL.rstrip("/") if GO2RTC_ENABLED else ""
    _g2_ws = _g2.replace("http://", "ws://").replace("https://", "wss://") if _g2 else ""
    _frame = f"'self' {_g2}" if _g2 else "'self'"
    _img   = f"'self' data: blob: https: {_g2}" if _g2 else "'self' data: blob: https:"
    _conn  = f"'self' {_g2} {_g2_ws}" if _g2 else "'self'"
    response.headers["Content-Security-Policy"] = (
        f"default-src 'self'; "
        f"script-src 'self' 'unsafe-inline'; "
        f"style-src 'self' 'unsafe-inline'; "
        f"img-src {_img}; "
        f"frame-src {_frame}; "
        f"connect-src {_conn}"
    )
    return response

class User(UserMixin):
    def __init__(self, row):
        self.user_id = row["id"]
        self.id = row["username"]
        self.username = row["username"]
        self.display_name = row.get("display_name") or row["username"]
        self.role = row["role"]
        self.active = bool(row.get("active", 1))

    def get_id(self):
        return str(self.user_id)

@login_manager.user_loader
def load_user(uid):
    row = db.get_user_by_id(uid)
    return User(row) if row and row.get("active") else None


ROLE_RANK = {"viewer": 1, "operator": 2, "admin": 3}

@app.before_request
def ensure_setup_complete():
    if request.endpoint is None:
        return
    allowed = {
        "static", "login", "logout", "sso_login",
        "setup", "api_setup_complete",
        "api_site_settings", "api_site_settings_register",
        "api_bulk_preview", "api_bulk_import",
        "central_sync", "api_audit"
    }
    if request.endpoint in allowed:
        return
    if not SETUP_COMPLETE and current_user.is_authenticated:
        return redirect(url_for("setup"))


def role_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapped(*args, **kwargs):
            if not current_user.is_authenticated:
                return login_manager.unauthorized()
            current_rank = ROLE_RANK.get(getattr(current_user, "role", "viewer"), 0)
            needed_rank = max(ROLE_RANK.get(role, 0) for role in roles)
            if current_rank < needed_rank:
                if request.path.startswith("/api/"):
                    return jsonify({"error": "Forbidden"}), 403
                flash("You do not have permission to access that page.")
                return redirect(url_for("dashboard"))
            return view(*args, **kwargs)
        return wrapped
    return decorator


def actor_name():
    return getattr(current_user, "display_name", getattr(current_user, "username", "system"))


def can_use_cached_central_user(user):
    if not user or user.get("source") != "central":
        return True
    if user.get("central_deleted") or not user.get("active"):
        return False
    valid_until = user.get("valid_until")
    if not valid_until:
        return False
    try:
        expires_at = datetime.fromisoformat(valid_until.replace("Z", "+00:00"))
        now = datetime.now(timezone.utc) if expires_at.tzinfo else datetime.now()
        return expires_at >= now
    except Exception:
        return False


def verify_login_user(username, password):
    user = db.get_user_by_username(username)
    if not user:
        return None
    if user.get("source") == "central":
        # Prefer fresh central truth whenever reachable so deletes/role changes
        # take effect immediately, but keep the offline-valid cache as fallback.
        central_sync.sync_users()
        user = db.get_user_by_username(username)
        if not user:
            return None
        if not can_use_cached_central_user(user):
            return None
    return db.verify_user(username, password)

# ── Auth ──────────────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET","POST"])
def login():
    error = request.args.get("sso_error")
    csrf_token = _get_csrf_token()
    if request.method == "POST":
        u = request.form.get("username","")
        p = request.form.get("password","")
        ip = request.remote_addr
        # CSRF check for login form
        provided_csrf = request.form.get("_csrf", "")
        if not (provided_csrf and hmac.compare_digest(csrf_token, provided_csrf)):
            error = "Invalid request. Please reload and try again."
            return render_template_string(LOGIN_HTML, error=error, csrf_token=csrf_token)
        if _check_login_rate_limit(ip):
            db.add_audit(u, "login", "Login blocked — rate limit exceeded", "System", ip, "failed")
            error = "Too many login attempts. Please wait a few minutes and try again."
            return render_template_string(LOGIN_HTML, error=error, csrf_token=csrf_token)
        user = verify_login_user(u, p)
        if user:
            _clear_login_attempts(ip)
            session.permanent = True
            login_user(User(user), remember=True)
            db.add_audit(u, "login", "User logged in successfully", "System", ip, "success")
            if not SETUP_COMPLETE:
                return redirect(url_for("setup"))
            return redirect(url_for("dashboard"))
        else:
            db.add_audit(u, "login", "Failed login attempt", "System", ip, "failed")
            error = "Invalid username or password"
    return render_template_string(LOGIN_HTML, error=error, csrf_token=csrf_token)


@app.route("/sso-login")
def sso_login():
    import traceback as _tb
    def _sso_fail(reason):
        return redirect(url_for("login", sso_error=reason))
    try:
        token = request.args.get("token", "")
        log.debug("SSO attempt: token_present=%s central_api_key_len=%d", bool(token), len(CENTRAL_API_KEY))
        if not token:
            return _sso_fail("No SSO token provided.")
        payload = verify_sso_token(token)
        if not payload:
            log.warning("SSO failed: invalid or expired token (key=%s...)", CENTRAL_API_KEY[:6])
            return _sso_fail("SSO token is invalid or expired. Please try again from the central dashboard.")
        username = payload.get("username", "")
        log.debug("SSO token valid for username=%s role=%s", username, payload.get("role"))
        central_sync.sync_users()
        user = db.get_user_by_username(username)
        if not user:
            log.warning("SSO failed: username '%s' not found in local DB", username)
            return _sso_fail(f"User '{username}' not found on this site. Please sync users from the central dashboard.")
        if user.get("source") == "central":
            if not can_use_cached_central_user(user):
                log.warning("SSO failed: central user '%s' cache expired or inactive", username)
                return _sso_fail(f"User '{username}' session has expired or is inactive. Please sync users.")
        elif not (
            user.get("source") == "local"
            and user.get("role") == "admin"
            and (user.get("username") or "").strip().lower() == USERNAME.strip().lower()
            and (payload.get("role") or "").strip().lower() == "admin"
        ):
            log.warning("SSO failed: local user '%s' not eligible (source=%s role=%s)", username, user.get("source"), user.get("role"))
            return _sso_fail("Your account is not eligible for single sign-on at this site.")
        session.permanent = True
        login_user(User(user), remember=True)
        db.add_audit(
            actor_name(),
            "login",
            f"Single sign-on login from central dashboard for {user.get('display_name') or user.get('username')}",
            "System",
            request.remote_addr,
            "success",
        )
        if not SETUP_COMPLETE:
            return redirect(url_for("setup"))
        return redirect(url_for("dashboard"))
    except Exception:
        log.error("SSO internal error:\n%s", _tb.format_exc())
        raise

@app.route("/logout")
@login_required
def logout():
    db.add_audit(actor_name(), "logout", "User logged out", "System", request.remote_addr, "success")
    logout_user()
    return redirect(url_for("login"))

# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route("/")
@login_required
def dashboard():
    zones = db.get_zones()
    locations = db.get_locations()
    nvrs  = db.get_nvrs()
    return render_template_string(
        DASHBOARD_HTML,
        zones=zones,
        locations=locations,
        nvrs=nvrs,
        user=actor_name(),
        role=current_user.role,
        site_name=SITE_NAME,
        csrf_token=_get_csrf_token(),
    )

# ── API ───────────────────────────────────────────────────────────────────────
@app.route("/api/health")
def api_health():
    warnings = get_runtime_warnings()
    status_code = 200 if not warnings else 503
    resp = {"ok": not warnings, "site_name": SITE_NAME}
    # Only expose config warnings to authenticated admins
    if current_user.is_authenticated and getattr(current_user, "role", "") == "admin":
        resp["warnings"] = warnings
    return jsonify(resp), status_code

@app.route("/api/cameras")
@login_required
def api_cameras():
    cameras = db.get_all_cameras()
    zone    = request.args.get("zone")
    location = request.args.get("location")
    nvr     = request.args.get("nvr")
    brand   = request.args.get("brand")
    status  = request.args.get("status")
    q       = request.args.get("q","").lower()
    if zone:   cameras = [c for c in cameras if c.get("zone") == zone]
    if location: cameras = [c for c in cameras if c.get("location") == location]
    if nvr:    cameras = [c for c in cameras if c.get("nvr_name") == nvr]
    if brand:  cameras = [c for c in cameras if c.get("brand","").lower() == brand.lower()]
    if status == "online":      cameras = [c for c in cameras if c.get("online") and not c.get("maintenance")]
    elif status == "offline":   cameras = [c for c in cameras if not c.get("online")]
    elif status == "maintenance": cameras = [c for c in cameras if c.get("maintenance")]
    if q: cameras = [c for c in cameras if q in c.get("name","").lower() or q in c.get("ip","")]
    total = len(cameras)
    try:
        page = max(1, int(request.args.get("page") or 1))
        page_size = min(400, max(50, int(request.args.get("page_size") or 100)))
    except (ValueError, TypeError):
        page, page_size = 1, 100
    start = (page - 1) * page_size
    end = start + page_size
    items = cameras[start:end]
    # Strip sensitive credential fields from list response for non-admins
    if getattr(current_user, "role", "") != "admin":
        items = [{k: v for k, v in c.items() if k not in ("password", "username")} for c in items]
    return jsonify({
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    })

@app.route("/api/stats")
@login_required
def api_stats():
    stats = db.get_stats()
    stats["central"] = central_sync.get_status()
    return jsonify(stats)

@app.route("/api/camera/<ip>")
@login_required
def api_camera_detail(ip):
    cam = db.get_camera(ip)
    if not cam: return jsonify({"error": "Not found"}), 404
    cam["history"] = db.get_camera_history(ip, 10)
    urls = get_stream_urls(ip, cam.get("brand",""), cam.get("username",""), cam.get("password",""))
    if cam.get("rtsp_url"):
        urls["rtsp"] = normalize_rtsp_url(cam["rtsp_url"])
    go2rtc_urls = get_go2rtc_urls(cam)
    if go2rtc_urls:
        urls["browser"] = f"/api/camera/{urllib.parse.quote(ip, safe='')}/live.mjpeg"
        urls["snapshot"] = f"/api/camera/{urllib.parse.quote(ip, safe='')}/snapshot.jpg"
        urls["player"] = get_go2rtc_player_path(cam)
        urls["player_local"] = get_go2rtc_local_player_path(cam)
    cam["stream_urls"] = urls
    # Strip credentials from detail response for non-admins
    if getattr(current_user, "role", "") != "admin":
        cam.pop("password", None)
        cam.pop("username", None)
        # Remove RTSP URL (contains credentials) for non-admins
        cam.get("stream_urls", {}).pop("rtsp", None)
    return jsonify(cam)


@app.route("/api/camera/<ip>/live.mjpeg")
@login_required
def api_camera_live_mjpeg(ip):
    cam = db.get_camera(ip)
    if not cam:
        return jsonify({"error": "Not found"}), 404
    go2rtc_urls = get_go2rtc_urls(cam)
    if not go2rtc_urls:
        return jsonify({"error": "Live stream bridge is unavailable"}), 503

    try:
        upstream = urllib.request.urlopen(go2rtc_urls["upstream_mjpeg"], timeout=10)
    except Exception as exc:
        log.warning("Live MJPEG proxy failed for %s: %s", ip, exc)
        return jsonify({"error": "Could not open live stream"}), 502

    def generate():
        try:
            while True:
                chunk = upstream.read(8192)
                if not chunk:
                    break
                yield chunk
        finally:
            upstream.close()

    content_type = upstream.headers.get("Content-Type", "multipart/x-mixed-replace")
    headers = {
        "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        "Pragma": "no-cache",
        "Expires": "0",
        "X-Accel-Buffering": "no",
    }
    return Response(stream_with_context(generate()), content_type=content_type, headers=headers)


@app.route("/api/camera/<ip>/warm", methods=["POST"])
@login_required
def api_warm_camera(ip):
    """Pre-warm the go2rtc RTSP connection for a camera.

    Called on hover/click before the player iframe loads. Triggers go2rtc to
    open the RTSP connection immediately so WebRTC negotiation can start sooner,
    reducing time-to-first-frame from ~2-3s to ~0.5-1s.
    Returns immediately — the actual warm-up happens in a background thread.
    """
    cam = db.get_camera(ip)
    if not cam:
        return jsonify({"ok": False}), 404
    urls = get_go2rtc_urls(cam)
    if not urls:
        return jsonify({"ok": False}), 404
    mjpeg_url = urls["upstream_mjpeg"]

    def _do_warm():
        try:
            # A short read is enough to trigger go2rtc's RTSP connection.
            # Timeout intentionally short — we just need to wake the stream.
            with urllib.request.urlopen(mjpeg_url, timeout=2) as resp:
                resp.read(4096)
        except Exception:
            pass  # timeout / error expected — go2rtc is now connecting in background

    threading.Thread(target=_do_warm, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/camera/<ip>/snapshot.jpg")
@login_required
def api_camera_snapshot(ip):
    cam = db.get_camera(ip)
    if not cam:
        return jsonify({"error": "Not found"}), 404
    go2rtc_urls = get_go2rtc_urls(cam)
    snapshot_url = go2rtc_urls["upstream_snapshot"] if go2rtc_urls else get_stream_urls(
        ip,
        cam.get("brand", ""),
        cam.get("username", ""),
        cam.get("password", ""),
    ).get("snapshot", "")
    if not snapshot_url:
        return jsonify({"error": "Snapshot unavailable"}), 503
    try:
        with urllib.request.urlopen(snapshot_url, timeout=10) as upstream:
            payload = upstream.read()
            content_type = upstream.headers.get("Content-Type", "image/jpeg")
    except Exception as exc:
        log.warning("Snapshot proxy failed for %s: %s", ip, exc)
        return jsonify({"error": "Could not load snapshot"}), 502
    return Response(payload, content_type=content_type)

@app.route("/api/camera/<ip>/maintenance", methods=["POST"])
@login_required
@role_required("operator")
def api_set_maintenance(ip):
    state = request.json.get("state", False)
    db.set_maintenance(ip, state, actor_name())
    return jsonify({"success": True, "ip": ip, "maintenance": state})

@app.route("/api/bulk/maintenance", methods=["POST"])
@login_required
@role_required("operator")
def api_bulk_maintenance():
    ips = request.json.get("ips", [])
    state = request.json.get("state", False)
    db.bulk_set_maintenance(ips, state, actor_name())
    return jsonify({"success": True, "count": len(ips)})

@app.route("/api/bulk/zone", methods=["POST"])
@login_required
@role_required("admin")
def api_bulk_zone():
    ips = request.json.get("ips", [])
    zone = request.json.get("zone", "")
    db.bulk_update_zone(ips, zone, actor_name())
    return jsonify({"success": True, "count": len(ips)})

@app.route("/api/bulk/nvr", methods=["POST"])
@login_required
@role_required("admin")
def api_bulk_nvr():
    ips = request.json.get("ips", [])
    nvr = request.json.get("nvr", "")
    db.bulk_update_nvr(ips, nvr, actor_name())
    return jsonify({"success": True, "count": len(ips)})


@app.route("/api/bulk/delete", methods=["POST"])
@login_required
@role_required("admin")
def api_bulk_delete():
    ips = [str(ip).strip() for ip in (request.json or {}).get("ips", []) if str(ip).strip()]
    if not ips:
        return jsonify({"error": "No cameras selected"}), 400

    deleted = []
    for ip in ips:
        cam = db.get_camera(ip)
        if not cam:
            continue
        db.deactivate_camera(ip)
        deleted.append({"ip": ip, "name": cam.get("name") or ip})

    if not deleted:
        return jsonify({"error": "No matching cameras found"}), 404

    db.add_audit(
        actor_name(),
        "bulk",
        f"Deleted {len(deleted)} cameras",
        ", ".join(item["ip"] for item in deleted[:10]) + ("..." if len(deleted) > 10 else ""),
        request.remote_addr,
        "success"
    )
    return jsonify({"success": True, "count": len(deleted), "deleted": deleted})

@app.route("/api/audit")
@login_required
def api_audit():
    event_type = request.args.get("type")
    search = request.args.get("q")
    result = request.args.get("result")
    audit_user = request.args.get("user")
    preset = request.args.get("preset", "7d")
    try:
        page = max(1, int(request.args.get("page") or 1))
        page_size = min(100, max(10, int(request.args.get("page_size") or 25)))
    except (ValueError, TypeError):
        page, page_size = 1, 25
    date_from, date_to = resolve_log_range(preset, request.args.get("from"), request.args.get("to"))
    if current_user.role != "admin":
        audit_user = actor_name()
    logs = db.get_audit_log(
        limit=page_size,
        offset=(page - 1) * page_size,
        event_type=event_type,
        search=search,
        result=result,
        user=audit_user,
        date_from=date_from,
        date_to=date_to,
    )
    return jsonify({
        "items": logs["items"],
        "total": logs["total"],
        "page": page,
        "page_size": page_size,
        "date_from": date_from,
        "date_to": date_to,
        "preset": preset,
    })


@app.route("/api/audit-users")
@login_required
@role_required("admin")
def api_audit_users():
    return jsonify(db.get_audit_users())

@app.route("/api/camera-logs")
@login_required
def api_camera_logs():
    event_type = request.args.get("event")
    zone = request.args.get("zone")
    nvr = request.args.get("nvr")
    search = request.args.get("q")
    preset = request.args.get("preset", "7d")
    try:
        page = max(1, int(request.args.get("page") or 1))
        page_size = min(100, max(10, int(request.args.get("page_size") or 25)))
    except (ValueError, TypeError):
        page, page_size = 1, 25
    date_from, date_to = resolve_log_range(preset, request.args.get("from"), request.args.get("to"))
    logs = db.get_camera_event_log(
        limit=page_size,
        offset=(page - 1) * page_size,
        event_type=event_type,
        zone=zone,
        nvr=nvr,
        search=search,
        date_from=date_from,
        date_to=date_to,
    )
    return jsonify({
        "items": logs["items"],
        "total": logs["total"],
        "page": page,
        "page_size": page_size,
        "date_from": date_from,
        "date_to": date_to,
        "preset": preset,
    })


def resolve_report_range(preset, from_str, to_str):
    today = datetime.now().date()
    if preset == "7d":
        start = today - timedelta(days=6)
        end = today
    elif preset == "30d":
        start = today - timedelta(days=29)
        end = today
    elif preset == "90d":
        start = today - timedelta(days=89)
        end = today
    elif preset == "3m":
        start = today - timedelta(days=89)
        end = today
    elif preset == "custom" and from_str and to_str:
        try:
            start = datetime.strptime(from_str, "%Y-%m-%d").date()
            end = datetime.strptime(to_str, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            start = today - timedelta(days=29)
            end = today
        if end < start:
            start, end = end, start
    else:
        start = today - timedelta(days=29)
        end = today
    return start.isoformat(), end.isoformat()


def _read_bulk_rows(upload_file):
    fname = (upload_file.filename or "").lower()
    rows = []
    if fname.endswith(".csv"):
        import csv, io as _io
        content = upload_file.read().decode("utf-8")
        rows = list(csv.DictReader(_io.StringIO(content)))
    elif fname.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(upload_file)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, [str(v or "") for v in row])))
    else:
        raise ValueError("Only .csv or .xlsx supported")
    return rows


def _is_valid_ip(value):
    try:
        ipaddress.ip_address((value or "").strip())
        return True
    except ValueError:
        return False


def _normalize_brand(value):
    brand = (value or "").strip().lower()
    if not brand:
        return ""
    alias_map = {
        "hik": "hikvision",
        "hikvision": "hikvision",
        "dahua": "dahua",
        "prama": "prama",
        "cp plus": "cpplus",
        "cp-plus": "cpplus",
        "cpplus": "cpplus",
        "other": "other",
    }
    return alias_map.get(brand, brand)


def _validate_bulk_rows(rows):
    existing_ips = {c["ip"] for c in db.get_all_cameras()}
    normalized = []
    preview = []
    errors = []
    warnings = []
    per_row_messages = {}
    file_ip_rows = {}
    nvr_meta = {}

    def add_msg(row_num, level, text):
        per_row_messages.setdefault(row_num, []).append({"level": level, "text": text})
        (errors if level == "error" else warnings).append({"row": row_num, "text": text})

    for idx, raw in enumerate(rows, start=2):
        row_num = idx
        ip = (raw.get("ip") or "").strip()
        if not ip and not any((str(v or "").strip()) for v in raw.values()):
            continue
        name = (raw.get("name") or "").strip()
        zone = (raw.get("zone") or "").strip()
        nvr_name = (raw.get("nvr_name") or "").strip()
        nvr_ip = (raw.get("nvr_ip") or "").strip()
        brand = _normalize_brand(raw.get("brand"))
        if not ip:
            add_msg(row_num, "error", "Camera IP is required.")
        elif not _is_valid_ip(ip):
            add_msg(row_num, "error", f"Camera IP '{ip}' is not a valid IP address.")
        if ip:
            file_ip_rows.setdefault(ip, []).append(row_num)
        if brand and brand not in ALLOWED_CAMERA_BRANDS:
            allowed = ", ".join(ALLOWED_CAMERA_BRANDS)
            add_msg(row_num, "error", f"Brand '{raw.get('brand')}' is not allowed. Use one of: {allowed}.")
        if nvr_ip and not _is_valid_ip(nvr_ip):
            add_msg(row_num, "error", f"NVR IP '{nvr_ip}' is not a valid IP address.")
        if nvr_name and not nvr_ip:
            add_msg(row_num, "warning", "NVR Name is set but NVR IP is missing. This NVR will show as unconfigured.")
        if nvr_ip:
            meta = nvr_meta.setdefault(nvr_ip, {"names": set(), "brands": set(), "rows": []})
            if nvr_name:
                meta["names"].add(nvr_name)
            if brand:
                meta["brands"].add(brand)
            meta["rows"].append(row_num)
        normalized.append({
            "ip": ip,
            "name": name,
            "location": (raw.get("location") or "").strip(),
            "zone": zone,
            "nvr_name": nvr_name,
            "nvr_ip": nvr_ip,
            "nvr_channel": raw.get("nvr_channel", 1),
            "brand": brand,
            "username": (raw.get("username") or "admin").strip() or "admin",
            "password": raw.get("password", ""),
            "notes": raw.get("notes", ""),
            "rtsp_url": raw.get("rtsp_url", ""),
            "_row_num": row_num,
        })

    for ip, row_nums in file_ip_rows.items():
        if len(row_nums) > 1:
            for row_num in row_nums:
                add_msg(row_num, "error", f"Camera IP '{ip}' is duplicated in this file (rows {', '.join(map(str, row_nums))}).")

    for nvr_ip, meta in nvr_meta.items():
        if len(meta["names"]) > 1:
            names = ", ".join(sorted(meta["names"]))
            for row_num in meta["rows"]:
                add_msg(row_num, "warning", f"NVR IP '{nvr_ip}' has inconsistent NVR names in this file: {names}.")
        if len(meta["brands"]) > 1:
            brands = ", ".join(sorted(meta["brands"]))
            for row_num in meta["rows"]:
                add_msg(row_num, "warning", f"NVR IP '{nvr_ip}' has inconsistent brands in this file: {brands}.")

    for row in normalized:
        if not row["ip"]:
            continue
        action = "update" if row["ip"] in existing_ips else "add"
        row_messages = per_row_messages.get(row["_row_num"], [])
        preview.append({
            "row": row["_row_num"],
            "ip": row["ip"],
            "name": row["name"],
            "zone": row["zone"],
            "brand": row["brand"],
            "nvr_name": row["nvr_name"],
            "nvr_ip": row["nvr_ip"],
            "action": action,
            "messages": row_messages,
        })

    blocking = bool(errors)
    return {
        "rows": normalized,
        "preview": preview,
        "errors": errors,
        "warnings": warnings,
        "blocking": blocking,
        "summary": {
            "total": len(preview),
            "new": sum(1 for row in preview if row["action"] == "add"),
            "updates": sum(1 for row in preview if row["action"] == "update"),
            "errors": len(errors),
            "warnings": len(warnings),
        },
    }


def resolve_log_range(preset, from_str, to_str):
    today = datetime.now().date()
    if preset == "7d":
        start = today - timedelta(days=6)
        end = today
    elif preset == "1m":
        start = today - timedelta(days=29)
        end = today
    elif preset == "3m":
        start = today - timedelta(days=89)
        end = today
    elif preset == "custom" and from_str and to_str:
        try:
            start = datetime.strptime(from_str, "%Y-%m-%d").date()
            end = datetime.strptime(to_str, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            start = today - timedelta(days=6)
            end = today
        if end < start:
            start, end = end, start
    else:
        start = today - timedelta(days=6)
        end = today
    return start.isoformat(), end.isoformat()


@app.route("/reports")
@login_required
def reports_page():
    return render_template_string(REPORTS_HTML, user=actor_name(), site_name=SITE_NAME)


@app.route("/nvr-monitor")
@login_required
def nvr_monitor_page():
    return render_template_string(
        NVR_HTML,
        user=actor_name(),
        site_name=SITE_NAME,
        monitor_settings=get_notification_settings(),
        zones=db.get_zones(),
    )


@app.route("/api/nvrs")
@login_required
def api_nvrs():
    status = (request.args.get("status") or "").strip().lower()
    zone = (request.args.get("zone") or "").strip()
    search = (request.args.get("q") or "").strip().lower()
    items = db.get_nvr_endpoints()
    nvr_status_map = db.get_nvr_status_map()

    for item in items:
        nvr_ip = (item.get("nvr_ip") or "").strip()
        if not nvr_ip:
            item["status"] = "unconfigured"
            item["status_summary"] = "NVR IP not configured yet."
            item["last_seen"] = None
            item["offline_since"] = None
        else:
            cached = nvr_status_map.get(nvr_ip)
            if cached is None:
                # Not yet polled — fall back to unknown
                item["status"] = "unknown"
                item["status_summary"] = "NVR not yet polled. Waiting for next monitor cycle."
                item["last_seen"] = None
                item["offline_since"] = None
            elif cached["online"]:
                item["status"] = "online"
                item["status_summary"] = "NVR responds to ping."
                item["last_seen"] = cached.get("last_seen")
                item["offline_since"] = None
            else:
                item["status"] = "offline"
                item["status_summary"] = "NVR is not responding to ping."
                item["last_seen"] = cached.get("last_seen")
                item["offline_since"] = cached.get("offline_since")

    if status:
        items = [item for item in items if item.get("status") == status]
    if zone:
        items = [item for item in items if zone in (item.get("zones") or [])]
    if search:
        items = [
            item for item in items
            if search in (item.get("nvr_name") or "").lower()
            or search in (item.get("nvr_ip") or "").lower()
            or any(search in (z or "").lower() for z in item.get("zones") or [])
            or any(search in (loc or "").lower() for loc in item.get("locations") or [])
            or any(search in (cam.get("name") or "").lower() or search in (cam.get("ip") or "").lower() for cam in item.get("cameras") or [])
        ]
    summary = {
        "total": len(items),
        "online": sum(1 for item in items if item["status"] == "online"),
        "offline": sum(1 for item in items if item["status"] == "offline"),
        "unconfigured": sum(1 for item in items if item["status"] == "unconfigured"),
        "unknown": sum(1 for item in items if item["status"] == "unknown"),
        "mapped_cameras": sum(item.get("total_cameras", 0) for item in items),
    }
    return jsonify({"summary": summary, "items": items})


@app.route("/api/reports")
@login_required
def api_reports():
    preset = request.args.get("preset", "30d")
    from_str = request.args.get("from")
    to_str = request.args.get("to")
    date_from, date_to = resolve_report_range(preset, from_str, to_str)
    overview = db.get_report_overview(date_from, date_to)
    return jsonify({
        "preset": preset,
        "date_from": date_from,
        "date_to": date_to,
        "overview": overview,
        "status": db.get_stats(),
        "daily_trend": db.get_report_daily_trend(date_from, date_to),
        "worst_cameras": db.get_report_worst_cameras(date_from, date_to, limit=10),
        "zone_summary": db.get_report_zone_summary(date_from, date_to),
        "nvr_summary": db.get_report_nvr_summary(date_from, date_to),
    })

@app.route("/api/refresh", methods=["POST"])
@login_required
@role_required("operator")
def api_refresh():
    monitor.poll_all()
    db.add_audit(actor_name(), "monitor", "Manual dashboard refresh triggered", "System", request.remote_addr, "success")
    return jsonify({"success": True, "stats": db.get_stats()})

@app.route("/api/central-refresh", methods=["POST"])
def api_central_refresh():
    if not hmac.compare_digest(request.headers.get("X-API-Key", ""), CENTRAL_API_KEY):
        return jsonify({"error": "Unauthorized"}), 401
    monitor.poll_all()
    central_sync.push_summary()
    return jsonify({"success": True, "stats": db.get_stats(), "site": cfg.get("central", "site_name", fallback="Local Site")})


@app.route("/api/central-sync-users", methods=["POST"])
def api_central_sync_users():
    if not hmac.compare_digest(request.headers.get("X-API-Key", ""), CENTRAL_API_KEY):
        return jsonify({"error": "Unauthorized"}), 401
    ok = central_sync.sync_users()
    return jsonify({"success": ok, "site": cfg.get("central", "site_name", fallback="Local Site")}), (200 if ok else 502)


@app.route("/api/site-settings", methods=["GET", "POST"])
@login_required
@role_required("admin")
def api_site_settings():
    if request.method == "GET":
        return jsonify(get_site_settings())

    data = request.get_json(silent=True) or {}
    site_name = (data.get("site_name") or "").strip()
    site_id = (data.get("site_id") or "").strip()
    if not site_name:
        return jsonify({"error": "Site name is required"}), 400
    if not site_id:
        return jsonify({"error": "Site ID is required"}), 400

    settings = save_site_settings(data)
    db.add_audit(
        actor_name(),
        "config",
        f"Updated site settings for {settings['site_name']}",
        settings["site_id"],
        request.remote_addr,
        "success"
    )
    central_sync.push_summary()
    return jsonify({"success": True, "settings": settings})


@app.route("/api/site-settings/verify", methods=["POST"])
@login_required
@role_required("admin")
def api_verify_site_settings():
    """Save + verify: push a summary to central to confirm the API key works."""
    settings = get_site_settings()
    api_key = (settings.get("api_key") or "").strip()
    if not api_key or api_key == "local-dev-key":
        return jsonify({"error": "No API key configured. Paste the key from the central dashboard's Manage Sites page and save first."}), 400
    api_url = (settings.get("api_url") or "").strip()
    if not api_url:
        return jsonify({"error": "Central API URL is missing. Set it and save first."}), 400
    ok = central_sync.push_summary()
    if ok:
        db.add_audit(
            actor_name(), "config",
            f"Verified central connection for {settings['site_name']}",
            settings["site_id"], request.remote_addr, "success"
        )
        return jsonify({
            "success": True,
            "message": f"Connection verified! Site '{settings['site_name']}' is synced with the central dashboard. SSO will work.",
        })
    status = central_sync.get_status()
    err = status.get("last_error") or "Central sync failed"
    return jsonify({"error": f"Could not reach the central dashboard: {err}. Check the API URL and API key."}), 502


@app.route("/api/site-settings/register", methods=["POST"])
@login_required
@role_required("admin")
def api_register_site_settings():
    settings = get_site_settings()
    site_id = settings["site_id"].strip()
    site_name = settings["site_name"].strip()
    registration_url = get_registration_url(settings["api_url"])
    existing_api_key = (settings.get("api_key") or "").strip()
    # Treat placeholder/default key as no key — force registration flow
    if existing_api_key in ("local-dev-key", ""):
        existing_api_key = ""
    if not site_id:
        return jsonify({"error": "Site ID is required before registration"}), 400
    if not site_name:
        return jsonify({"error": "Site name is required before registration"}), 400
    if existing_api_key:
        if central_sync.push_summary():
            db.add_audit(
                actor_name(),
                "config",
                f"Verified central sync for {settings['site_name']}",
                settings["site_id"],
                request.remote_addr,
                "success"
            )
            return jsonify({
                "success": True,
                "settings": settings,
                "api_key": existing_api_key,
                "central_api_url": settings["api_url"],
                "message": "Site connection verified with central dashboard.",
            })
        return jsonify({
            "error": "This site already has an API key, but central sync failed. Check the Central API URL or generate a fresh key from the central dashboard."
        }), 502
    if not registration_url:
        return jsonify({"error": "Central API URL is missing or not valid"}), 400

    payload = json.dumps({
        "site_id": site_id,
        "site_name": site_name,
        "campus": settings["campus"],
    }).encode("utf-8")
    req = urllib.request.Request(
        registration_url,
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=15, context=SSL_CONTEXT) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        try:
            body = json.loads(exc.read().decode("utf-8"))
            msg = body.get("error") or f"Central registration failed with HTTP {exc.code}"
        except Exception:
            msg = f"Central registration failed with HTTP {exc.code}"
        return jsonify({"error": msg}), 502
    except Exception as exc:
        return jsonify({"error": f"Could not reach central server: {exc}"}), 502

    updated = dict(settings)
    updated["enabled"] = True
    updated["api_key"] = body.get("api_key") or settings["api_key"]
    updated["api_url"] = body.get("central_api_url") or settings["api_url"]
    updated["site_id"] = body.get("site_id") or site_id
    updated["site_name"] = body.get("site_name") or site_name
    updated["campus"] = body.get("campus") or settings["campus"]
    saved = save_site_settings(updated)
    db.add_audit(
        actor_name(),
        "config",
        f"Registered site {saved['site_name']} with central dashboard",
        saved["site_id"],
        request.remote_addr,
        "success"
    )
    central_sync.push_summary()
    return jsonify({
        "success": True,
        "settings": saved,
        "api_key": saved["api_key"],
        "central_api_url": saved["api_url"],
    })


@app.route("/api/notification-settings", methods=["GET", "POST"])
@login_required
@role_required("admin")
def api_notification_settings():
    if request.method == "GET":
        return jsonify(get_notification_settings())

    data = request.get_json(silent=True) or {}
    saved = save_notification_settings(data)
    db.add_audit(
        actor_name(),
        "config",
        "Updated notification settings",
        "Notifications",
        request.remote_addr,
        "success"
    )
    return jsonify({"success": True, "settings": saved})


@app.route("/api/users", methods=["GET", "POST"])
@login_required
@role_required("admin")
def api_users():
    if request.method == "GET":
        return jsonify([u for u in db.list_users() if u.get("source") == "local"])
    return jsonify({"error": "Create shared users from the central dashboard. This local site keeps only the fallback admin account."}), 400


@app.route("/api/users/<int:user_id>", methods=["PATCH"])
@login_required
@role_required("admin")
def api_update_user(user_id):
    return jsonify({"error": "Update shared users from the central dashboard. Local login changes are limited to the fallback admin account."}), 400

# ── Bulk Upload ───────────────────────────────────────────────────────────────
@app.route("/api/bulk/preview", methods=["POST"])
@login_required
@role_required("admin")
def api_bulk_preview():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    try:
        rows = _read_bulk_rows(f)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    result = _validate_bulk_rows(rows)
    return jsonify(result)

@app.route("/api/bulk/import", methods=["POST"])
@login_required
@role_required("admin")
def api_bulk_import():
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    f = request.files["file"]
    try:
        rows = _read_bulk_rows(f)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    validation = _validate_bulk_rows(rows)
    if validation["blocking"]:
        return jsonify({
            "error": "Import blocked. Fix the file errors shown in preview and try again.",
            "errors": validation["errors"],
            "warnings": validation["warnings"],
            "summary": validation["summary"],
        }), 400

    added = updated = 0
    for r in validation["rows"]:
        ip = r.get("ip","").strip()
        if not ip: continue
        existing = db.get_camera(ip)
        db.upsert_camera({
            "ip": ip, "name": r.get("name",""), "location": r.get("location",""),
            "zone": r.get("zone",""), "nvr_name": r.get("nvr_name",""), "nvr_ip": r.get("nvr_ip",""),
            "nvr_channel": r.get("nvr_channel", 1), "brand": _normalize_brand(r.get("brand","")),
            "username": r.get("username","admin"), "password": r.get("password",""),
            "notes": r.get("notes",""), "rtsp_url": r.get("rtsp_url","")
        })
        if existing: updated += 1
        else: added += 1

    db.add_audit(actor_name(), "bulk",
                 f"Import: {added} added, {updated} updated", f"{added+updated} cameras",
                 request.remote_addr, "success")
    return jsonify({"success": True, "added": added, "updated": updated, "warnings": validation["warnings"]})

@app.route("/api/camera", methods=["POST"])
@login_required
@role_required("admin")
def api_add_camera():
    data = request.get_json(silent=True) or {}
    ip = (data.get("ip") or "").strip()
    original_ip = (data.get("original_ip") or ip).strip()
    if not ip:
        return jsonify({"error": "IP address is required"}), 400
    if not _is_valid_ip(ip):
        return jsonify({"error": f"'{ip}' is not a valid IP address"}), 400
    nvr_ip = (data.get("nvr_ip") or "").strip()
    if nvr_ip and not _is_valid_ip(nvr_ip):
        return jsonify({"error": f"NVR IP '{nvr_ip}' is not a valid IP address"}), 400
    brand = (data.get("brand") or "").strip().lower()
    if brand and brand not in ALLOWED_CAMERA_BRANDS:
        return jsonify({"error": f"Brand '{brand}' is not allowed. Use one of: {', '.join(ALLOWED_CAMERA_BRANDS)}"}), 400
    try:
        nvr_channel = int(data.get("nvr_channel") or 1)
    except (ValueError, TypeError):
        nvr_channel = 1

    payload = {
        "ip": ip,
        "name": (data.get("name") or "").strip(),
        "location": (data.get("location") or "").strip(),
        "zone": (data.get("zone") or "").strip(),
        "nvr_name": (data.get("nvr_name") or "").strip(),
        "nvr_ip": nvr_ip,
        "nvr_channel": nvr_channel,
        "brand": brand,
        "username": ((data.get("username") or "admin").strip() or "admin"),
        "password": data.get("password") or "",
        "notes": (data.get("notes") or "").strip(),
        "rtsp_url": (data.get("rtsp_url") or "").strip(),
    }
    try:
        if original_ip and db.get_camera(original_ip):
            existing = db.get_camera(original_ip)
            db.update_camera(original_ip, payload)
        else:
            existing = db.get_camera(ip)
            db.upsert_camera(payload)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 409
    db.add_audit(
        actor_name(),
        "config",
        f"{'Updated' if existing else 'Added'} camera {original_ip if existing else ip}{f' -> {ip}' if existing and original_ip != ip else ''}",
        ip,
        request.remote_addr,
        "success"
    )
    return jsonify({"success": True, "mode": "updated" if existing else "added", "ip": ip})


@app.route("/api/camera/<ip>", methods=["DELETE"])
@login_required
@role_required("admin")
def api_delete_camera(ip):
    cam = db.get_camera(ip)
    if not cam:
        return jsonify({"error": "Camera not found"}), 404
    db.deactivate_camera(ip)
    db.add_audit(
        actor_name(),
        "config",
        f"Deleted camera {cam.get('name') or ip}",
        ip,
        request.remote_addr,
        "success"
    )
    return jsonify({"success": True, "ip": ip})

# ── Export ────────────────────────────────────────────────────────────────────
@app.route("/export/cameras/excel")
@login_required
def export_cameras_excel():
    cameras = db.get_all_cameras()
    out = exporter.export_cameras_excel(cameras)
    db.add_audit(actor_name(), "export", "Full camera list exported to Excel",
                 "System", request.remote_addr, "success")
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"cameras_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

@app.route("/export/cameras/csv")
@login_required
@role_required("admin")
def export_cameras_csv():
    import csv
    cameras = db.get_all_cameras()
    out = io.StringIO()
    fields = ["ip","name","location","zone","nvr_name","nvr_ip","nvr_channel","brand","username","password","notes","rtsp_url"]
    writer = csv.DictWriter(out, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(cameras)
    db.add_audit(actor_name(), "export", "Camera list (with credentials) exported to CSV",
                 "System", request.remote_addr, "success")
    return send_file(io.BytesIO(out.getvalue().encode()), mimetype="text/csv",
                     as_attachment=True, download_name=f"cameras_{datetime.now().strftime('%Y%m%d_%H%M')}.csv")

@app.route("/export/offline/excel")
@login_required
def export_offline_excel():
    cameras = db.get_offline_cameras()
    out = exporter.export_offline_excel(cameras)
    db.add_audit(actor_name(), "export", "Offline camera list exported to Excel",
                 "System", request.remote_addr, "success")
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"offline_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

@app.route("/export/offline/pdf")
@login_required
def export_offline_pdf():
    cameras = db.get_offline_cameras()
    out = exporter.export_offline_pdf(cameras)
    db.add_audit(actor_name(), "export", "Offline camera list exported to PDF",
                 "System", request.remote_addr, "success")
    return send_file(out, mimetype="application/pdf",
                     as_attachment=True, download_name=f"offline_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf")

@app.route("/export/template")
@login_required
def export_template():
    out = exporter.export_template_excel()
    return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="camera_import_template.xlsx")

# ── HTML Templates ────────────────────────────────────────────────────────────
LOGIN_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>CCTV Monitor — Kumarans</title>
<style>*{box-sizing:border-box;margin:0;padding:0}body{font-family:system-ui,sans-serif;background:linear-gradient(180deg,#eef4f8 0%,#dfe8ee 100%);display:flex;align-items:center;justify-content:center;min-height:100vh;color:#1f2937;padding:24px}.card{background:#fff;border-radius:22px;padding:30px;width:460px;max-width:100%;box-shadow:0 24px 60px rgba(15,23,42,.12);border:1px solid #d8e3ea}.brand{display:grid;justify-items:center;text-align:center;margin-bottom:22px}.brand img{width:92px;height:auto;display:block}.eyebrow{font-size:11px;font-weight:700;letter-spacing:.14em;text-transform:uppercase;color:#0f766e;margin-top:14px}.logo{font-size:24px;font-weight:800;line-height:1.2;margin-top:10px;color:#0f172a}.sub{color:#64748b;font-size:13px;margin-top:10px;margin-bottom:22px}label{display:block;font-size:12px;font-weight:600;color:#475569;margin-bottom:6px}input{width:100%;padding:10px 12px;border:1px solid #cbd5e1;border-radius:10px;font-size:14px;margin-bottom:16px;outline:none}input:focus{border-color:#3498db}button{width:100%;padding:11px;background:#1f4ed8;color:#fff;border:none;border-radius:10px;font-size:14px;cursor:pointer;font-weight:700}button:hover{background:#1d4ed8}.err{color:#b91c1c;font-size:12px;margin-bottom:14px;background:#fff1f2;border:1px solid #fecdd3;padding:10px 12px;border-radius:10px}.warn{font-size:12px;color:#7c2d12;background:#fff7ed;border:1px solid #fed7aa;padding:12px 14px;border-radius:12px;margin-top:18px;line-height:1.5}</style></head>
<body><div class="card"><div class="brand"><img src="https://kumarans.org/images/loader_logo.png?v=ats-cms.1.0" alt="Sri Kumarans logo"><div class="eyebrow">Authorized Access</div><div class="logo">Sri Kumarans Childrens Home Educational Council</div><p class="sub">Local camera monitoring dashboard access for approved operations and administrative users only.</p></div>
{% if error %}<div class="err">{{ error }}</div>{% endif %}
<form method="POST"><input type="hidden" name="_csrf" value="{{ csrf_token }}"><label>Username</label><input name="username" type="text" autofocus><label>Password</label><input name="password" type="password"><button type="submit">Sign In</button></form><div class="warn">Warning: This system is restricted to authorized personnel of Sri Kumarans Childrens Home Educational Council. Unauthorized access, use, or distribution of information from this dashboard is prohibited and may be monitored and investigated.</div></div></body></html>"""

DASHBOARD_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>CCTV Monitor — Kumarans</title>
<style>
:root{--bg:#f3f6fb;--surface:#ffffff;--surface-soft:#f8fbff;--line:#dbe3ef;--line-strong:#c8d4e5;--text:#1f2a37;--muted:#6b7a90;--primary:#1f6feb;--primary-soft:#e8f1ff;--ok:#16a34a;--shadow:0 10px 30px rgba(15,23,42,.06)}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:system-ui,sans-serif;background:linear-gradient(180deg,#f8fafc 0%,var(--bg) 100%);color:var(--text);font-size:13px}
.topbar{padding:14px 20px;background:rgba(255,255,255,.96);border-bottom:1px solid var(--line);position:sticky;top:0;z-index:100;backdrop-filter:blur(12px)}
.top-row{display:grid;grid-template-columns:1fr auto 1fr;align-items:center;gap:14px}
.brand{display:flex;align-items:center;gap:12px;min-width:0}
.brand-mark{display:flex;align-items:center;justify-content:center;flex:0 0 auto}
.brand-mark img{width:342px;max-width:100%;height:auto;object-fit:contain;filter:drop-shadow(0 2px 6px rgba(15,23,42,.10))}
.brand-copy{display:flex;flex-direction:column;gap:2px;min-width:0}
.school{font-size:16px;font-weight:700;color:var(--text);line-height:1.2}
.site-pill{font-size:11px;font-weight:700;color:#2159b3;background:var(--primary-soft);border:1px solid #cfe0ff;border-radius:999px;padding:5px 12px;line-height:1}
.titlebar{justify-self:center;display:flex;flex-direction:column;align-items:center;gap:8px;font-size:18px;font-weight:700;color:var(--text);letter-spacing:-.02em;text-align:center}
.title-main{line-height:1}
.title-sub{display:flex;align-items:center;justify-content:center}
@keyframes livePulse{0%{box-shadow:0 0 0 0 rgba(22,163,74,.45)}70%{box-shadow:0 0 0 9px rgba(22,163,74,0)}100%{box-shadow:0 0 0 0 rgba(22,163,74,0)}}
.top-actions{display:flex;justify-content:flex-end}
.subbar{display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-top:12px;padding-top:12px;border-top:1px solid #edf2f7}
.status-rail{display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-left:auto}
.status-chip{font-size:11px;color:#475569;display:flex;align-items:center;gap:6px;font-weight:700;padding:6px 10px;border-radius:999px;background:#fff;border:1px solid var(--line)}
.status-chip .dot{width:8px;height:8px;border-radius:50%;display:inline-block;background:#94a3b8}
.status-chip.ok{color:#166534;background:#effcf3;border-color:#ccefd7}
.status-chip.ok .dot{background:var(--ok);box-shadow:0 0 0 0 rgba(22,163,74,.45);animation:livePulse 1.8s infinite}
.status-chip.warn{color:#8a3b12;background:#fff4e8;border-color:#f8d8b0}
.status-chip.warn .dot{background:#f59f0b}
.status-chip.info{color:#2159b3;background:var(--primary-soft);border-color:#cfe0ff}
.status-chip.info .dot{background:var(--primary)}
.search{flex:0 0 230px;padding:9px 12px;border:1px solid var(--line);border-radius:10px;font-size:12px;outline:none;background:#fff;color:var(--text)}
.search:focus{border-color:var(--primary);box-shadow:0 0 0 3px rgba(31,111,235,.10)}
.nav-link{padding:8px 12px;border-radius:10px;font-size:12px;cursor:pointer;border:1px solid var(--line);background:#fff;color:#4b5563;text-decoration:none;font-weight:600}
.nav-link:hover{background:#f8fbff;border-color:var(--line-strong)}
.stats-row{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;padding:14px 20px}
.scard{background:var(--surface);border-radius:14px;padding:14px 16px;border:1px solid var(--line);box-shadow:var(--shadow)}
.slbl{font-size:11px;color:var(--muted);margin-bottom:4px}
.sval{font-size:26px;font-weight:700}
.body{display:flex;gap:0;min-height:calc(100vh - 140px)}
.sidebar{width:170px;flex-shrink:0;padding:12px 8px;background:var(--surface);border-right:1px solid var(--line)}
.sg{margin-bottom:16px}
.st{font-size:10px;font-weight:700;color:#94a3b8;letter-spacing:0.08em;padding:0 8px;margin-bottom:4px;display:block}
.sb{display:block;width:100%;text-align:left;padding:7px 9px;border-radius:9px;border:none;background:none;cursor:pointer;color:#64748b;font-size:12px}
.sb:hover,.sb.active{background:var(--primary-soft);color:#174ea6;font-weight:600}
.main{flex:1;display:flex;flex-direction:column}
.fbar{display:flex;gap:8px;align-items:center;padding:12px 16px;background:var(--surface);border-bottom:1px solid var(--line);flex-wrap:wrap}
.chip{padding:6px 12px;border-radius:20px;border:1px solid var(--line);font-size:11px;cursor:pointer;background:#fff;color:#64748b;font-weight:600}
.chip.active{background:var(--primary);color:#fff;border-color:var(--primary)}
.gwrap{flex:1;padding:12px 16px;overflow-y:auto}
.table-wrap{background:var(--surface);border:1px solid var(--line);border-radius:16px;overflow:auto;box-shadow:var(--shadow)}
.cam-table{width:100%;border-collapse:separate;border-spacing:0;background:#fff;font-size:12px}
.cam-table thead th{position:sticky;top:0;background:#f8fafc;z-index:2;white-space:nowrap;font-size:11px;font-weight:700;color:#5d6d83;text-transform:uppercase;letter-spacing:.04em}
.cam-table tbody tr{cursor:pointer}
.cam-table tbody tr.offline{background:#fff5f5}
.cam-table tbody tr.maintenance{background:#fff9f0}
.cam-table tbody tr.selected{background:#ebf5fb}
.cam-table tbody tr:hover{background:#f8fbff}
.cam-table td,.cam-table th{padding:10px 12px;border-bottom:1px solid #eef2f7;text-align:left;vertical-align:middle}
.empty-state{padding:24px;text-align:center;color:#888;font-size:12px}
.dot{width:8px;height:8px;border-radius:50%}
.hb{font-size:10px;font-weight:700;padding:3px 9px;border-radius:999px;display:inline-block;letter-spacing:.02em}
.tn{font-size:12px;font-weight:600}
.tip{font-size:10px;color:var(--muted);font-family:ui-monospace,SFMono-Regular,monospace}
.bb{font-size:10px;font-weight:600;padding:3px 9px;border-radius:999px;display:inline-block}
.status-pill{display:inline-flex;align-items:center;gap:6px;padding:3px 8px;border-radius:999px;font-size:10px;font-weight:600;text-transform:capitalize}
.status-pill.online{background:#eafaf1;color:#1e8449}
.status-pill.offline{background:#fdecea;color:#922b21}
.status-pill.maintenance{background:#fff3e0;color:#9c640c}
.row-check{width:14px;height:14px}
.bbar{display:flex;align-items:center;gap:6px;padding:10px 16px;background:var(--surface);border-top:1px solid var(--line);flex-wrap:wrap}
.btn{padding:8px 12px;border-radius:10px;border:1px solid var(--line);background:#fff;cursor:pointer;font-size:11px;color:#475569;font-weight:600}
.btn:hover{background:#f8fbff;border-color:var(--line-strong)}
.btn.danger{border-color:#f5c6c6;color:#c0392b}
.footer-note{padding:12px 18px 18px;color:#7a879a;font-size:11px;line-height:1.6;text-align:center}
.sel-bar{display:none;padding:4px 8px;background:#ebf5fb;border:1px solid #aed6f1;border-radius:12px;align-items:center;gap:8px;font-size:12px;flex-wrap:wrap}
.sel-bar.show{display:flex}
/* Modal */
.overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,0.5);z-index:200;align-items:flex-start;justify-content:center;padding:40px 12px;overflow-y:auto}
.overlay.show{display:flex}
.modal{background:#fff;border-radius:14px;width:500px;max-width:100%}
.mhdr{display:flex;align-items:flex-start;justify-content:space-between;padding:16px 20px;border-bottom:1px solid #eee}
.mbody{padding:16px 20px}
.pvbox{width:100%;height:auto;min-height:0;background:transparent;border:none;display:block;margin-bottom:16px;overflow:visible;position:relative}
.pvbox img{display:block;width:100%;height:auto;object-fit:contain}
.irow{display:grid;grid-template-columns:118px 1fr;font-size:12px;padding:9px 0;border-bottom:1px solid #f1f5f9;align-items:baseline;gap:10px}
.ilbl{color:var(--muted);font-weight:600;font-size:11px;text-transform:uppercase;letter-spacing:.04em}
.tog{width:36px;height:20px;border-radius:10px;background:#ddd;position:relative;cursor:pointer;transition:background 0.2s}
.tog.on{background:#e67e22}
.tog:after{content:'';position:absolute;width:16px;height:16px;border-radius:50%;background:white;top:2px;left:2px;transition:left 0.15s}
.tog.on:after{left:18px}
/* Audit */
.audit-wrap{padding:20px}
table{width:100%;border-collapse:collapse;background:#fff;border-radius:12px;overflow:hidden;border:1px solid var(--line);font-size:12px}
thead th{padding:10px 12px;text-align:left;background:#f8fafc;font-weight:700;color:#5d6d83;border-bottom:1px solid #edf2f7;font-size:11px;letter-spacing:.04em;text-transform:uppercase;white-space:nowrap}
tbody tr{border-bottom:1px solid #f1f5f9}
tbody tr:hover{background:#f8fbff}
td{padding:9px 12px}
.badge{display:inline-block;padding:3px 9px;border-radius:999px;font-size:10px;font-weight:600}
/* Stat value colors */
.sval-on{color:#16a34a}.sval-off{color:#dc2626}.sval-mt{color:#d97706}
/* Local user table */
.user-select{appearance:none;-webkit-appearance:none;border:1px solid transparent;background:transparent;font-size:12px;font-weight:600;padding:5px 22px 5px 10px;border-radius:8px;cursor:pointer;transition:border-color .15s,background .15s;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6' viewBox='0 0 10 6'%3E%3Cpath d='M1 1l4 4 4-4' stroke='%236b7a90' stroke-width='1.5' fill='none' stroke-linecap='round'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 7px center}
.user-select:hover{border-color:var(--line);background-color:#f8fafc;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6' viewBox='0 0 10 6'%3E%3Cpath d='M1 1l4 4 4-4' stroke='%236b7a90' stroke-width='1.5' fill='none' stroke-linecap='round'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 7px center}
.user-select:focus{border-color:var(--primary);background-color:#fff;outline:none;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6' viewBox='0 0 10 6'%3E%3Cpath d='M1 1l4 4 4-4' stroke='%236b7a90' stroke-width='1.5' fill='none' stroke-linecap='round'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 7px center}
.role-admin{color:#7c3aed}.role-operator{color:#0369a1}.role-viewer{color:#475569}
.status-active{color:#16a34a}.status-disabled{color:#b91c1c}
.uname-badge{font-size:11px;font-family:ui-monospace,monospace;color:var(--muted);background:#f1f5f9;padding:3px 9px;border-radius:999px;white-space:nowrap;display:inline-block}
.pwd-cell{display:flex;align-items:center;gap:6px}
.pwd-input{border:1px solid var(--line);background:#f8fafc;font-size:12px;padding:5px 8px;border-radius:8px;width:110px;transition:border-color .15s}
.pwd-input:focus{border-color:var(--primary);background:#fff;outline:none;box-shadow:0 0 0 3px rgba(31,111,235,.08)}
.btn-set{font-size:11px;font-weight:600;padding:5px 10px;border-radius:8px;border:1px solid var(--line);background:#fff;cursor:pointer;color:#475569;white-space:nowrap}
.btn-set:hover{background:#f0f7ff;border-color:var(--primary);color:var(--primary)}
/* Upload Modal */
.upload-zone{border:2px dashed #ddd;border-radius:10px;padding:30px;text-align:center;cursor:pointer;margin-bottom:14px}
.upload-zone:hover{border-color:#3498db}
.diff-table{width:100%;font-size:11px;border-collapse:collapse;margin-top:10px}
.diff-table th,.diff-table td{padding:5px 8px;border:1px solid #eee;text-align:left}
.diff-table th{background:#f5f5f5}
.form-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px}
.form-field{display:flex;flex-direction:column;gap:5px}
.form-field.full{grid-column:1 / -1}
.form-field label{font-size:11px;color:#666;font-weight:600}
.form-field input,.form-field select,.form-field textarea{width:100%;padding:9px 10px;border:1px solid #d9e0e8;border-radius:10px;font-size:12px;background:#fff;outline:none}
.form-field input:focus,.form-field select:focus,.form-field textarea:focus{border-color:var(--primary);box-shadow:0 0 0 3px rgba(31,111,235,.10)}
.form-field textarea{min-height:76px;resize:vertical}
@media (max-width: 980px){.top-row{grid-template-columns:1fr;justify-items:start}.titlebar{justify-self:start;align-items:flex-start}.title-sub{justify-content:flex-start}.top-actions{justify-content:flex-start}.status-rail{margin-left:0}}
@media (max-width: 860px){.stats-row{grid-template-columns:repeat(2,1fr)}.body{flex-direction:column}.sidebar{width:100%;border-right:none;border-bottom:1px solid var(--line)}.form-grid{grid-template-columns:1fr}}
</style></head>
<body>

<div class="topbar">
  <div class="top-row">
    <div class="brand">
      <div class="brand-mark"><img src="https://kumarans.org/images/Sri%20Kumaran%20Childrens%20Home.png" alt="Sri Kumaran logo"></div>
      <div class="brand-copy">
      </div>
    </div>
    <div class="titlebar">
      <span class="title-main">CCTV Monitoring Dashboard</span>
      <div class="title-sub"><strong class="site-pill" id="siteNameLabel">{{ site_name }}</strong></div>
    </div>
    <div class="top-actions">
      <a href="/logout" class="nav-link">Sign out ({{ user }})</a>
    </div>
  </div>
  <div class="subbar">
    <input class="search" type="search" placeholder="Search name or IP..." id="si" oninput="fc()">
    {% if role in ['operator','admin'] %}
    <button class="btn" id="refreshBtn" onclick="manualRefresh()">Refresh Now</button>
    {% endif %}
    {% if role == 'admin' %}
    <button class="btn" onclick="openSiteSettingsModal()">Edit Site Details</button>
    <button class="btn" onclick="openNotificationSettingsModal()">Notification Settings</button>
    {% endif %}
    <a href="/nvr-monitor" class="nav-link">NVR Monitor</a>
    <a href="/reports" class="nav-link">Reports</a>
    <a href="/camera-logs" class="nav-link">Camera Logs</a>
    <a href="/audit" class="nav-link">Audit Log</a>
    <div class="status-rail">
      <span class="status-chip ok" id="localInd"><span class="dot"></span><span id="localTxt">Local Server: Live</span></span>
      <span class="status-chip info" id="tunnelInd" title="Public site access status"><span class="dot"></span><span id="tunnelTxt">Tunnel: Checking</span></span>
      <span class="status-chip info" id="consoleInd" title="Central reporting status"><span class="dot"></span><span id="consoleTxt">Console: Checking</span></span>
    </div>
  </div>
</div>

<div class="stats-row">
  <div class="scard"><div class="slbl">Total Cameras</div><div class="sval" id="sT">—</div></div>
  <div class="scard"><div class="slbl">Online</div><div class="sval sval-on" id="sOn">—</div></div>
  <div class="scard"><div class="slbl">Offline</div><div class="sval sval-off" id="sOff">—</div></div>
  <div class="scard"><div class="slbl">Maintenance</div><div class="sval sval-mt" id="sMt">—</div></div>
</div>

<div class="body">
  <div class="sidebar">
    <div class="sg">
      <span class="st">ZONE</span>
      <button class="sb active" onclick="sf('zone','',this)">All Zones</button>
      {% for z in zones %}
      <button class="sb" onclick="sf('zone','{{ z }}',this)">{{ z }}</button>
      {% endfor %}
    </div>
    <div class="sg">
      <span class="st">LOCATION</span>
      <button class="sb active" onclick="sf('location','',this)">All Locations</button>
      {% for loc in locations %}
      <button class="sb" onclick="sf('location','{{ loc }}',this)">{{ loc }}</button>
      {% endfor %}
    </div>
    <div class="sg">
      <span class="st">NVR</span>
      <button class="sb active" onclick="sf('nvr','',this)">All NVRs</button>
      {% for n in nvrs %}
      <button class="sb" onclick="sf('nvr','{{ n }}',this)">{{ n }}</button>
      {% endfor %}
    </div>
    <div class="sg">
      <span class="st">BRAND</span>
      <button class="sb active" onclick="sf('brand','',this)">All Brands</button>
      <button class="sb" onclick="sf('brand','hikvision',this)">Hikvision</button>
      <button class="sb" onclick="sf('brand','dahua',this)">Dahua</button>
      <button class="sb" onclick="sf('brand','prama',this)">Prama</button>
    </div>
  </div>

  <div class="main">
    <div class="fbar">
      <span style="font-size:11px;color:#888">Status:</span>
      <button class="chip active" onclick="sf('status','',this)">All</button>
      <button class="chip" onclick="sf('status','online',this)">Online</button>
      <button class="chip" onclick="sf('status','offline',this)">Offline</button>
      <button class="chip" onclick="sf('status','maintenance',this)">Maintenance</button>
      <select class="search" id="zoneFilter" style="flex:0 0 140px" onchange="setQuickFilter('zone', this.value)">
        <option value="">All Zones</option>
        {% for z in zones %}<option value="{{ z }}">{{ z }}</option>{% endfor %}
      </select>
      <select class="search" id="locationFilter" style="flex:0 0 170px" onchange="setQuickFilter('location', this.value)">
        <option value="">All Locations</option>
        {% for loc in locations %}<option value="{{ loc }}">{{ loc }}</option>{% endfor %}
      </select>
      <select class="search" id="nvrFilter" style="flex:0 0 140px" onchange="setQuickFilter('nvr', this.value)">
        <option value="">All NVRs</option>
        {% for n in nvrs %}<option value="{{ n }}">{{ n }}</option>{% endfor %}
      </select>
      <select class="search" id="brandFilter" style="flex:0 0 130px" onchange="setQuickFilter('brand', this.value)">
        <option value="">All Brands</option>
        <option value="hikvision">Hikvision</option>
        <option value="dahua">Dahua</option>
        <option value="prama">Prama</option>
      </select>
      <select class="search" id="cameraPageSize" style="flex:0 0 118px" onchange="updateCameraPageSize()">
        <option value="50">50 / page</option>
        <option value="100" selected>100 / page</option>
        <option value="200">200 / page</option>
        <option value="300">300 / page</option>
        <option value="400">400 / page</option>
      </select>
      <div style="flex:1"></div>
      <button class="btn" onclick="clearFilters()">Clear Filters</button>
      <span style="font-size:11px;color:#aaa" id="ccnt"></span>
      <div class="sel-bar" id="selBar">
        <span id="selCount"></span>
        {% if role in ['operator','admin'] %}
        <button class="btn" onclick="bulkMaint(true)">Set Maintenance ON</button>
        <button class="btn" onclick="bulkMaint(false)">Set Maintenance OFF</button>
        {% endif %}
        {% if role == 'admin' %}
        <button class="btn" onclick="showBulkZone()">Assign Zone</button>
        <button class="btn" onclick="showBulkNvr()">Assign NVR</button>
        <button class="btn danger" onclick="deleteSelectedCameras()">Delete Selected</button>
        {% endif %}
        <button class="btn" onclick="clearSel()">Clear Selection</button>
      </div>
    </div>

    <div class="gwrap">
      <div class="table-wrap" id="cg"></div>
    </div>
  </div>
</div>

<div class="bbar">
  <span style="font-size:11px;color:#aaa" id="lastPoll">Last poll: —</span>
  <span style="font-size:11px;color:#70839b" id="cameraPageInfo">Page 1 of 1 • 0 shown • 0 matched</span>
  <div style="flex:1"></div>
  <button class="btn" id="cameraPrevBtn" onclick="changeCameraPage(-1)">Previous</button>
  <div id="cameraPageNums" style="display:flex;align-items:center;gap:6px;flex-wrap:wrap"></div>
  <button class="btn" id="cameraNextBtn" onclick="changeCameraPage(1)">Next</button>
  {% if role == 'admin' %}
  <button class="btn" onclick="openAddCameraModal()">Add Camera</button>
  <button class="btn" onclick="document.getElementById('uploadModal').classList.add('show')">Bulk Upload</button>
  {% endif %}
  <button class="btn" onclick="window.location='/export/cameras/excel'">Export Excel</button>
  <button class="btn" onclick="window.location='/export/cameras/csv'">Export CSV</button>
  <button class="btn" onclick="window.location='/export/offline/excel'">Offline Excel</button>
  <button class="btn" onclick="window.location='/export/offline/pdf'">Offline PDF</button>
  {% if role == 'admin' %}
  <button class="btn" onclick="window.location='/export/template'">Import Template</button>
  {% endif %}
</div>

<div class="footer-note">
  © Sri Kumaran Childrens Home Educational Council. All rights reserved. Authorized operational use only. Activity on this monitoring system may be logged and reviewed.
</div>

<!-- Camera Detail Modal -->
<div class="overlay" id="ov" onclick="if(event.target===this)co()">
  <div class="modal">
    <div class="mhdr">
      <div>
        <div style="font-size:16px;font-weight:600;display:flex;align-items:center;gap:8px"><span id="mNmDot" style="width:10px;height:10px;border-radius:50%;display:inline-block;flex:0 0 10px;background:#bbb"></span><span id="mNm"></span></div>
        <div style="font-size:12px;color:var(--muted);margin-top:3px" id="mSb"></div>
      </div>
      <div style="display:flex;gap:6px">
        {% if role == 'admin' %}
        <button class="btn" onclick="openEditCameraModal()">Edit Camera</button>
        <button class="btn danger" onclick="deleteSelectedCamera()">Delete Camera</button>
        {% endif %}
        <button class="btn" onclick="co()">Close</button>
      </div>
    </div>
    <div class="mbody">
      <div class="pvbox" id="pvb"><div style="text-align:center;color:#aaa">Loading preview...</div></div>
      <div id="mRws"></div>
      {% if role in ['operator','admin'] %}
      <div style="display:flex;align-items:center;justify-content:space-between;margin:14px 0 10px;padding-top:10px;border-top:1px solid #f5f5f5">
        <div>
          <div style="font-size:13px;font-weight:600">Maintenance Mode</div>
          <div style="font-size:11px;color:var(--muted);margin-top:2px">Suppresses all alerts for this camera</div>
        </div>
        <div class="tog" id="mtog" onclick="tm()"></div>
      </div>
      {% endif %}
      <div style="border-top:1px solid #f5f5f5;padding-top:12px">
        <div style="font-size:11px;font-weight:700;color:var(--muted);letter-spacing:.06em;margin-bottom:8px">RECENT EVENTS</div>
        <div id="mEv"></div>
      </div>
    </div>
  </div>
</div>

<!-- Upload Modal -->
<div class="overlay" id="uploadModal" onclick="if(event.target===this)this.classList.remove('show')">
  <div class="modal">
    <div class="mhdr">
      <div style="font-size:16px;font-weight:600">Bulk Camera Upload</div>
      <button class="btn" onclick="document.getElementById('uploadModal').classList.remove('show')">Close</button>
    </div>
    <div class="mbody">
      <p style="font-size:12px;color:#666;margin-bottom:12px">Upload a CSV or Excel file to add or update cameras. <a href="/export/template" style="color:#3498db">Download template</a></p>
      <div class="upload-zone" onclick="document.getElementById('fileInput').click()">
        <div style="font-size:24px;margin-bottom:8px">📁</div>
        <div style="font-size:13px;font-weight:500">Click to choose file</div>
        <div style="font-size:11px;color:#aaa;margin-top:4px">.csv or .xlsx supported</div>
      </div>
      <input type="file" id="fileInput" accept=".csv,.xlsx" style="display:none" onchange="previewUpload(this)">
      <div id="diffArea"></div>
      <div id="uploadActions" style="display:none;margin-top:12px;display:none">
        <button class="btn" style="background:#27ae60;color:#fff;border-color:#27ae60" onclick="confirmUpload()">Confirm Import</button>
      </div>
    </div>
  </div>
</div>

<div class="overlay" id="addCameraModal" onclick="if(event.target===this)closeAddCameraModal()">
  <div class="modal" style="width:620px">
    <div class="mhdr">
      <div>
        <div style="font-size:16px;font-weight:600" id="cameraFormTitle">Add Camera Manually</div>
        <div style="font-size:12px;color:#888;margin-top:3px" id="cameraFormSubtitle">Create or update a single camera without uploading a file.</div>
      </div>
      <button class="btn" onclick="closeAddCameraModal()">Close</button>
    </div>
    <div class="mbody">
      <form id="addCameraForm" onsubmit="submitCameraForm(event)">
        <input id="camOriginalIp" name="original_ip" type="hidden">
        <div class="form-grid">
          <div class="form-field">
            <label for="camIp">IP Address</label>
            <input id="camIp" name="ip" required placeholder="192.168.1.10">
          </div>
          <div class="form-field">
            <label for="camName">Camera Name</label>
            <input id="camName" name="name" required placeholder="Gate Camera 01">
          </div>
          <div class="form-field">
            <label for="camLocation">Location</label>
            <input id="camLocation" name="location" placeholder="Main gate">
          </div>
          <div class="form-field">
            <label for="camZone">Zone</label>
            <input id="camZone" name="zone" placeholder="Zone A">
          </div>
          <div class="form-field">
            <label for="camNvr">NVR Name</label>
            <input id="camNvr" name="nvr_name" placeholder="NVR-01">
          </div>
          <div class="form-field">
            <label for="camNvrIp">NVR IP</label>
            <input id="camNvrIp" name="nvr_ip" placeholder="192.168.1.200">
          </div>
          <div class="form-field">
            <label for="camChannel">NVR Channel</label>
            <input id="camChannel" name="nvr_channel" type="number" min="1" value="1">
          </div>
          <div class="form-field">
            <label for="camBrand">Brand</label>
            <select id="camBrand" name="brand">
              <option value="">Select brand</option>
              <option value="hikvision">Hikvision</option>
              <option value="dahua">Dahua</option>
              <option value="prama">Prama</option>
            </select>
          </div>
          <div class="form-field">
            <label for="camUser">Username</label>
            <input id="camUser" name="username" value="admin" placeholder="admin">
          </div>
          <div class="form-field">
            <label for="camPassword">Password</label>
            <input id="camPassword" name="password" type="password" placeholder="Camera password">
          </div>
          <div class="form-field full">
            <label for="camRtsp">RTSP URL</label>
            <input id="camRtsp" name="rtsp_url" placeholder="Optional custom RTSP URL">
          </div>
          <div class="form-field full">
            <label for="camNotes">Notes</label>
            <textarea id="camNotes" name="notes" placeholder="Optional notes"></textarea>
          </div>
        </div>
        <div style="display:flex;justify-content:space-between;align-items:center;margin-top:14px;gap:10px">
          <span id="addCameraMsg" style="font-size:12px;color:#888"></span>
          <div style="display:flex;gap:8px">
            <button type="button" class="btn" onclick="closeAddCameraModal()">Cancel</button>
            <button type="submit" class="btn" style="background:#2c3e50;color:#fff;border-color:#2c3e50">Save Camera</button>
          </div>
        </div>
      </form>
    </div>
  </div>
</div>

<div class="overlay" id="siteSettingsModal" onclick="if(event.target===this)closeSiteSettingsModal()">
  <div class="modal" style="width:720px">
    <div class="mhdr">
      <div>
        <div style="font-size:16px;font-weight:600">Site Details</div>
        <div style="font-size:12px;color:#888;margin-top:3px">Update the site identity and central dashboard connection from here.</div>
      </div>
      <button class="btn" onclick="closeSiteSettingsModal()">Close</button>
    </div>
    <div class="mbody">
      <form id="siteSettingsForm" onsubmit="saveSiteSettings(event)">
        <div class="form-grid">
          <div class="form-field full" style="flex-direction:row;align-items:center;gap:10px">
            <input id="siteEnabled" name="enabled" type="checkbox" style="width:auto">
            <label for="siteEnabled" style="margin:0">Enable central sync for this site</label>
          </div>
          <div class="form-field">
            <label for="siteId">Site ID</label>
            <input id="siteId" name="site_id" required placeholder="blr-campus-1">
          </div>
          <div class="form-field">
            <label for="siteName">Site Name</label>
            <input id="siteName" name="site_name" required placeholder="Bangalore Campus">
          </div>
          <div class="form-field">
            <label for="siteCampus">Campus</label>
            <input id="siteCampus" name="campus" placeholder="South Campus">
          </div>
          <div class="form-field">
            <label for="siteAddress">Site Address</label>
            <input id="siteAddress" name="site_address" placeholder="Admin Block, Bengaluru">
          </div>
          <div class="form-field">
            <label for="siteContactName">Point of Contact</label>
            <input id="siteContactName" name="contact_name" placeholder="Security Control Room">
          </div>
          <div class="form-field">
            <label for="siteContactPhone">Contact Phone</label>
            <input id="siteContactPhone" name="contact_phone" placeholder="+91 98765 43210">
          </div>
          <div class="form-field full">
            <label for="siteContactEmail">Contact Email</label>
            <input id="siteContactEmail" name="contact_email" placeholder="controlroom@example.com">
          </div>
          <div class="form-field full">
            <label for="siteDashboardUrl">Dashboard URL</label>
            <input id="siteDashboardUrl" name="dashboard_url" placeholder="https://blr-campus.example.com">
          </div>
          <div class="form-field full">
            <label for="siteRefreshUrl">Refresh URL</label>
            <input id="siteRefreshUrl" name="refresh_url" placeholder="https://blr-campus.example.com/api/central-refresh">
          </div>
          <div class="form-field full">
            <label for="siteApiUrl">Central API URL</label>
            <input id="siteApiUrl" name="api_url" placeholder="https://monitor.example.com/api/site-summary">
          </div>
          <div class="form-field full">
            <label for="siteApiKey">Central API Key</label>
            <input id="siteApiKey" name="api_key" placeholder="Paste the API key from the central dashboard">
            <small style="color:#64748b;margin-top:4px;display:block">Copy the API key from the central dashboard's Manage Sites page and paste it here, then click Save &amp; Verify.</small>
          </div>
        </div>
        <div style="display:flex;justify-content:space-between;align-items:center;margin-top:14px;gap:10px">
          <span id="siteSettingsMsg" style="font-size:12px;color:#888"></span>
          <div style="display:flex;gap:8px;flex-wrap:wrap">
            <button type="button" class="btn" id="registerSiteBtn" onclick="registerThisSite()">Register This Site</button>
            <button type="button" class="btn" id="verifySiteBtn" onclick="saveAndVerifySite()" style="background:#0f766e;color:#fff;border-color:#0f766e">Save &amp; Verify</button>
            <button type="button" class="btn" onclick="closeSiteSettingsModal()">Cancel</button>
            <button type="submit" class="btn" style="background:#2c3e50;color:#fff;border-color:#2c3e50">Save Site Details</button>
          </div>
        </div>
      </form>
    </div>
  </div>
</div>

<div class="overlay" id="notificationSettingsModal" onclick="if(event.target===this)closeNotificationSettingsModal()">
  <div class="modal" style="width:900px">
    <div class="mhdr">
      <div>
        <div style="font-size:16px;font-weight:600">Notification Settings</div>
        <div style="font-size:12px;color:#888;margin-top:3px">Manage channels, recipients, greetings, and alert timing.</div>
      </div>
      <button class="btn" onclick="closeNotificationSettingsModal()">Close</button>
    </div>
    <div class="mbody">
      <form id="notificationSettingsForm" onsubmit="saveNotificationSettings(event)">
        <div class="form-grid">
          <div class="form-field">
            <label for="notifPollInterval">Status Poll Interval (sec)</label>
            <input id="notifPollInterval" type="number" min="5" name="poll_interval">
          </div>
          <div class="form-field">
            <label for="notifStatusRetries">Status Retry Count</label>
            <input id="notifStatusRetries" type="number" min="1" name="status_retries">
          </div>
          <div class="form-field">
            <label for="notifAlertRetries">Alert Retry Count</label>
            <input id="notifAlertRetries" type="number" min="1" name="alert_ping_retries">
          </div>
          <div class="form-field">
            <label for="notifCooldown">Cooldown Period (minutes)</label>
            <input id="notifCooldown" type="number" min="0" name="alert_cooldown_minutes">
          </div>
          <div class="form-field full" style="flex-direction:row;align-items:center;gap:16px;flex-wrap:wrap">
            <label style="margin:0;display:flex;align-items:center;gap:8px"><input id="notifOfflineToggle" type="checkbox"> Notify on offline</label>
            <label style="margin:0;display:flex;align-items:center;gap:8px"><input id="notifRecoveryToggle" type="checkbox"> Notify on recovery</label>
            <label style="margin:0;display:flex;align-items:center;gap:8px"><input id="notifDailyToggle" type="checkbox"> Send daily summary</label>
          </div>
          <div class="form-field">
            <label for="notifDailyTime">Daily Summary Time</label>
            <input id="notifDailyTime" name="daily_report_time" placeholder="08:00">
          </div>
          <div class="form-field full">
            <label for="notifGreeting">Greeting Template</label>
            <input id="notifGreeting" name="greeting_template" placeholder="Dear {name},">
          </div>

          <div class="form-field full">
            <label style="font-size:12px;color:#2c3e50">Email Settings</label>
          </div>
          <div class="form-field full" style="flex-direction:row;align-items:center;gap:10px">
            <input id="notifEmailEnabled" type="checkbox" style="width:auto">
            <label for="notifEmailEnabled" style="margin:0">Enable email notifications</label>
          </div>
          <div class="form-field">
            <label for="notifSmtpHost">SMTP Host</label>
            <input id="notifSmtpHost" name="smtp_host" placeholder="smtp.gmail.com">
          </div>
          <div class="form-field">
            <label for="notifSmtpPort">SMTP Port</label>
            <input id="notifSmtpPort" type="number" name="smtp_port" placeholder="587">
          </div>
          <div class="form-field full" style="flex-direction:row;align-items:center;gap:10px">
            <input id="notifSmtpTls" type="checkbox" style="width:auto">
            <label for="notifSmtpTls" style="margin:0">Use TLS</label>
          </div>
          <div class="form-field">
            <label for="notifSenderEmail">Sender Email</label>
            <input id="notifSenderEmail" name="sender_email" placeholder="alerts@example.com">
          </div>
          <div class="form-field">
            <label for="notifSenderPassword">Sender Password / App Password</label>
            <input id="notifSenderPassword" type="password" name="sender_password">
          </div>
          <div class="form-field full">
            <label for="notifSubjectPrefix">Subject Prefix</label>
            <input id="notifSubjectPrefix" name="subject_prefix" placeholder="[CAM ALERT]">
          </div>

          <div class="form-field full">
            <label style="font-size:12px;color:#2c3e50">WhatsApp Settings</label>
          </div>
          <div class="form-field full" style="flex-direction:row;align-items:center;gap:10px">
            <input id="notifWhatsappEnabled" type="checkbox" style="width:auto">
            <label for="notifWhatsappEnabled" style="margin:0">Enable WhatsApp notifications</label>
          </div>
          <div class="form-field">
            <label for="notifSid">Account SID</label>
            <input id="notifSid" name="account_sid">
          </div>
          <div class="form-field">
            <label for="notifToken">Auth Token</label>
            <input id="notifToken" type="password" name="auth_token">
          </div>
          <div class="form-field full">
            <label for="notifFromNumber">From Number</label>
            <input id="notifFromNumber" name="from_number" placeholder="whatsapp:+14155238886">
          </div>

          <div class="form-field full">
            <label style="font-size:12px;color:#2c3e50">Recipients</label>
            <div id="recipientRows" style="display:flex;flex-direction:column;gap:10px"></div>
            <button type="button" class="btn" onclick="addRecipientRow()">Add Recipient</button>
          </div>

          <div class="form-field full">
            <label for="offlineTemplate">Offline Template</label>
            <textarea id="offlineTemplate" style="min-height:90px"></textarea>
          </div>
          <div class="form-field full">
            <label for="recoveryTemplate">Recovery Template</label>
            <textarea id="recoveryTemplate" style="min-height:90px"></textarea>
          </div>
          <div class="form-field full">
            <label for="dailyTemplate">Daily Summary Template</label>
            <textarea id="dailyTemplate" style="min-height:90px"></textarea>
          </div>
        </div>
        <div style="display:flex;justify-content:space-between;align-items:center;margin-top:14px;gap:10px">
          <span id="notificationSettingsMsg" style="font-size:12px;color:#888"></span>
          <div style="display:flex;gap:8px">
            <button type="button" class="btn" onclick="closeNotificationSettingsModal()">Cancel</button>
            <button type="submit" class="btn" style="background:#2c3e50;color:#fff;border-color:#2c3e50">Save Notification Settings</button>
          </div>
        </div>
      </form>
    </div>
  </div>
</div>

<div class="overlay" id="userManagementModal" onclick="if(event.target===this)closeUserManagementModal()">
  <div class="modal" style="width:860px">
    <div class="mhdr">
      <div>
        <div style="font-size:16px;font-weight:600">User Management</div>
        <div style="font-size:12px;color:#888;margin-top:3px">Create users and assign viewer, operator, or admin access.</div>
      </div>
      <button class="btn" onclick="closeUserManagementModal()">Close</button>
    </div>
    <div class="mbody">
      <div class="form-grid" style="margin-bottom:14px">
        <div class="form-field">
          <label for="newUsername">Username</label>
          <input id="newUsername" placeholder="e.g. rajesh.kumar">
        </div>
        <div class="form-field">
          <label for="newUserPassword">Initial Password</label>
          <input id="newUserPassword" type="password" placeholder="Set a temporary password">
        </div>
        <div class="form-field">
          <label for="newUserRole">Role</label>
          <select id="newUserRole">
            <option value="viewer">Viewer</option>
            <option value="operator">Operator</option>
            <option value="admin">Admin</option>
          </select>
        </div>
        <div class="form-field" style="justify-content:flex-end">
          <label>&nbsp;</label>
          <button class="btn" type="button" onclick="createUser()">Create User</button>
        </div>
      </div>
      <div id="userMgmtMsg" style="font-size:12px;color:#888;margin-bottom:10px"></div>
      <div class="table-wrap">
        <table class="cam-table">
          <thead>
            <tr>
              <th>Username</th>
              <th>Role</th>
              <th>Status</th>
              <th>Set Password</th>
              <th></th>
            </tr>
          </thead>
          <tbody id="userRows"></tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<script>
const _CSRF_TOKEN = {{ csrf_token | tojson }};

// ── CSRF-aware fetch wrappers ───────────────────────────────────────────────
function apiFetch(url, opts){
  opts = opts || {};
  opts.headers = opts.headers || {};
  if(opts.method && opts.method.toUpperCase() !== 'GET'){
    opts.headers['X-CSRF-Token'] = _CSRF_TOKEN;
  }
  return fetch(url, opts);
}

let flt={zone:'',location:'',nvr:'',brand:'',status:''}, q='', selIps=new Set(), selCam=null, uploadFile=null, lastTs=Date.now(), formMode='add';
let cameraPage=1, cameraPageSize=100, cameraTotal=0;
let notificationRecipients=[];
let userRowsCache=[];

// ── XSS-safe HTML escaping ──────────────────────────────────────────────────
function escHtml(v){
  if(v==null)return '';
  return String(v).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&#x27;');
}

function hc(h){return h>=95?{bg:'#eafaf1',c:'#1e8449'}:h>=80?{bg:'#fef9e7',c:'#d68910'}:{bg:'#fdecea',c:'#c0392b'};}
function bc(b){b=(b||'').toLowerCase();return b==='hikvision'?{bg:'#ebf5fb',c:'#1a5276',l:'HIK'}:b==='dahua'?{bg:'#e8f8f5',c:'#0e6655',l:'DAH'}:{bg:'#fef5e7',c:'#9c640c',l:'PRA'};}
function dc(c){return c.maintenance?'#e67e22':c.online?'#27ae60':'#e74c3c';}
function ts(c){return c.maintenance?'maintenance':c.online?'online':'offline';}
function fmtDateTime(v){
  if(!v)return '—';
  const d=new Date(v);
  if(Number.isNaN(d.getTime()))return v;
  const dd=String(d.getDate()).padStart(2,'0');
  const mm=String(d.getMonth()+1).padStart(2,'0');
  const yyyy=d.getFullYear();
  let hh=d.getHours();
  const min=String(d.getMinutes()).padStart(2,'0');
  const ap=hh>=12?'PM':'AM';
  hh=hh%12||12;
  return `${dd}-${mm}-${yyyy} ${String(hh).padStart(2,'0')}:${min} ${ap}`;
}
function fmtDuration(s){
  if(!s)return '';
  const total=Math.max(0, parseInt(s,10) || 0);
  const d=Math.floor(total/86400);
  const h=Math.floor((total%86400)/3600);
  const m=Math.floor((total%3600)/60);
  const sec=total%60;
  const parts=[];
  if(d)parts.push(`${d}d`);
  if(h)parts.push(`${h}h`);
  if(m)parts.push(`${m}m`);
  if(sec)parts.push(`${sec}s`);
  return parts.join(' ') || '0s';
}
function statusHtml(c){
  if(c.maintenance) return '<span style="color:#e67e22;font-weight:500">Maintenance</span>';
  if(c.online) return '<span style="color:#27ae60;font-weight:500">Online</span>';
  return '<span style="color:#c0392b;font-weight:500">Offline</span>';
}

function sf(k,v,btn){
  flt[k]=v;
  cameraPage=1;
  const sel=
    k==='zone'?'.sidebar .sg:nth-child(1) .sb':
    k==='location'?'.sidebar .sg:nth-child(2) .sb':
    k==='nvr'?'.sidebar .sg:nth-child(3) .sb':
    k==='brand'?'.sidebar .sg:nth-child(4) .sb':
    '.fbar .chip';
  document.querySelectorAll(sel).forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');
  syncQuickFilters();
  loadCameras();
}
function fc(){q=document.getElementById('si').value.toLowerCase();cameraPage=1;loadCameras();}
function setQuickFilter(k,v){
  flt[k]=v;
  cameraPage=1;
  const sel=
    k==='zone'?'.sidebar .sg:nth-child(1) .sb':
    k==='location'?'.sidebar .sg:nth-child(2) .sb':
    k==='nvr'?'.sidebar .sg:nth-child(3) .sb':
    '.sidebar .sg:nth-child(4) .sb';
  document.querySelectorAll(sel).forEach(b=>b.classList.remove('active'));
  const allBtn=document.querySelector(
    k==='zone'?'.sidebar .sg:nth-child(1) .sb':
    k==='location'?'.sidebar .sg:nth-child(2) .sb':
    k==='nvr'?'.sidebar .sg:nth-child(3) .sb':
    '.sidebar .sg:nth-child(4) .sb'
  );
  let matched=[...document.querySelectorAll(sel)].find(b=>{
    const txt=(b.textContent || '').trim().toLowerCase();
    return v && (txt===v.toLowerCase());
  });
  (matched || allBtn)?.classList.add('active');
  syncQuickFilters();
  loadCameras();
}
function syncQuickFilters(){
  const zone=document.getElementById('zoneFilter');
  const location=document.getElementById('locationFilter');
  const nvr=document.getElementById('nvrFilter');
  const brand=document.getElementById('brandFilter');
  if(zone)zone.value=flt.zone || '';
  if(location)location.value=flt.location || '';
  if(nvr)nvr.value=flt.nvr || '';
  if(brand)brand.value=flt.brand || '';
}
function clearFilters(){
  flt={zone:'',location:'',nvr:'',brand:'',status:''};
  q='';
  cameraPage=1;
  document.getElementById('si').value='';
  document.querySelectorAll('.sidebar .sb, .fbar .chip').forEach(b=>b.classList.remove('active'));
  document.querySelector('.sidebar .sg:nth-child(1) .sb')?.classList.add('active');
  document.querySelector('.sidebar .sg:nth-child(2) .sb')?.classList.add('active');
  document.querySelector('.sidebar .sg:nth-child(3) .sb')?.classList.add('active');
  document.querySelector('.sidebar .sg:nth-child(4) .sb')?.classList.add('active');
  document.querySelector('.fbar .chip')?.classList.add('active');
  syncQuickFilters();
  loadCameras();
}
function updateCameraPageSize(){
  cameraPageSize=parseInt(document.getElementById('cameraPageSize').value || '100',10);
  cameraPage=1;
  loadCameras();
}
function changeCameraPage(delta){
  const totalPages=Math.max(1, Math.ceil(cameraTotal / cameraPageSize));
  const next=cameraPage+delta;
  if(next<1 || next>totalPages) return;
  cameraPage=next;
  loadCameras();
}
function setCameraPage(page){
  const totalPages=Math.max(1, Math.ceil(cameraTotal / cameraPageSize));
  if(page<1 || page>totalPages || page===cameraPage) return;
  cameraPage=page;
  loadCameras();
}
function buildPageTokens(current, total){
  if(total<=7){
    const out=[];
    for(let i=1;i<=total;i++) out.push(i);
    return out;
  }
  if(current<=4) return [1,2,3,4,5,'...',total];
  if(current>=total-3) return [1,'...',total-4,total-3,total-2,total-1,total];
  return [1,'...',current-1,current,current+1,'...',total];
}
function renderCameraPageNumbers(totalPages){
  const holder=document.getElementById('cameraPageNums');
  if(!holder) return;
  const tokens=buildPageTokens(cameraPage,totalPages);
  holder.innerHTML=tokens.map(tok=>{
    if(tok==='...') return `<span style="font-size:11px;color:#94a3b8;padding:0 2px">...</span>`;
    const active=tok===cameraPage;
    return `<button class="btn ${active?'active':''}" ${active?'disabled':''} onclick="setCameraPage(${tok})" style="${active?'background:#1f6feb;color:#fff;border-color:#1f6feb;':''}">${tok}</button>`;
  }).join('');
}
function syncCameraPager(total, itemsShown){
  const totalPages=Math.max(1, Math.ceil(total / cameraPageSize));
  document.getElementById('cameraPageInfo').textContent=`Page ${cameraPage} of ${totalPages} • ${itemsShown} shown • ${total} matched`;
  document.getElementById('cameraPrevBtn').disabled=cameraPage<=1;
  document.getElementById('cameraNextBtn').disabled=cameraPage>=totalPages;
  renderCameraPageNumbers(totalPages);
}

async function loadStats(){
  const r=await fetch('/api/stats');const s=await r.json();
  document.getElementById('sT').textContent=s.total;
  document.getElementById('sOn').textContent=s.online;
  document.getElementById('sOff').textContent=s.offline;
  document.getElementById('sMt').textContent=s.maintenance;
  const central=s.central||{};
  const consoleInd=document.getElementById('consoleInd');
  const consoleTxt=document.getElementById('consoleTxt');
  if(consoleTxt && consoleInd){
    consoleInd.classList.remove('ok','warn','info');
    if(!central.enabled){
      consoleInd.classList.add('warn');
      consoleTxt.textContent='Console: Disabled';
    }else if(central.healthy){
      consoleInd.classList.add('ok');
      consoleTxt.textContent='Console: Reporting';
    }else{
      consoleInd.classList.add('warn');
      consoleTxt.textContent='Console: Not Reporting';
    }
    const detail = central.last_success_at ? ('Last success: ' + new Date(central.last_success_at).toLocaleString()) : (central.last_error || 'No successful sync yet');
    consoleInd.title = detail;
  }
  const tunnelInd=document.getElementById('tunnelInd');
  const tunnelTxt=document.getElementById('tunnelTxt');
  if(tunnelInd && tunnelTxt){
    tunnelInd.classList.remove('ok','warn','info');
    const apiUrl = (central.api_url || '').trim();
    const publicTunnel = apiUrl && !apiUrl.includes('127.0.0.1') && !apiUrl.includes('localhost');
    if(publicTunnel){
      tunnelInd.classList.add('ok');
      tunnelTxt.textContent='Tunnel: Active';
    }else{
      tunnelInd.classList.add('warn');
      tunnelTxt.textContent='Tunnel: Local Only';
    }
    tunnelInd.title = apiUrl || 'No public tunnel URL configured';
  }
}

async function manualRefresh(){
  const btn=document.getElementById('refreshBtn');
  const txt=document.getElementById('pollTxt');
  btn.disabled=true;
  txt.textContent='Refreshing...';
  try{
    const r=await apiFetch('/api/refresh',{method:'POST'});
    const d=await r.json();
    if(!r.ok) throw new Error(d.error || 'Refresh failed');
    await loadCameras();
    await refreshMainModal();
    txt.textContent='Live';
  }catch(err){
    txt.textContent='Refresh failed';
    alert(err.message || 'Could not refresh camera status');
  }finally{
    btn.disabled=false;
  }
}

async function loadCameras(){
  const p=new URLSearchParams();
  if(flt.zone)p.set('zone',flt.zone);
  if(flt.location)p.set('location',flt.location);
  if(flt.nvr)p.set('nvr',flt.nvr);
  if(flt.brand)p.set('brand',flt.brand);
  if(flt.status)p.set('status',flt.status);
  if(q)p.set('q',q);
  p.set('page', String(cameraPage));
  p.set('page_size', String(cameraPageSize));
  const r=await fetch('/api/cameras?'+p);
  const data=await r.json();
  const cams=data.items || [];
  cameraTotal=data.total || 0;
  const totalPages=Math.max(1, Math.ceil(cameraTotal / cameraPageSize));
  if(cameraPage>totalPages){
    cameraPage=totalPages;
    return loadCameras();
  }
  document.getElementById('ccnt').textContent=`${cams.length} of ${cameraTotal} cameras`;
  document.getElementById('cg').innerHTML=!cams.length
    ? `<div class="empty-state">No cameras match the current filter.</div>`
    : `<table class="cam-table">
        <thead>
          <tr>
            <th style="width:36px"><input type="checkbox" class="row-check" onclick="toggleVisibleSelection(event, ${cams.map(c=>`'${c.ip}'`).join(',')})"></th>
            <th>Status</th>
            <th>Name</th>
            <th>IP Address</th>
            <th>Zone</th>
            <th>NVR</th>
            <th>Brand</th>
            <th>Health</th>
          </tr>
        </thead>
        <tbody>
          ${cams.map(c=>{
            const h=hc(c.health_7d||100),b=bc(c.brand),s=ts(c);
            const sel=selIps.has(c.ip)?'selected':'';
            return `<tr class="${s} ${sel}" data-row-ip="${escHtml(c.ip)}" onmouseenter="warmCamera('${escHtml(c.ip)}')" onclick="openModal('${escHtml(c.ip)}')">
              <td><input type="checkbox" class="row-check" ${selIps.has(c.ip)?'checked':''} onclick="event.stopPropagation();toggleSel('${escHtml(c.ip)}')"></td>
              <td><span class="status-pill ${s}"><span class="dot" style="background:${dc(c)}"></span>${s}</span></td>
              <td><div class="tn">${escHtml(c.name)||'—'}</div><div class="tip">${escHtml(c.location)||''}</div></td>
              <td class="tip">${escHtml(c.ip)}</td>
              <td>${escHtml(c.zone)||'—'}</td>
              <td>${c.nvr_name?`${escHtml(c.nvr_name)} <span style="color:#aaa">Ch.${escHtml(c.nvr_channel||1)}</span>`:'—'}</td>
              <td><span class="bb" style="background:${b.bg};color:${b.c}">${b.l}</span></td>
              <td><span class="hb" style="background:${h.bg};color:${h.c}">${Math.round(c.health_7d||100)}%</span></td>
            </tr>`;
          }).join('')}
        </tbody>
      </table>`;
  syncCameraPager(cameraTotal, cams.length);
  loadStats();
  document.getElementById('lastPoll').textContent='Last updated: '+fmtDateTime(new Date());
}

function toggleSel(ip){
  if(selIps.has(ip))selIps.delete(ip);else selIps.add(ip);
  const row=document.querySelector(`[data-row-ip="${ip}"]`);
  row?.classList.toggle('selected', selIps.has(ip));
  const checkbox=row?.querySelector('.row-check');
  if(checkbox)checkbox.checked=selIps.has(ip);
  const bar=document.getElementById('selBar');
  bar.classList.toggle('show',selIps.size>0);
  document.getElementById('selCount').textContent=selIps.size+' cameras selected';
}
function toggleVisibleSelection(event,...ips){
  event.stopPropagation();
  const shouldSelect=event.target.checked;
  ips.forEach(ip=>shouldSelect?selIps.add(ip):selIps.delete(ip));
  document.querySelectorAll('[data-row-ip]').forEach(row=>{
    const ip=row.getAttribute('data-row-ip');
    row.classList.toggle('selected', selIps.has(ip));
    const checkbox=row.querySelector('.row-check');
    if(checkbox)checkbox.checked=selIps.has(ip);
  });
  const bar=document.getElementById('selBar');
  bar.classList.toggle('show',selIps.size>0);
  document.getElementById('selCount').textContent=selIps.size+' cameras selected';
}
function clearSel(){
  selIps.clear();
  document.querySelectorAll('[data-row-ip]').forEach(row=>row.classList.remove('selected'));
  document.querySelectorAll('.row-check').forEach(box=>box.checked=false);
  document.getElementById('selBar').classList.remove('show');
}

function openAddCameraModal(){
  formMode='add';
  document.getElementById('addCameraMsg').textContent='';
  document.getElementById('addCameraForm').reset();
  document.getElementById('cameraFormTitle').textContent='Add Camera Manually';
  document.getElementById('cameraFormSubtitle').textContent='Create or update a single camera without uploading a file.';
  document.getElementById('camChannel').value=1;
  document.getElementById('camUser').value='admin';
  document.getElementById('camNvrIp').value='';
  document.getElementById('camOriginalIp').value='';
  document.getElementById('camIp').readOnly=false;
  document.getElementById('addCameraModal').classList.add('show');
}
function closeAddCameraModal(){
  document.getElementById('addCameraModal').classList.remove('show');
}
function openEditCameraModal(){
  if(!selCam)return;
  formMode='edit';
  document.getElementById('addCameraMsg').textContent='';
  document.getElementById('cameraFormTitle').textContent='Edit Camera';
  document.getElementById('cameraFormSubtitle').textContent='Update this camera and save the changes.';
  document.getElementById('camOriginalIp').value=selCam.ip || '';
  document.getElementById('camIp').value=selCam.ip || '';
  document.getElementById('camIp').readOnly=false;
  document.getElementById('camName').value=selCam.name || '';
  document.getElementById('camLocation').value=selCam.location || '';
  document.getElementById('camZone').value=selCam.zone || '';
  document.getElementById('camNvr').value=selCam.nvr_name || '';
  document.getElementById('camNvrIp').value=selCam.nvr_ip || '';
  document.getElementById('camChannel').value=selCam.nvr_channel || 1;
  document.getElementById('camBrand').value=(selCam.brand || '').toLowerCase();
  document.getElementById('camUser').value=selCam.username || 'admin';
  document.getElementById('camPassword').value=selCam.password || '';
  document.getElementById('camRtsp').value=selCam.rtsp_url || '';
  document.getElementById('camNotes').value=selCam.notes || '';
  document.getElementById('addCameraModal').classList.add('show');
}

async function openSiteSettingsModal(){
  const msg=document.getElementById('siteSettingsMsg');
  msg.style.color='#888';
  msg.textContent='Loading site settings...';
  document.getElementById('siteSettingsModal').classList.add('show');
  const r=await fetch('/api/site-settings');
  const d=await r.json();
  document.getElementById('siteEnabled').checked=!!d.enabled;
  document.getElementById('siteId').value=d.site_id || '';
  document.getElementById('siteName').value=d.site_name || '';
  document.getElementById('siteCampus').value=d.campus || '';
  document.getElementById('siteAddress').value=d.site_address || '';
  document.getElementById('siteContactName').value=d.contact_name || '';
  document.getElementById('siteContactPhone').value=d.contact_phone || '';
  document.getElementById('siteContactEmail').value=d.contact_email || '';
  document.getElementById('siteDashboardUrl').value=d.dashboard_url || '';
  document.getElementById('siteRefreshUrl').value=d.refresh_url || '';
  document.getElementById('siteApiUrl').value=d.api_url || '';
  document.getElementById('siteApiKey').value=d.api_key || '';
  const regBtn=document.getElementById('registerSiteBtn');
  const verBtn=document.getElementById('verifySiteBtn');
  const hasKey=d.api_key && d.api_key !== 'local-dev-key';
  if(regBtn){
    regBtn.textContent=hasKey ? 'Re-register Site' : 'Register This Site';
  }
  if(verBtn){
    verBtn.style.display=hasKey ? '' : 'none';
  }
  msg.textContent=hasKey ? 'API key is set. You can Save & Verify to test the connection.' : 'Paste the API key from the central dashboard and click Save & Verify.';
}
function closeSiteSettingsModal(){
  document.getElementById('siteSettingsModal').classList.remove('show');
}
async function saveSiteSettings(event){
  event.preventDefault();
  const form=event.target;
  const msg=document.getElementById('siteSettingsMsg');
  const fd=new FormData(form);
  const payload=Object.fromEntries(fd.entries());
  payload.enabled=document.getElementById('siteEnabled').checked;
  msg.style.color='#888';
  msg.textContent='Saving site settings...';
  const r=await apiFetch('/api/site-settings',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify(payload)
  });
  const d=await r.json();
  if(!r.ok){
    msg.style.color='#c0392b';
    msg.textContent=d.error || 'Could not save site settings';
    return;
  }
  document.getElementById('siteNameLabel').textContent=d.settings.site_name || 'Local Site';
  msg.style.color='#27ae60';
  msg.textContent='Saved. Central sync details are updated.';
  setTimeout(()=>closeSiteSettingsModal(), 500);
}
async function registerThisSite(){
  const msg=document.getElementById('siteSettingsMsg');
  const regBtn=document.getElementById('registerSiteBtn');
  regBtn.disabled=true;
  msg.style.color='#888';
  msg.textContent='Saving settings...';
  // Save form data first so config.ini has the latest API key before registering
  const form=document.getElementById('siteSettingsForm');
  const fd=new FormData(form);
  const payload=Object.fromEntries(fd.entries());
  payload.enabled=document.getElementById('siteEnabled').checked;
  const saveR=await apiFetch('/api/site-settings',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
  if(!saveR.ok){
    regBtn.disabled=false;
    const sd=await saveR.json();
    msg.style.color='#c0392b';
    msg.textContent=sd.error || 'Could not save settings before registering';
    return;
  }
  msg.textContent='Connecting to central dashboard...';
  const r=await apiFetch('/api/site-settings/register',{method:'POST'});
  const d=await r.json();
  if(!r.ok){
    regBtn.disabled=false;
    msg.style.color='#c0392b';
    msg.textContent=d.error || 'Could not register this site';
    return;
  }
  document.getElementById('siteApiUrl').value=d.central_api_url || '';
  document.getElementById('siteApiKey').value=d.api_key || '';
  document.getElementById('siteName').value=d.settings.site_name || '';
  document.getElementById('siteCampus').value=d.settings.campus || '';
  document.getElementById('siteId').value=d.settings.site_id || '';
  document.getElementById('siteEnabled').checked=!!d.settings.enabled;
  document.getElementById('siteNameLabel').textContent=d.settings.site_name || 'Local Site';
  regBtn.textContent='Re-register Site';
  regBtn.disabled=false;
  msg.style.color='#27ae60';
  msg.textContent=d.message || 'Site registered with central dashboard and API key saved.';
}
async function saveAndVerifySite(){
  const msg=document.getElementById('siteSettingsMsg');
  const verBtn=document.getElementById('verifySiteBtn');
  verBtn.disabled=true;
  msg.style.color='#888';
  msg.textContent='Saving settings...';
  const form=document.getElementById('siteSettingsForm');
  const fd=new FormData(form);
  const payload=Object.fromEntries(fd.entries());
  payload.enabled=document.getElementById('siteEnabled').checked;
  const saveR=await apiFetch('/api/site-settings',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
  if(!saveR.ok){
    verBtn.disabled=false;
    const sd=await saveR.json();
    msg.style.color='#c0392b';
    msg.textContent=sd.error || 'Could not save settings';
    return;
  }
  msg.textContent='Verifying connection to central dashboard...';
  const r=await apiFetch('/api/site-settings/verify',{method:'POST'});
  const d=await r.json();
  verBtn.disabled=false;
  if(!r.ok){
    msg.style.color='#c0392b';
    msg.textContent=d.error || 'Verification failed';
    return;
  }
  const regBtn=document.getElementById('registerSiteBtn');
  if(regBtn) regBtn.textContent='Re-register Site';
  msg.style.color='#27ae60';
  msg.textContent=d.message || 'Connection verified! SSO and sync are working.';
}

function renderRecipientRows(){
  const wrap=document.getElementById('recipientRows');
  wrap.innerHTML=notificationRecipients.map((r,idx)=>`
    <div style="border:1px solid #e5e7eb;border-radius:10px;padding:10px;background:#fafbfd">
      <div class="form-grid">
        <div class="form-field">
          <label>Name</label>
          <input value="${(r.name||'').replace(/"/g,'&quot;')}" oninput="updateRecipient(${idx},'name',this.value)" placeholder="Security Manager">
        </div>
        <div class="form-field">
          <label>Email</label>
          <input value="${(r.email||'').replace(/"/g,'&quot;')}" oninput="updateRecipient(${idx},'email',this.value)" placeholder="security@example.com">
        </div>
        <div class="form-field">
          <label>WhatsApp</label>
          <input value="${(r.whatsapp||'').replace(/"/g,'&quot;')}" oninput="updateRecipient(${idx},'whatsapp',this.value)" placeholder="whatsapp:+9198xxxxxxx">
        </div>
        <div class="form-field" style="justify-content:flex-end">
          <label>&nbsp;</label>
          <button type="button" class="btn" onclick="removeRecipient(${idx})">Remove</button>
        </div>
        <div class="form-field full" style="flex-direction:row;align-items:center;gap:16px;flex-wrap:wrap">
          <label style="margin:0;display:flex;align-items:center;gap:8px"><input type="checkbox" ${r.email_enabled?'checked':''} onchange="updateRecipient(${idx},'email_enabled',this.checked)"> Email</label>
          <label style="margin:0;display:flex;align-items:center;gap:8px"><input type="checkbox" ${r.whatsapp_enabled?'checked':''} onchange="updateRecipient(${idx},'whatsapp_enabled',this.checked)"> WhatsApp</label>
        </div>
      </div>
    </div>
  `).join('') || '<div style="font-size:12px;color:#888">No recipients added yet.</div>';
}
function addRecipientRow(){
  notificationRecipients.push({name:'',email:'',whatsapp:'',email_enabled:true,whatsapp_enabled:false});
  renderRecipientRows();
}
function removeRecipient(idx){
  notificationRecipients.splice(idx,1);
  renderRecipientRows();
}
function updateRecipient(idx,key,val){
  notificationRecipients[idx][key]=val;
}
async function openNotificationSettingsModal(){
  const msg=document.getElementById('notificationSettingsMsg');
  msg.style.color='#888';
  msg.textContent='Loading notification settings...';
  document.getElementById('notificationSettingsModal').classList.add('show');
  const r=await fetch('/api/notification-settings');
  const d=await r.json();
  document.getElementById('notifPollInterval').value=d.poll_interval ?? 10;
  document.getElementById('notifStatusRetries').value=d.status_retries ?? 1;
  document.getElementById('notifAlertRetries').value=d.alert_ping_retries ?? 6;
  document.getElementById('notifCooldown').value=d.alert_cooldown_minutes ?? 30;
  document.getElementById('notifOfflineToggle').checked=!!d.notify_offline;
  document.getElementById('notifRecoveryToggle').checked=!!d.notify_recovery;
  document.getElementById('notifDailyToggle').checked=!!d.daily_summary_enabled;
  document.getElementById('notifDailyTime').value=d.daily_report_time || '08:00';
  document.getElementById('notifGreeting').value=d.greeting_template || 'Dear {name},';
  document.getElementById('notifEmailEnabled').checked=!!d.email_enabled;
  document.getElementById('notifSmtpHost').value=d.smtp_host || '';
  document.getElementById('notifSmtpPort').value=d.smtp_port ?? 587;
  document.getElementById('notifSmtpTls').checked=!!d.smtp_use_tls;
  document.getElementById('notifSenderEmail').value=d.sender_email || '';
  document.getElementById('notifSenderPassword').value=d.sender_password || '';
  document.getElementById('notifSubjectPrefix').value=d.subject_prefix || '[CAM ALERT]';
  document.getElementById('notifWhatsappEnabled').checked=!!d.whatsapp_enabled;
  document.getElementById('notifSid').value=d.account_sid || '';
  document.getElementById('notifToken').value=d.auth_token || '';
  document.getElementById('notifFromNumber').value=d.from_number || '';
  document.getElementById('offlineTemplate').value=(d.templates||{}).offline || '';
  document.getElementById('recoveryTemplate').value=(d.templates||{}).recovery || '';
  document.getElementById('dailyTemplate').value=(d.templates||{}).daily || '';
  notificationRecipients=(d.recipients || []).map(x=>({...x}));
  renderRecipientRows();
  msg.textContent='Configure timing, channels, and recipients.';
}
function closeNotificationSettingsModal(){
  document.getElementById('notificationSettingsModal').classList.remove('show');
}
async function saveNotificationSettings(event){
  event.preventDefault();
  const msg=document.getElementById('notificationSettingsMsg');
  msg.style.color='#888';
  msg.textContent='Saving notification settings...';
  const payload={
    poll_interval: parseInt(document.getElementById('notifPollInterval').value || '10', 10),
    status_retries: parseInt(document.getElementById('notifStatusRetries').value || '1', 10),
    alert_ping_retries: parseInt(document.getElementById('notifAlertRetries').value || '6', 10),
    alert_cooldown_minutes: parseInt(document.getElementById('notifCooldown').value || '30', 10),
    notify_offline: document.getElementById('notifOfflineToggle').checked,
    notify_recovery: document.getElementById('notifRecoveryToggle').checked,
    daily_summary_enabled: document.getElementById('notifDailyToggle').checked,
    daily_report_time: document.getElementById('notifDailyTime').value,
    greeting_template: document.getElementById('notifGreeting').value,
    email_enabled: document.getElementById('notifEmailEnabled').checked,
    smtp_host: document.getElementById('notifSmtpHost').value,
    smtp_port: parseInt(document.getElementById('notifSmtpPort').value || '587', 10),
    smtp_use_tls: document.getElementById('notifSmtpTls').checked,
    sender_email: document.getElementById('notifSenderEmail').value,
    sender_password: document.getElementById('notifSenderPassword').value,
    subject_prefix: document.getElementById('notifSubjectPrefix').value,
    whatsapp_enabled: document.getElementById('notifWhatsappEnabled').checked,
    account_sid: document.getElementById('notifSid').value,
    auth_token: document.getElementById('notifToken').value,
    from_number: document.getElementById('notifFromNumber').value,
    templates: {
      offline: document.getElementById('offlineTemplate').value,
      recovery: document.getElementById('recoveryTemplate').value,
      daily: document.getElementById('dailyTemplate').value,
    },
    recipients: notificationRecipients,
  };
  const r=await apiFetch('/api/notification-settings',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify(payload)
  });
  const d=await r.json();
  if(!r.ok){
    msg.style.color='#c0392b';
    msg.textContent=d.error || 'Could not save notification settings';
    return;
  }
  msg.style.color='#27ae60';
  msg.textContent='Notification settings saved.';
  setTimeout(()=>closeNotificationSettingsModal(), 600);
}

function renderUserRows(users){
  userRowsCache=users || [];
  const tbody=document.getElementById('userRows');
  const roleColor={admin:'role-admin',operator:'role-operator',viewer:'role-viewer'};
  tbody.innerHTML=userRowsCache.map(u=>`
    <tr>
      <td><span class="uname-badge">${escHtml(u.username)}</span></td>
      <td>
        <select class="user-select ${roleColor[u.role]||'role-viewer'}" onchange="this.className='user-select '+({admin:'role-admin',operator:'role-operator',viewer:'role-viewer'}[this.value]||'role-viewer');updateUser(${u.id},{role:this.value})">
          <option value="viewer" ${u.role==='viewer'?'selected':''}>Viewer</option>
          <option value="operator" ${u.role==='operator'?'selected':''}>Operator</option>
          <option value="admin" ${u.role==='admin'?'selected':''}>Admin</option>
        </select>
      </td>
      <td>
        <select class="user-select ${u.active?'status-active':'status-disabled'}" onchange="this.className='user-select '+(this.value==='true'?'status-active':'status-disabled');updateUser(${u.id},{active:this.value==='true'})">
          <option value="true" ${u.active?'selected':''}>Active</option>
          <option value="false" ${!u.active?'selected':''}>Disabled</option>
        </select>
      </td>
      <td><div class="pwd-cell"><input id="pwd_${u.id}" class="pwd-input" type="password" placeholder="New password"><button type="button" class="btn-set" onclick="saveUserPassword(${u.id})">Set</button></div></td>
      <td></td>
    </tr>
  `).join('') || '<tr><td colspan="5" class="empty-state">No users found.</td></tr>';
}
async function openUserManagementModal(){
  const msg=document.getElementById('userMgmtMsg');
  msg.style.color='#888';
  msg.textContent='Loading users...';
  document.getElementById('userManagementModal').classList.add('show');
  const r=await fetch('/api/users');
  const d=await r.json();
  if(!r.ok){
    msg.style.color='#c0392b';
    msg.textContent=d.error || 'Could not load users';
    return;
  }
  renderUserRows(d);
  msg.textContent='Manage user roles and passwords from here.';
}
function closeUserManagementModal(){
  document.getElementById('userManagementModal').classList.remove('show');
}
async function createUser(){
  const msg=document.getElementById('userMgmtMsg');
  msg.style.color='#888';
  msg.textContent='Creating user...';
  const payload={
    username: document.getElementById('newUsername').value.trim(),
    password: document.getElementById('newUserPassword').value,
    role: document.getElementById('newUserRole').value,
    active: true,
  };
  const r=await apiFetch('/api/users',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
  const d=await r.json();
  if(!r.ok){
    msg.style.color='#c0392b';
    msg.textContent=d.error || 'Could not create user';
    return;
  }
  document.getElementById('newUsername').value='';
  document.getElementById('newUserPassword').value='';
  document.getElementById('newUserRole').value='viewer';
  msg.style.color='#27ae60';
  msg.textContent=`User ${d.user.username} created.`;
  const rows=await fetch('/api/users');
  renderUserRows(await rows.json());
}
async function updateUser(id, changes){
  const msg=document.getElementById('userMgmtMsg');
  const r=await apiFetch(`/api/users/${id}`,{method:'PATCH',headers:{'Content-Type':'application/json'},body:JSON.stringify(changes)});
  const d=await r.json();
  if(!r.ok){
    msg.style.color='#c0392b';
    msg.textContent=d.error || 'Could not update user';
    return;
  }
  msg.style.color='#27ae60';
  msg.textContent=`Updated ${d.user.username}.`;
}
async function saveUserPassword(id){
  const input=document.getElementById(`pwd_${id}`);
  const password=input.value;
  if(!password)return;
  await updateUser(id, {password});
  input.value='';
}

async function submitCameraForm(event){
  event.preventDefault();
  const form=event.target;
  const msg=document.getElementById('addCameraMsg');
  const payload=Object.fromEntries(new FormData(form).entries());
  payload.nvr_channel=parseInt(payload.nvr_channel || '1', 10);
  msg.style.color='#888';
  msg.textContent='Saving camera...';
  let r, d;
  try {
    r=await apiFetch('/api/camera',{
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify(payload)
    });
    // If session expired, @login_required redirects to /login (HTML).
    // Detect this before trying to parse JSON.
    if(r.url && r.url.includes('/login')){
      msg.style.color='#c0392b';
      msg.textContent='Session expired — please reload the page and log in again.';
      return;
    }
    const ct=r.headers.get('Content-Type')||'';
    if(!ct.includes('application/json')){
      const body=await r.text().catch(()=>'(unreadable)');
      msg.style.color='#c0392b';
      msg.textContent=`Server error (HTTP ${r.status}): ${body.slice(0,200)}`;
      console.error('Non-JSON camera save response:', r.status, r.url, body);
      return;
    }
    d=await r.json();
  } catch(err) {
    msg.style.color='#c0392b';
    msg.textContent='Network error — could not reach server. Is the local server running?';
    console.error('Camera save error:', err);
    return;
  }
  if(!r.ok){
    msg.style.color='#c0392b';
    msg.textContent=d.error || 'Could not save camera';
    return;
  }
  closeAddCameraModal();
  if(selCam && selCam.ip === payload.ip){
    selCam = {...selCam, ...payload};
  }
  await loadCameras();
  if(formMode==='edit'){
    await openModal(payload.ip);
  } else {
    window.location.reload();
  }
}

function copyTextValue(value){
  if(!value) return;
  if(navigator.clipboard && navigator.clipboard.writeText){
    navigator.clipboard.writeText(value);
  }
}
function resolvePlayerStream(streamUrls){
  if(!streamUrls) return '';
  const host=window.location.hostname;
  if(host==='127.0.0.1'||host==='localhost'){
    return streamUrls.player_local || streamUrls.player || '';
  }
  return streamUrls.player || streamUrls.player_local || '';
}
async function toggleFullscreen(elementId){
  const el=document.getElementById(elementId);
  if(!el) return;
  const img=el.querySelector('img');
  if(document.fullscreenElement){
    await document.exitFullscreen();
    if(img){
      img.style.width='100%';
      img.style.height='auto';
      img.style.maxHeight='65vh';
      img.style.objectFit='contain';
    }
    return;
  }
  if(el.requestFullscreen){
    await el.requestFullscreen();
    if(img){
      img.style.width='100vw';
      img.style.height='100vh';
      img.style.maxHeight='100vh';
      img.style.objectFit='contain';
    }
  }
}
let mainPreviewTimer=null;
function stopMainPreview(){
  if(mainPreviewTimer){
    window.clearInterval(mainPreviewTimer);
    mainPreviewTimer=null;
  }
}
function startMainPreview(snapshotUrl){
  stopMainPreview();
  if(!snapshotUrl) return;
  const img=document.getElementById('mainPreviewImg');
  if(!img) return;
  const refresh=()=>{ img.src=`${snapshotUrl}${snapshotUrl.includes('?')?'&':'?'}t=${Date.now()}`; };
  refresh();
  mainPreviewTimer=window.setInterval(refresh, 500);
}

function renderMainModal(c){
  selCam=c;
  document.getElementById('mNm').textContent=c.name;
  document.getElementById('mNmDot').style.background=dc(c);
  document.getElementById('mSb').textContent=`${c.ip}  •  ${c.location||''}, ${c.zone||''}  •  ${c.nvr_name||''} Ch.${c.nvr_channel||1}`;
  const pb=document.getElementById('pvb');
  pb.style.height='auto';
  pb.style.minHeight='0';
  pb.style.padding='0';
  pb.style.background='transparent';
  pb.style.border='none';
  const browserStream=(c.stream_urls||{}).browser || (c.stream_urls||{}).mjpeg || '';
  const snapshotStream=(c.stream_urls||{}).snapshot || '';
  const playerStream=resolvePlayerStream(c.stream_urls||{});
  if(c.online&&!c.maintenance&&c.stream_urls){
    if(playerStream){
      stopMainPreview();
      pb.innerHTML=`<div id="mainPreviewFrame" ondblclick="toggleFullscreen('mainPreviewFrame')" style="width:100%;max-width:100%;aspect-ratio:4 / 3;border:1px solid #cbd5e1;border-radius:12px;overflow:hidden;cursor:zoom-in;background:#fff">
        <iframe src="${playerStream}" allow="autoplay; fullscreen; picture-in-picture" allowfullscreen style="display:block;width:100%;height:100%;border:0;background:#fff"></iframe>
      </div>`;
    }else{
      pb.innerHTML=`<div id="mainPreviewFrame" ondblclick="toggleFullscreen('mainPreviewFrame')" style="width:100%;border:1px solid #cbd5e1;border-radius:12px;overflow:hidden;cursor:zoom-in;background:#fff">
        <img id="mainPreviewImg" style="width:100%;height:auto;max-height:65vh;object-fit:contain;background:#fff;display:block">
      </div>`;
      const img=document.getElementById('mainPreviewImg');
      if(browserStream){
        stopMainPreview();
        img.onerror=()=>{
          img.onerror=null;
          if(snapshotStream){
            startMainPreview(snapshotStream);
          }else{
            img.parentElement.innerHTML=`<div style="color:#aaa;font-size:12px;padding:24px;text-align:center;border:1px solid #e5e7eb;border-radius:12px;background:#fff">Live preview unavailable</div>`;
          }
        };
        img.src=browserStream;
      }else if(snapshotStream){
        img.onerror=()=>{
          img.onerror=null;
          img.parentElement.innerHTML=`<div style="color:#aaa;font-size:12px;padding:24px;text-align:center;border:1px solid #e5e7eb;border-radius:12px;background:#fff">Live preview unavailable</div>`;
        };
        startMainPreview(snapshotStream);
      }else{
        img.parentElement.innerHTML=`<div style="color:#aaa;font-size:12px;padding:24px;text-align:center;border:1px solid #e5e7eb;border-radius:12px;background:#fff">Live preview unavailable</div>`;
      }
    }
  } else if(!c.online){
    stopMainPreview();
    pb.innerHTML=`<div style="text-align:center;color:#c0392b;padding:28px;border:1px solid #e5e7eb;border-radius:12px;background:#fff"><div style="font-size:32px;margin-bottom:8px">🔴</div><div style="font-weight:600">Camera Offline</div><div style="font-size:11px;color:#aaa;margin-top:6px">Offline since ${fmtDateTime(c.offline_since)}</div></div>`;
  } else {
    stopMainPreview();
    pb.innerHTML=`<div style="text-align:center;color:#e67e22;padding:28px;border:1px solid #e5e7eb;border-radius:12px;background:#fff"><div style="font-size:32px;margin-bottom:8px">🟡</div><div style="font-weight:600">Maintenance Mode</div></div>`;
  }
  const h=hc(c.health_7d||100);
  document.getElementById('mRws').innerHTML=[
    ['Zone',escHtml(c.zone)||'—'],['Location',escHtml(c.location)||'—'],
    ['Brand',escHtml((c.brand||'').charAt(0).toUpperCase()+(c.brand||'').slice(1))],
    ['NVR / Channel',escHtml(c.nvr_name||'')+'  /  Ch.'+escHtml(c.nvr_channel||1)],
    ['NVR IP',escHtml(c.nvr_ip)||'—'],
    ['7-day Health',`<span style="background:${h.bg};color:${h.c};padding:2px 7px;border-radius:4px;font-weight:600">${Math.round(c.health_7d||100)}%</span>`],
    ['Status',statusHtml(c)],
    ['Notes',escHtml(c.notes)||'—'],
  ].map(([l,v])=>`<div class="irow"><span class="ilbl">${l}</span><span>${v}</span></div>`).join('');
  {% if role == 'admin' %}
  document.getElementById('mRws').innerHTML += `<div class="irow"><span class="ilbl">RTSP URL</span><span style="font-family:ui-monospace,SFMono-Regular,monospace;font-size:10px;color:var(--primary);word-break:break-all;overflow-wrap:anywhere">${escHtml((c.stream_urls||{}).rtsp||'—')}</span></div>`;
  {% endif %}
  document.getElementById('mtog').className='tog'+(c.maintenance?' on':'');
  const ev=(c.history||[]).slice(0,5).map(h=>`<div style="font-size:11px;padding:4px 0;border-bottom:1px solid #f5f5f5;color:${h.event==='offline'?'#c0392b':h.event==='online'?'#27ae60':'#888'}">${h.event==='offline'?'🔴':'🟢'} ${h.event.charAt(0).toUpperCase()+h.event.slice(1)} — ${fmtDateTime(h.ts)}${h.event==='online'&&h.duration_s?` <span style="color:#888">(${fmtDuration(h.duration_s)} downtime)</span>`:''}</div>`).join('');
  document.getElementById('mEv').innerHTML=ev||'<div style="color:#aaa;font-size:11px">No events recorded yet</div>';
  document.getElementById('ov').classList.add('show');
}
function warmCamera(ip){
  // Fire-and-forget: tells go2rtc to open the RTSP connection now,
  // so it's ready by the time the player iframe loads.
  apiFetch('/api/camera/'+encodeURIComponent(ip)+'/warm',{method:'POST'}).catch(()=>{});
}
async function openModal(ip){
  warmCamera(ip); // start RTSP warm-up immediately, parallel to API fetch
  const r=await fetch('/api/camera/'+ip);
  const c=await r.json();
  renderMainModal(c);
}
async function refreshMainModal(){
  if(!selCam || !document.getElementById('ov').classList.contains('show')) return;
  const r=await fetch('/api/camera/'+selCam.ip);
  const c=await r.json();
  renderMainModal(c);
}

function co(){stopMainPreview();document.getElementById('ov').classList.remove('show');selCam=null;}

async function deleteSelectedCamera(){
  if(!selCam) return;
  const label=selCam.name || selCam.ip;
  if(!confirm(`Delete ${label}?`)) return;
  const typed=prompt(`Type DELETE to confirm removing ${label}`);
  if(typed !== 'DELETE') return;
  const r=await apiFetch('/api/camera/'+encodeURIComponent(selCam.ip), {method:'DELETE'});
  const d=await r.json();
  if(!r.ok){
    alert(d.error || 'Could not delete camera');
    return;
  }
  co();
  clearSel();
  loadCameras();
}

function applyMaintenanceStateLocally(ip, maintenance){
  if(selCam && selCam.ip===ip){
    selCam.maintenance=maintenance;
    document.getElementById('mtog').className='tog'+(maintenance?' on':'');
    const statusEl=document.getElementById('mSt');
    if(statusEl)statusEl.innerHTML=statusHtml(selCam);
    const dot=document.getElementById('mNmDot');
    if(dot)dot.style.background=dc(selCam);
  }
}

async function tm(){
  if(!selCam)return;
  const toggle=document.getElementById('mtog');
  if(toggle?.dataset.busy==='1')return;
  const newState=!selCam.maintenance;
  const prevState=selCam.maintenance;
  if(toggle)toggle.dataset.busy='1';
  applyMaintenanceStateLocally(selCam.ip, newState);
  try{
    const r=await apiFetch('/api/camera/'+selCam.ip+'/maintenance',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({state:newState})});
    const d=await r.json();
    if(!r.ok) throw new Error(d.error || 'Could not update maintenance mode');
    loadCameras();
    refreshMainModal();
  }catch(err){
    applyMaintenanceStateLocally(selCam.ip, prevState);
    alert(err.message || 'Could not update maintenance mode');
  }finally{
    if(toggle)toggle.dataset.busy='0';
  }
}

async function bulkMaint(state){
  if(!selIps.size)return;
  const ips=[...selIps];
  if(selCam && ips.includes(selCam.ip)){
    applyMaintenanceStateLocally(selCam.ip, state);
  }
  await apiFetch('/api/bulk/maintenance',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({ips,state})});
  clearSel();loadCameras();refreshMainModal();
}
async function showBulkZone(){
  const z=prompt('Enter zone name:');if(!z)return;
  await apiFetch('/api/bulk/zone',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({ips:[...selIps],zone:z})});
  clearSel();loadCameras();
}
async function showBulkNvr(){
  const n=prompt('Enter NVR name:');if(!n)return;
  await apiFetch('/api/bulk/nvr',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({ips:[...selIps],nvr:n})});
  clearSel();loadCameras();
}

async function deleteSelectedCameras(){
  if(!selIps.size) return;
  const count=selIps.size;
  const selected=[...selIps].map(ip=>{
    const row=document.querySelector(`[data-row-ip="${ip}"]`);
    const name=row?.querySelector('.tn')?.textContent?.trim();
    return name && name !== '—' ? `${name} (${ip})` : ip;
  });
  const preview=selected.slice(0,5).join('\\n');
  const extra=count>5 ? `\\n...and ${count-5} more` : '';
  if(!confirm(`Delete ${count} selected camera${count===1?'':'s'}?\\n\\n${preview}${extra}`)) return;
  const typed=prompt(`Type DELETE to confirm removing:\\n\\n${preview}${extra}`);
  if(typed !== 'DELETE') return;
  const r=await apiFetch('/api/bulk/delete',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({ips:[...selIps]})
  });
  const d=await r.json();
  if(!r.ok){
    alert(d.error || 'Could not delete selected cameras');
    return;
  }
  if(selCam && selIps.has(selCam.ip)){
    co();
  }
  clearSel();
  loadCameras();
}

async function previewUpload(input){
  const file=input.files[0];if(!file)return;
  uploadFile=file;
  const fd=new FormData();fd.append('file',file);
  const r=await apiFetch('/api/bulk/preview',{method:'POST',body:fd});
  const d=await r.json();
  if(!r.ok){
    document.getElementById('diffArea').innerHTML=`<div style="padding:12px;border:1px solid #fecaca;background:#fff1f2;color:#b91c1c;border-radius:12px;font-size:12px">${d.error||'Preview failed'}</div>`;
    document.getElementById('uploadActions').style.display='none';
    return;
  }
  const rows=d.preview||[];
  const summary=d.summary||{};
  const errors=d.errors||[];
  const warnings=d.warnings||[];
  const errorHtml=errors.length?`<div style="margin-bottom:10px;padding:12px;border:1px solid #fecaca;background:#fff1f2;color:#991b1b;border-radius:12px;font-size:12px;line-height:1.55">
    <div style="font-weight:700;margin-bottom:6px">Errors to fix before import</div>
    <ul style="margin-left:18px">${errors.slice(0,8).map(x=>`<li>Row ${x.row}: ${x.text}</li>`).join('')}${errors.length>8?`<li>...and ${errors.length-8} more</li>`:''}</ul>
  </div>`:'';
  const warningHtml=warnings.length?`<div style="margin-bottom:10px;padding:12px;border:1px solid #fed7aa;background:#fff7ed;color:#9a3412;border-radius:12px;font-size:12px;line-height:1.55">
    <div style="font-weight:700;margin-bottom:6px">Warnings to review</div>
    <ul style="margin-left:18px">${warnings.slice(0,8).map(x=>`<li>Row ${x.row}: ${x.text}</li>`).join('')}${warnings.length>8?`<li>...and ${warnings.length-8} more</li>`:''}</ul>
  </div>`:'';
  const html=`<div style="font-size:12px;margin-bottom:8px"><strong>${summary.total||rows.length}</strong> cameras found — <span style="color:#27ae60">${summary.new||0} new</span>, <span style="color:#e67e22">${summary.updates||0} updates</span>, <span style="color:#b91c1c">${summary.errors||0} errors</span>, <span style="color:#b45309">${summary.warnings||0} warnings</span></div>
  ${errorHtml}
  ${warningHtml}
  <div style="max-height:240px;overflow-y:auto"><table class="diff-table"><thead><tr><th>Row</th><th>IP</th><th>Name</th><th>NVR</th><th>NVR IP</th><th>Brand</th><th>Action</th><th>Checks</th></tr></thead><tbody>
  ${rows.slice(0,30).map(r=>`<tr>
    <td>${r.row}</td>
    <td style="font-family:monospace">${r.ip}</td>
    <td>${r.name||'—'}</td>
    <td>${r.nvr_name||'—'}</td>
    <td style="font-family:monospace">${r.nvr_ip||'—'}</td>
    <td>${r.brand||'—'}</td>
    <td style="color:${r.action==='add'?'#27ae60':'#e67e22'};font-weight:600">${r.action}</td>
    <td>${(r.messages||[]).length? (r.messages||[]).map(m=>`<div style="color:${m.level==='error'?'#b91c1c':'#9a3412'}">${m.level}: ${m.text}</div>`).join('') : '<span style="color:#15803d">ok</span>'}</td>
  </tr>`).join('')}
  ${rows.length>30?`<tr><td colspan="8" style="color:#aaa;text-align:center">...and ${rows.length-30} more</td></tr>`:''}
  </tbody></table></div>`;
  document.getElementById('diffArea').innerHTML=html;
  document.getElementById('uploadActions').style.display=d.blocking?'none':'block';
}

async function confirmUpload(){
  if(!uploadFile)return;
  const fd=new FormData();fd.append('file',uploadFile);
  const r=await apiFetch('/api/bulk/import',{method:'POST',body:fd});
  const d=await r.json();
  if(!r.ok){
    alert(d.error || 'Import failed');
    return;
  }
  document.getElementById('uploadModal').classList.remove('show');
  alert(`Import complete: ${d.added} added, ${d.updated} updated${(d.warnings||[]).length?`, ${(d.warnings||[]).length} warnings reviewed`:''}`);
  loadCameras();
}

// Keep dashboard stats/table fresh without re-rendering the open video iframe.
// Re-rendering the modal forces go2rtc renegotiation (MSE -> RTC) repeatedly.
setInterval(async()=>{await loadCameras();}, 30000);
loadCameras();
</script>

<!-- Audit Log Route -->
<script>
// Handled server-side at /audit
</script>
</body></html>"""

# Audit log page
@app.route("/setup")
@login_required
def setup():
    if SETUP_COMPLETE:
        return redirect(url_for("dashboard"))
    return render_template("setup.html", user=actor_name(), site_name=SITE_NAME, csrf_token=_get_csrf_token())

@app.route("/api/setup/complete", methods=["POST"])
@login_required
@role_required("admin")
def api_setup_complete():
    mark_setup_complete(True)
    return jsonify({"success": True})
@app.route("/audit")
@login_required
def audit_page():
    return render_template_string(
        AUDIT_HTML,
        user=actor_name(),
        site_name=SITE_NAME,
        role=current_user.role,
    )

@app.route("/camera-logs")
@login_required
def camera_logs_page():
    return render_template_string(CAMERA_LOG_HTML, user=actor_name(), zones=db.get_zones(), nvrs=db.get_nvrs(), site_name=SITE_NAME, role=current_user.role)

NVR_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>NVR Monitor — Kumarans</title>
<style>
:root{--bg:#f3f6fb;--surface:#ffffff;--line:#dbe3ef;--line-strong:#c8d4e5;--text:#1f2a37;--muted:#6b7a90;--primary:#1f6feb;--primary-soft:#e8f1ff;--ok:#16a34a;--warn:#d97706;--danger:#dc2626;--shadow:0 10px 30px rgba(15,23,42,.06)}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:system-ui,sans-serif;background:linear-gradient(180deg,#f8fafc 0%,var(--bg) 100%);font-size:13px;color:var(--text);min-height:100vh;display:flex;flex-direction:column}
.topbar{padding:14px 20px;background:rgba(255,255,255,.96);border-bottom:1px solid var(--line);position:sticky;top:0;z-index:100;backdrop-filter:blur(12px)}
.top-row{display:grid;grid-template-columns:1fr auto 1fr;align-items:center;gap:14px}
.brand img{width:342px;max-width:100%;height:auto;object-fit:contain;filter:drop-shadow(0 2px 6px rgba(15,23,42,.10))}
.titlebar{justify-self:center;display:flex;flex-direction:column;align-items:center;gap:8px;text-align:center}
.title-main{font-size:20px;font-weight:700;line-height:1.1;color:var(--text)}
.site-pill{font-size:11px;font-weight:700;color:#2159b3;background:var(--primary-soft);border:1px solid #cfe0ff;border-radius:999px;padding:5px 12px;line-height:1}
.top-actions{display:flex;justify-content:flex-end;gap:8px;flex-wrap:wrap}
.nav-link,.btn,.chip{padding:8px 12px;border-radius:10px;font-size:12px;border:1px solid var(--line);background:#fff;color:#475569;text-decoration:none;font-weight:600;cursor:pointer}
.nav-link:hover,.btn:hover,.chip:hover{background:#f8fbff;border-color:var(--line-strong)}
.chip.active,.btn.primary{background:var(--primary);color:#fff;border-color:var(--primary)}
.shell{padding:18px 20px 22px;flex:1}
.hero{display:grid;grid-template-columns:1.3fr .9fr;gap:14px;margin-bottom:14px}
.panel{background:var(--surface);border:1px solid var(--line);border-radius:16px;box-shadow:var(--shadow)}
.panel.pad{padding:18px}
.hero h1{font-size:24px;line-height:1.1;margin-bottom:8px}
.hero p{font-size:13px;color:var(--muted);max-width:760px;line-height:1.6}
.info-list{display:grid;gap:8px;margin-top:12px}
.info-item{padding:10px 12px;border-radius:12px;background:#f8fafc;border:1px solid #edf2f7;color:#516174}
.info-item strong{color:#0f172a}
.filters{display:flex;gap:8px;align-items:center;flex-wrap:wrap;margin-bottom:14px}
.search,.select{padding:9px 12px;border:1px solid var(--line);border-radius:10px;font-size:12px;background:#fff;color:var(--text);outline:none}
.search:focus,.select:focus{border-color:var(--primary);box-shadow:0 0 0 3px rgba(31,111,235,.10)}
.kpi-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin-bottom:14px}
.kpi{padding:16px;border-radius:16px;background:var(--surface);border:1px solid var(--line);box-shadow:var(--shadow)}
.kpi .k{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em}
.kpi .v{font-size:28px;font-weight:800;margin-top:8px;color:var(--text)}
.kpi .s{font-size:12px;color:#70839b;margin-top:6px}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(320px,1fr));gap:14px}
.card{background:var(--surface);border:1px solid var(--line);border-radius:18px;padding:18px;box-shadow:var(--shadow);cursor:pointer;transition:transform .15s ease, box-shadow .15s ease}
.card:hover{transform:translateY(-2px);box-shadow:0 14px 28px rgba(15,23,42,.10)}
.card-head{display:flex;justify-content:space-between;align-items:flex-start;gap:12px;margin-bottom:14px}
.card-title{font-size:18px;font-weight:700}
.card-meta{font-size:12px;color:#70839b;margin-top:6px;display:grid;gap:4px}
.status-badge{display:inline-flex;align-items:center;gap:6px;padding:5px 10px;border-radius:999px;font-size:11px;font-weight:700;text-transform:capitalize}
.status-badge .dot{width:8px;height:8px;border-radius:50%}
.status-online{background:#ebf8ef;color:#157f42}.status-online .dot{background:#22c55e}
.status-offline{background:#fef2f2;color:#b91c1c}.status-offline .dot{background:#ef4444}
.status-unconfigured{background:#fff7ed;color:#b45309}.status-unconfigured .dot{background:#f59e0b}
.metric-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:10px;margin-bottom:12px}
.metric{padding:12px;border-radius:12px;background:#f8fafc;border:1px solid #edf2f7}
.metric .m{font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em}
.metric .n{font-size:20px;font-weight:800;margin-top:6px}
.pill-row{display:flex;gap:6px;flex-wrap:wrap}
.pill{display:inline-flex;align-items:center;padding:4px 9px;border-radius:999px;background:#eef4ff;color:#365f9b;font-size:11px;font-weight:600}
.subtext{font-size:12px;color:#6b7a90;line-height:1.55}
.overlay{display:none;position:fixed;inset:0;background:rgba(15,23,42,.45);z-index:240;align-items:flex-start;justify-content:center;padding:40px 12px;overflow-y:auto}
.overlay.show{display:flex}
.modal{background:#fff;border-radius:18px;width:860px;max-width:100%;border:1px solid var(--line);box-shadow:0 24px 60px rgba(15,23,42,.16)}
.mhdr{display:flex;align-items:flex-start;justify-content:space-between;padding:18px 22px;border-bottom:1px solid #edf2f7}
.mbody{padding:18px 22px;display:grid;gap:16px}
.detail-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:10px}
.detail{padding:12px;border-radius:12px;background:#f8fafc;border:1px solid #edf2f7}
.detail .k{font-size:10px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em}
.detail .v{font-size:18px;font-weight:800;margin-top:6px}
.section-title{font-size:13px;font-weight:700;color:#0f172a;margin-bottom:8px}
.camera-table{width:100%;border-collapse:collapse;background:#fff;border:1px solid #edf2f7;border-radius:14px;overflow:hidden}
.camera-table th,.camera-table td{padding:10px 12px;text-align:left;font-size:12px;border-bottom:1px solid #edf2f7}
.camera-table th{background:#f8fafc;color:#5d6d83;font-weight:700;text-transform:uppercase;font-size:11px;letter-spacing:.08em}
.empty{padding:28px;color:#7a879a;text-align:center;background:#fff;border:1px dashed #cbd5e1;border-radius:16px}
.footer-note{padding:0 18px 18px;color:#7a879a;font-size:11px;line-height:1.6;text-align:center}
@media (max-width: 1080px){.hero{grid-template-columns:1fr}.kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))}.detail-grid{grid-template-columns:repeat(2,minmax(0,1fr))}}
@media (max-width: 860px){.top-row{grid-template-columns:1fr;justify-items:start}.titlebar{justify-self:start;align-items:flex-start}.top-actions{justify-content:flex-start}}
@media (max-width: 640px){.shell,.topbar{padding-left:14px;padding-right:14px}.kpi-grid,.detail-grid{grid-template-columns:1fr}.filters,.top-actions{align-items:stretch}.btn,.nav-link,.chip,.search,.select{width:100%}}
</style></head>
<body>
<div class="topbar">
  <div class="top-row">
    <div class="brand"><img src="https://kumarans.org/images/Sri%20Kumaran%20Childrens%20Home.png" alt="Sri Kumaran logo"></div>
    <div class="titlebar">
      <span class="title-main">NVR Monitoring</span>
      <strong class="site-pill">{{ site_name }}</strong>
    </div>
    <div class="top-actions">
      <a href="/" class="nav-link">Dashboard</a>
      <a href="/reports" class="nav-link">Reports</a>
      <a href="/camera-logs" class="nav-link">Camera Logs</a>
      <a href="/audit" class="nav-link">Audit Log</a>
      <a href="/logout" class="nav-link">Sign out ({{ user }})</a>
    </div>
  </div>
</div>

<div class="shell">
  <div class="hero">
    <div class="panel pad">
      <h1>NVR Reachability</h1>
      <p>This first version checks only whether each NVR responds to direct ping. It does not infer health from cameras and does not yet pull HDD, recording, or vendor diagnostics.</p>
      <div class="info-list">
        <div class="info-item"><strong>Online</strong> means the configured NVR IP responds to ping.</div>
        <div class="info-item"><strong>Offline</strong> means the configured NVR IP does not respond to ping.</div>
        <div class="info-item"><strong>Unconfigured</strong> means the NVR name exists but no NVR IP is configured yet.</div>
      </div>
    </div>
    <div class="panel pad">
      <div class="section-title">Configuration Note</div>
      <div class="subtext">Set the <strong>NVR IP</strong> in camera details for every camera mapped to that NVR. Cameras with the same <strong>NVR Name</strong> and <strong>NVR IP</strong> are grouped into one NVR card here.</div>
      <div class="subtext" style="margin-top:12px">The page refreshes every 10 seconds so operators can quickly see if an NVR is reachable from the local server.</div>
    </div>
  </div>

  <div class="kpi-grid" id="nvrKpis">
    <div class="kpi"><div class="k">NVRs</div><div class="v">—</div><div class="s">Configured groups</div></div>
    <div class="kpi"><div class="k">Online</div><div class="v" style="color:#15803d">—</div><div class="s">Responding to ping</div></div>
    <div class="kpi"><div class="k">Offline</div><div class="v" style="color:#b91c1c">—</div><div class="s">Not responding to ping</div></div>
    <div class="kpi"><div class="k">Unconfigured</div><div class="v" style="color:#b45309">—</div><div class="s">Missing NVR IP</div></div>
  </div>

  <div class="filters">
    <input id="nvrSearch" class="search" type="search" placeholder="Search NVR, IP, zone, location, camera name or IP..." oninput="debouncedLoadNvrs()">
    <select id="nvrStatus" class="select" style="width:170px" onchange="loadNvrs()">
      <option value="">All Statuses</option>
      <option value="online">Online</option>
      <option value="offline">Offline</option>
      <option value="unconfigured">Unconfigured</option>
    </select>
    <select id="nvrZone" class="select" style="width:170px" onchange="loadNvrs()">
      <option value="">All Zones</option>
      {% for z in zones %}<option value="{{ z }}">{{ z }}</option>{% endfor %}
    </select>
    <button class="btn" onclick="clearNvrFilters()">Clear Filters</button>
    <div style="flex:1"></div>
    <span id="nvrCount" style="font-size:12px;color:#70839b">Loading NVRs...</span>
  </div>

  <div class="grid" id="nvrGrid"></div>
</div>

<div class="footer-note">
  © Sri Kumaran Childrens Home Educational Council. All rights reserved. Authorized operational use only. Activity on this monitoring system may be logged and reviewed.
</div>

<div class="overlay" id="nvrOv" onclick="if(event.target===this)closeNvrModal()">
  <div class="modal">
    <div class="mhdr">
      <div>
        <div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap">
          <div id="nvrModalTitle" style="font-size:18px;font-weight:700">NVR</div>
          <span id="nvrModalBadge" class="status-badge status-online"><span class="dot"></span><span>online</span></span>
        </div>
        <div id="nvrModalSub" style="font-size:12px;color:#70839b;margin-top:6px"></div>
      </div>
      <button class="btn" onclick="closeNvrModal()">Close</button>
    </div>
    <div class="mbody">
      <div class="detail-grid" id="nvrDetailGrid"></div>
      <div>
        <div class="section-title">Mapped Cameras</div>
        <div style="overflow:auto">
          <table class="camera-table">
            <thead><tr><th>Camera</th><th>IP</th><th>Zone</th><th>Location</th><th>Channel</th></tr></thead>
            <tbody id="nvrCameraRows"></tbody>
          </table>
        </div>
      </div>
    </div>
  </div>
</div>

<script>
function escHtml(v){if(v==null)return '';return String(v).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&#x27;');}
let nvrSearchTimer=null;
let nvrItems=[];
let activeNvr=null;
function fmtDateTime(v){
  if(!v) return '—';
  const d=new Date(v);
  if(Number.isNaN(d.getTime())) return v;
  const dd=String(d.getDate()).padStart(2,'0');
  const mm=String(d.getMonth()+1).padStart(2,'0');
  const yyyy=d.getFullYear();
  let hh=d.getHours();
  const min=String(d.getMinutes()).padStart(2,'0');
  const ap=hh>=12?'PM':'AM';
  hh=hh%12||12;
  return `${dd}-${mm}-${yyyy} ${String(hh).padStart(2,'0')}:${min} ${ap}`;
}
function statusBadge(status){
  return `<span class="status-badge status-${status}"><span class="dot"></span><span>${status}</span></span>`;
}
function debouncedLoadNvrs(){
  window.clearTimeout(nvrSearchTimer);
  nvrSearchTimer=window.setTimeout(loadNvrs, 250);
}
function clearNvrFilters(){
  document.getElementById('nvrSearch').value='';
  document.getElementById('nvrStatus').value='';
  document.getElementById('nvrZone').value='';
  loadNvrs();
}
function renderKpis(summary){
  const cards=document.querySelectorAll('#nvrKpis .kpi');
  cards[0].querySelector('.v').textContent=summary.total || 0;
  cards[1].querySelector('.v').textContent=summary.online || 0;
  cards[2].querySelector('.v').textContent=summary.offline || 0;
  cards[3].querySelector('.v').textContent=summary.unconfigured || 0;
}
function renderNvrs(items){
  const grid=document.getElementById('nvrGrid');
  document.getElementById('nvrCount').textContent=`${items.length} NVR${items.length===1?'':'s'} shown`;
  if(!items.length){
    grid.innerHTML='<div class="empty">No NVR groups match the current filters.</div>';
    return;
  }
  grid.innerHTML=items.map((item, idx)=>`<article class="card" onclick="openNvrModal(${idx})">
    <div class="card-head">
      <div>
        <div class="card-title">${escHtml(item.nvr_name)}</div>
        <div class="card-meta">
          <div>IP: ${item.nvr_ip || 'Not configured'}</div>
          <div>${(item.zones||[]).join(' • ') || 'Unassigned zone'}</div>
        </div>
      </div>
      ${statusBadge(item.status)}
    </div>
    <div class="metric-grid">
      <div class="metric"><div class="m">Mapped Cameras</div><div class="n">${item.total_cameras}</div></div>
      <div class="metric"><div class="m">Last Checked</div><div class="n" style="font-size:14px">${fmtDateTime(item.last_checked)}</div></div>
    </div>
    <div class="pill-row">
      ${(item.brands||[]).map(brand=>`<span class="pill">${brand}</span>`).join('') || '<span class="pill">No brand</span>'}
    </div>
    <div class="subtext" style="margin-top:12px">${item.status_summary}</div>
  </article>`).join('');
}
async function loadNvrs(){
  const p=new URLSearchParams();
  const q=document.getElementById('nvrSearch').value.trim();
  const status=document.getElementById('nvrStatus').value;
  const zone=document.getElementById('nvrZone').value;
  if(q) p.set('q', q);
  if(status) p.set('status', status);
  if(zone) p.set('zone', zone);
  const r=await fetch('/api/nvrs?'+p.toString());
  const data=await r.json();
  nvrItems=data.items || [];
  renderKpis(data.summary || {});
  renderNvrs(nvrItems);
}
function openNvrModal(index){
  activeNvr=nvrItems[index];
  if(!activeNvr) return;
  document.getElementById('nvrModalTitle').textContent=activeNvr.nvr_name;
  document.getElementById('nvrModalBadge').className=`status-badge status-${activeNvr.status}`;
  document.getElementById('nvrModalBadge').innerHTML=`<span class="dot"></span><span>${activeNvr.status}</span>`;
  document.getElementById('nvrModalSub').textContent=`${activeNvr.status_summary} • ${activeNvr.total_cameras} mapped cameras`;
  document.getElementById('nvrDetailGrid').innerHTML=[
    ['NVR IP', activeNvr.nvr_ip || 'Not configured'],
    ['Status', activeNvr.status],
    ['Mapped Cameras', activeNvr.total_cameras],
    ['Last Checked', fmtDateTime(activeNvr.last_checked)],
    ['Zones', (activeNvr.zones||[]).join(', ') || '—'],
    ['Locations', (activeNvr.locations||[]).join(', ') || '—'],
    ['Brands', (activeNvr.brands||[]).join(', ') || '—'],
    ['Next Step', activeNvr.nvr_ip ? 'Ready for deeper NVR health checks later' : 'Set NVR IP in camera details'],
  ].map(([k,v])=>`<div class="detail"><div class="k">${k}</div><div class="v" style="font-size:${k==='Locations' || k==='Zones' || k==='Brands' || k==='Next Step' ? '14px' : '18px'}">${v}</div></div>`).join('');
  document.getElementById('nvrCameraRows').innerHTML=(activeNvr.cameras||[]).map(cam=>`<tr>
    <td><div style="font-weight:700">${cam.name||'—'}</div><div style="font-size:11px;color:#70839b">${cam.brand||'—'}</div></td>
    <td style="font-family:monospace;font-size:11px;color:#70839b">${cam.ip||''}</td>
    <td>${cam.zone||'—'}</td>
    <td>${cam.location||'—'}</td>
    <td>${cam.nvr_channel||1}</td>
  </tr>`).join('') || '<tr><td colspan="5" style="padding:18px;text-align:center;color:#7a879a">No mapped cameras.</td></tr>';
  document.getElementById('nvrOv').classList.add('show');
}
function closeNvrModal(){
  document.getElementById('nvrOv').classList.remove('show');
  activeNvr=null;
}
loadNvrs();
setInterval(loadNvrs, 10000);
</script>
</body></html>"""

REPORTS_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Reports — Kumarans</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
:root{--bg:#f3f6fb;--surface:#ffffff;--surface-soft:#f8fbff;--line:#dbe3ef;--line-strong:#c8d4e5;--text:#1f2a37;--muted:#6b7a90;--primary:#1f6feb;--primary-soft:#e8f1ff;--ok:#16a34a;--shadow:0 10px 30px rgba(15,23,42,.06)}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:system-ui,sans-serif;background:linear-gradient(180deg,#f8fafc 0%,var(--bg) 100%);color:var(--text);font-size:13px;min-height:100vh}
.topbar{padding:14px 20px;background:rgba(255,255,255,.96);border-bottom:1px solid var(--line);position:sticky;top:0;z-index:100;backdrop-filter:blur(12px)}
.top-row{display:grid;grid-template-columns:1fr auto 1fr;align-items:center;gap:14px}
.brand{display:flex;align-items:center;gap:12px;min-width:0}
.brand img{width:342px;max-width:100%;height:auto;object-fit:contain;filter:drop-shadow(0 2px 6px rgba(15,23,42,.10))}
.titlebar{justify-self:center;display:flex;flex-direction:column;align-items:center;gap:8px;text-align:center}
.title-main{font-size:20px;font-weight:700;line-height:1.1;color:var(--text)}
.site-pill{font-size:11px;font-weight:700;color:#2159b3;background:var(--primary-soft);border:1px solid #cfe0ff;border-radius:999px;padding:5px 12px;line-height:1}
.top-actions{display:flex;justify-content:flex-end;gap:8px;flex-wrap:wrap}
.nav-link,.btn,.chip{padding:8px 12px;border-radius:10px;font-size:12px;border:1px solid var(--line);background:#fff;color:#475569;text-decoration:none;font-weight:600;cursor:pointer}
.nav-link:hover,.btn:hover,.chip:hover{background:#f8fbff;border-color:var(--line-strong)}
.chip.active,.btn.primary{background:var(--primary);color:#fff;border-color:var(--primary)}
.subbar{display:flex;align-items:center;gap:8px;flex-wrap:wrap;margin-top:12px;padding-top:12px;border-top:1px solid #edf2f7}
.date-field{padding:9px 11px;border:1px solid var(--line);border-radius:10px;font-size:12px;background:#fff;color:var(--text);outline:none}
.date-field:focus{border-color:var(--primary);box-shadow:0 0 0 3px rgba(31,111,235,.10)}
.shell{padding:18px 20px 22px}
.hero{display:grid;grid-template-columns:1.5fr 1fr;gap:14px;margin-bottom:14px}
.panel{background:var(--surface);border:1px solid var(--line);border-radius:16px;box-shadow:var(--shadow)}
.panel.pad{padding:18px}
.hero h1{font-size:24px;line-height:1.1;margin-bottom:8px}
.hero p{font-size:13px;color:var(--muted);max-width:720px;line-height:1.6}
.range-note{margin-top:12px;font-size:12px;color:#4f6280}
.toolbar{display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.kpi-grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px;margin-bottom:14px}
.kpi{padding:16px;border-radius:16px;background:var(--surface);border:1px solid var(--line);box-shadow:var(--shadow)}
.kpi .k{font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.08em}
.kpi .v{font-size:28px;font-weight:800;margin-top:8px;color:var(--text)}
.kpi .s{font-size:12px;color:#70839b;margin-top:6px}
.grid{display:grid;grid-template-columns:1.3fr 1fr;gap:14px;margin-bottom:14px}
.stack{display:grid;gap:14px}
.chart-card{padding:16px}
.card-head{display:flex;justify-content:space-between;align-items:center;margin-bottom:14px;gap:12px}
.card-title{font-size:16px;font-weight:700;color:var(--text)}
.card-sub{font-size:12px;color:#70839b}
.chart-wrap{height:320px}
.small-chart{height:260px}
.table-wrap{overflow:auto}
table{width:100%;border-collapse:collapse}
th,td{padding:10px 12px;text-align:left;font-size:12px;border-bottom:1px solid #edf2f7}
th{color:#5d6d83;font-weight:700;text-transform:uppercase;font-size:11px;letter-spacing:.08em;background:#f8fafc}
td{color:var(--text)}
.rank{display:inline-flex;align-items:center;justify-content:center;width:24px;height:24px;border-radius:999px;background:var(--primary-soft);color:#2159b3;font-size:11px;font-weight:700}
.pill{display:inline-block;padding:4px 8px;border-radius:999px;background:#ebf8ef;color:#147a3f;font-size:11px}
.empty{padding:28px;color:#7a879a;text-align:center}
.footer-note{padding:0 18px 18px;color:#7a879a;font-size:11px;line-height:1.6;text-align:center}
@media (max-width: 1080px){.hero,.grid{grid-template-columns:1fr}.kpi-grid{grid-template-columns:repeat(2,minmax(0,1fr))}}
@media (max-width: 860px){.top-row{grid-template-columns:1fr;justify-items:start}.titlebar{justify-self:start;align-items:flex-start}.top-actions{justify-content:flex-start}.kpi-grid{grid-template-columns:1fr 1fr}}
@media (max-width: 640px){.shell,.topbar{padding-left:14px;padding-right:14px}.kpi-grid{grid-template-columns:1fr}.toolbar,.subbar,.top-actions{align-items:stretch}.btn,.nav-link,.chip,.date-field{width:100%}}
</style></head>
<body>
<div class="topbar">
  <div class="top-row">
    <div class="brand">
      <img src="https://kumarans.org/images/Sri%20Kumaran%20Childrens%20Home.png" alt="Sri Kumaran logo">
    </div>
    <div class="titlebar">
      <span class="title-main">Performance Reports</span>
      <strong class="site-pill">{{ site_name }}</strong>
    </div>
    <div class="top-actions">
      <a href="/" class="nav-link">Dashboard</a>
      <a href="/nvr-monitor" class="nav-link">NVR Monitor</a>
      <a href="/camera-logs" class="nav-link">Camera Logs</a>
      <a href="/audit" class="nav-link">Audit Log</a>
      <a href="/logout" class="nav-link">Sign out ({{ user }})</a>
    </div>
  </div>
  <div class="subbar">
    <button class="chip active" data-preset="7d" onclick="pickPreset('7d')">Last 7 Days</button>
    <button class="chip" data-preset="1m" onclick="pickPreset('1m')">1 Month</button>
    <button class="chip" data-preset="3m" onclick="pickPreset('3m')">3 Months</button>
    <input id="fromDate" class="date-field" type="date">
    <input id="toDate" class="date-field" type="date">
    <button class="btn" onclick="applyCustomRange()">Apply Custom Range</button>
  </div>
</div>

<div class="shell">
  <div class="hero">
    <div class="panel pad">
      <h1>Operational Reporting</h1>
      <p>Review camera availability, downtime impact, and outage patterns across the selected period with the same monitoring style used on the main console.</p>
      <div class="range-note" id="rangeNote">Loading selected range...</div>
    </div>
    <div class="panel pad">
      <div class="card-title">Reporting Window</div>
      <div class="card-sub" id="reportWindow" style="margin-top:8px">Preparing range...</div>
      <div class="card-sub" style="margin-top:14px">Use presets for quick review, or choose a custom date range for investigations and monthly reporting.</div>
    </div>
  </div>

  <div class="kpi-grid" id="kpis">
    <div class="kpi"><div class="k">Uptime</div><div class="v">—</div><div class="s">Selected range</div></div>
    <div class="kpi"><div class="k">Offline Events</div><div class="v">—</div><div class="s">Selected range</div></div>
    <div class="kpi"><div class="k">Downtime</div><div class="v">—</div><div class="s">Selected range</div></div>
    <div class="kpi"><div class="k">Current Fleet</div><div class="v">—</div><div class="s">Live status snapshot</div></div>
  </div>

  <div class="grid">
    <div class="panel chart-card">
      <div class="card-head">
        <div>
          <div class="card-title">Outage Trend</div>
          <div class="card-sub">Offline events and downtime across the selected range</div>
        </div>
      </div>
      <div class="chart-wrap"><canvas id="trendChart"></canvas></div>
    </div>
    <div class="stack">
      <div class="panel chart-card">
        <div class="card-head">
          <div>
            <div class="card-title">Current Status Split</div>
            <div class="card-sub">Live online, offline, and maintenance breakdown</div>
          </div>
        </div>
        <div class="small-chart"><canvas id="statusChart"></canvas></div>
      </div>
      <div class="panel chart-card">
        <div class="card-head">
          <div>
            <div class="card-title">Top 10 Worst Cameras</div>
            <div class="card-sub">Ordered by downtime and repeated outages</div>
          </div>
        </div>
        <div class="small-chart"><canvas id="worstChart"></canvas></div>
      </div>
    </div>
  </div>

  <div class="grid">
    <div class="panel chart-card">
      <div class="card-head">
        <div>
          <div class="card-title">Zone Health</div>
          <div class="card-sub">Compare downtime pressure across zones</div>
        </div>
      </div>
      <div class="small-chart"><canvas id="zoneChart"></canvas></div>
    </div>
    <div class="panel chart-card">
      <div class="card-head">
        <div>
          <div class="card-title">Worst Cameras Table</div>
          <div class="card-sub">Detailed ranking for the selected range</div>
        </div>
      </div>
      <div class="table-wrap"><table><thead><tr><th>#</th><th>Camera</th><th>Downtime</th><th>Events</th><th>Uptime</th></tr></thead><tbody id="worstTable"></tbody></table></div>
    </div>
  </div>

  <div class="grid">
    <div class="panel chart-card">
      <div class="card-head">
        <div>
          <div class="card-title">Zone Summary</div>
          <div class="card-sub">Downtime and outage distribution by zone</div>
        </div>
      </div>
      <div class="table-wrap"><table><thead><tr><th>Zone</th><th>Cameras</th><th>Downtime</th><th>Events</th><th>Uptime</th></tr></thead><tbody id="zoneTable"></tbody></table></div>
    </div>
    <div class="panel chart-card">
      <div class="card-head">
        <div>
          <div class="card-title">NVR Summary</div>
          <div class="card-sub">Find NVRs driving the highest downtime impact</div>
        </div>
      </div>
      <div class="table-wrap"><table><thead><tr><th>NVR</th><th>Cameras</th><th>Downtime</th><th>Events</th><th>Uptime</th></tr></thead><tbody id="nvrTable"></tbody></table></div>
    </div>
  </div>

  <div class="footer-note">© Sri Kumaran Childrens Home Educational Council. All rights reserved. Authorized operational use only. Activity on this monitoring system may be logged and reviewed.</div>
</div>

<script>
function escHtml(v){if(v==null)return '';return String(v).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&#x27;');}
let currentPreset='7d';
let trendChart,statusChart,worstChart,zoneChart;

function fmtDowntime(mins){
  const total=Math.max(0,parseInt(mins||0,10));
  const d=Math.floor(total/1440);
  const h=Math.floor((total%1440)/60);
  const m=total%60;
  const out=[];
  if(d)out.push(`${d}d`);
  if(h)out.push(`${h}h`);
  if(m || !out.length)out.push(`${m}m`);
  return out.join(' ');
}
function fmtPct(v){ return `${Math.round((parseFloat(v)||0)*10)/10}%`; }
function syncPresetButtons(active){
  document.querySelectorAll('.chip[data-preset]').forEach(btn=>btn.classList.toggle('active', btn.dataset.preset===active));
}
function pickPreset(preset){
  currentPreset=preset;
  syncPresetButtons(preset);
  loadReports();
}
function applyCustomRange(){
  currentPreset='custom';
  syncPresetButtons('');
  loadReports();
}
function destroyChart(instance){ if(instance)instance.destroy(); }
function setTableRows(id, rows, render){
  const el=document.getElementById(id);
  if(!rows.length){ el.innerHTML=`<tr><td colspan="5" class="empty">No data in this range.</td></tr>`; return; }
  el.innerHTML=rows.map(render).join('');
}
const chartText='#506279';
const gridLine='rgba(148,163,184,.15)';
async function loadReports(){
  const p=new URLSearchParams();
  p.set('preset', currentPreset);
  if(currentPreset==='custom'){
    p.set('from', document.getElementById('fromDate').value);
    p.set('to', document.getElementById('toDate').value);
  }
  const r=await fetch('/api/reports?'+p.toString());
  const d=await r.json();
  document.getElementById('rangeNote').textContent=`Showing ${d.date_from} to ${d.date_to}`;
  document.getElementById('reportWindow').textContent=`${d.date_from} to ${d.date_to}`;
  if(currentPreset!=='custom'){
    document.getElementById('fromDate').value=d.date_from;
    document.getElementById('toDate').value=d.date_to;
  }

  const kpis=document.querySelectorAll('#kpis .kpi');
  kpis[0].querySelector('.v').textContent=fmtPct(d.overview.uptime_pct);
  kpis[1].querySelector('.v').textContent=d.overview.offline_events || 0;
  kpis[2].querySelector('.v').textContent=fmtDowntime(d.overview.downtime_min);
  kpis[3].querySelector('.v').textContent=`${d.status.online}/${d.status.total}`;
  kpis[3].querySelector('.s').textContent=`Offline ${d.status.offline} • Maintenance ${d.status.maintenance}`;

  destroyChart(trendChart);
  trendChart=new Chart(document.getElementById('trendChart'),{
    type:'bar',
    data:{
      labels:d.daily_trend.map(x=>x.date),
      datasets:[
        {type:'bar',label:'Offline Events',data:d.daily_trend.map(x=>x.offline_events||0),backgroundColor:'rgba(31,111,235,.45)',borderRadius:8,maxBarThickness:28},
        {type:'line',label:'Downtime (min)',data:d.daily_trend.map(x=>x.downtime_min||0),borderColor:'#16a34a',backgroundColor:'rgba(22,163,74,.08)',tension:.35,fill:true,yAxisID:'y1'}
      ]
    },
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{labels:{color:chartText}}},scales:{x:{ticks:{color:chartText},grid:{color:gridLine}},y:{ticks:{color:chartText},grid:{color:gridLine}},y1:{position:'right',ticks:{color:'#147a3f'},grid:{drawOnChartArea:false}}}}
  });

  destroyChart(statusChart);
  statusChart=new Chart(document.getElementById('statusChart'),{
    type:'doughnut',
    data:{labels:['Online','Offline','Maintenance'],datasets:[{data:[d.status.online||0,d.status.offline||0,d.status.maintenance||0],backgroundColor:['#22c55e','#ef4444','#f59e0b'],borderWidth:0}]},
    options:{responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'bottom',labels:{color:chartText}}},cutout:'68%'}
  });

  destroyChart(worstChart);
  worstChart=new Chart(document.getElementById('worstChart'),{
    type:'bar',
    data:{labels:d.worst_cameras.map(x=>x.name||x.ip),datasets:[{label:'Downtime (min)',data:d.worst_cameras.map(x=>x.total_downtime_min||0),backgroundColor:'rgba(239,68,68,.55)',borderRadius:10}]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{x:{ticks:{color:chartText},grid:{color:gridLine}},y:{ticks:{color:chartText},grid:{display:false}}}}
  });

  destroyChart(zoneChart);
  zoneChart=new Chart(document.getElementById('zoneChart'),{
    type:'bar',
    data:{labels:d.zone_summary.slice(0,8).map(x=>x.zone),datasets:[{label:'Downtime (min)',data:d.zone_summary.slice(0,8).map(x=>x.downtime_min||0),backgroundColor:'rgba(14,165,233,.55)',borderRadius:10}]},
    options:{indexAxis:'y',responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{x:{ticks:{color:chartText},grid:{color:gridLine}},y:{ticks:{color:chartText},grid:{display:false}}}}
  });

  setTableRows('worstTable', d.worst_cameras, (row, idx)=>`<tr><td><span class="rank">${idx+1}</span></td><td><div>${escHtml(row.name)||'—'}</div><div style="color:#70839b;font-size:11px">${escHtml(row.ip)||''}</div></td><td>${fmtDowntime(row.total_downtime_min)}</td><td>${row.offline_events||0}</td><td><span class="pill">${fmtPct(row.uptime_pct||0)}</span></td></tr>`);
  setTableRows('zoneTable', d.zone_summary, row=>`<tr><td>${escHtml(row.zone)}</td><td>${row.camera_count||0}</td><td>${fmtDowntime(row.downtime_min)}</td><td>${row.offline_events||0}</td><td>${fmtPct(row.uptime_pct||0)}</td></tr>`);
  setTableRows('nvrTable', d.nvr_summary, row=>`<tr><td>${escHtml(row.nvr_name)}</td><td>${row.camera_count||0}</td><td>${fmtDowntime(row.downtime_min)}</td><td>${row.offline_events||0}</td><td>${fmtPct(row.uptime_pct||0)}</td></tr>`);
}
loadReports();
</script>
</body></html>"""

AUDIT_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Audit Log — Kumarans</title>
<style>
:root{--bg:#f3f6fb;--surface:#ffffff;--line:#dbe3ef;--line-strong:#c8d4e5;--text:#1f2a37;--muted:#6b7a90;--primary:#1f6feb;--primary-soft:#e8f1ff;--shadow:0 10px 30px rgba(15,23,42,.06)}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:system-ui,sans-serif;background:linear-gradient(180deg,#f8fafc 0%,var(--bg) 100%);font-size:13px;color:var(--text);min-height:100vh;display:flex;flex-direction:column}
.topbar{padding:14px 20px;background:rgba(255,255,255,.96);border-bottom:1px solid var(--line);position:sticky;top:0;z-index:100;backdrop-filter:blur(12px)}
.top-row{display:grid;grid-template-columns:1fr auto 1fr;align-items:center;gap:14px}
.brand img{width:342px;max-width:100%;height:auto;object-fit:contain;filter:drop-shadow(0 2px 6px rgba(15,23,42,.10))}
.titlebar{justify-self:center;display:flex;flex-direction:column;align-items:center;gap:8px;text-align:center}
.title-main{font-size:20px;font-weight:700;line-height:1.1;color:var(--text)}
.site-pill{font-size:11px;font-weight:700;color:#2159b3;background:var(--primary-soft);border:1px solid #cfe0ff;border-radius:999px;padding:5px 12px;line-height:1}
.top-actions{display:flex;justify-content:flex-end;gap:8px;flex-wrap:wrap}
.nav-link,.btn,.chip{padding:8px 12px;border-radius:10px;font-size:12px;border:1px solid var(--line);background:#fff;color:#475569;text-decoration:none;font-weight:600;cursor:pointer}
.nav-link:hover,.btn:hover,.chip:hover{background:#f8fbff;border-color:var(--line-strong)}
.chip.active{background:var(--primary);color:#fff;border-color:var(--primary)}
.page{flex:1;display:flex;flex-direction:column}
.stats-row{display:grid;grid-template-columns:repeat(5,1fr);gap:12px;padding:16px 20px 0}
.scard{background:var(--surface);border-radius:14px;padding:14px 16px;border:1px solid var(--line);box-shadow:var(--shadow)}
.slbl{font-size:11px;color:var(--muted);margin-bottom:4px}.sval{font-size:24px;font-weight:700}
.fbar{display:flex;gap:8px;align-items:center;padding:14px 20px;background:transparent;flex-wrap:wrap}
.srch{padding:9px 12px;border:1px solid var(--line);border-radius:10px;font-size:12px;outline:none;background:#fff;color:var(--text)}
.srch:focus{border-color:var(--primary);box-shadow:0 0 0 3px rgba(31,111,235,.10)}
.panel{margin:0 20px;background:var(--surface);border-radius:16px;border:1px solid var(--line);box-shadow:var(--shadow);overflow:hidden}
.table-wrap{overflow:auto}
table{width:100%;border-collapse:collapse;background:#fff;font-size:12px}
thead th{padding:10px 12px;text-align:left;background:#f8fafc;font-weight:700;color:#5d6d83;border-bottom:1px solid #edf2f7;white-space:nowrap}
tbody tr{border-bottom:1px solid #f1f5f9}tbody tr:hover{background:#f8fbff}td{padding:9px 12px}
.badge{display:inline-block;padding:3px 8px;border-radius:999px;font-size:10px;font-weight:700}
.bbar{display:flex;align-items:center;gap:8px;padding:12px 20px 18px;background:transparent;flex-wrap:wrap}
.page-info{font-size:11px;color:#70839b}
.footer-note{padding:0 18px 18px;color:#7a879a;font-size:11px;line-height:1.6;text-align:center;margin-top:auto}
@media (max-width: 980px){.top-row{grid-template-columns:1fr;justify-items:start}.titlebar{justify-self:start;align-items:flex-start}.top-actions{justify-content:flex-start}.stats-row{grid-template-columns:repeat(2,1fr)}}
@media (max-width: 640px){.topbar,.stats-row,.fbar,.bbar{padding-left:14px;padding-right:14px}.stats-row{grid-template-columns:1fr}.srch,.btn,.nav-link,.chip{width:100%}.panel{margin:0 14px}}
</style></head><body>
<div class="topbar">
  <div class="top-row">
    <div class="brand"><img src="https://kumarans.org/images/Sri%20Kumaran%20Childrens%20Home.png" alt="Sri Kumaran logo"></div>
    <div class="titlebar">
      <div class="title-main">Audit Log</div>
      <strong class="site-pill">{{ site_name }}</strong>
    </div>
    <div class="top-actions">
      <a href="/" class="nav-link">Dashboard</a>
      <a href="/nvr-monitor" class="nav-link">NVR Monitor</a>
      <a href="/camera-logs" class="nav-link">Camera Logs</a>
      <a href="/logout" class="nav-link">Sign out ({{ user }})</a>
    </div>
  </div>
</div>

<div class="page">
  <div class="stats-row">
    <div class="scard"><div class="slbl">Matched Records</div><div class="sval" id="sTot">—</div></div>
    <div class="scard"><div class="slbl">Logins On Page</div><div class="sval" id="sLg">—</div></div>
    <div class="scard"><div class="slbl">Maintenance On Page</div><div class="sval" id="sMt">—</div></div>
    <div class="scard"><div class="slbl">Bulk Changes On Page</div><div class="sval" id="sBk">—</div></div>
    <div class="scard"><div class="slbl">Exports On Page</div><div class="sval" id="sEx">—</div></div>
  </div>

  <div class="fbar">
    <input class="srch" type="search" id="sq" placeholder="Search user, camera, action..." oninput="debouncedLoad()">
    <select class="srch" id="auditType" style="width:150px" onchange="goFirstPage()">
      <option value="all">All Types</option>
      <option value="login">Login</option>
      <option value="logout">Logout</option>
      <option value="maintenance">Maintenance</option>
      <option value="bulk">Bulk Change</option>
      <option value="alert">Alert</option>
      <option value="export">Export</option>
      <option value="config">Config</option>
      <option value="monitor">Monitor</option>
    </select>
    <select class="srch" id="auditResult" style="width:140px" onchange="goFirstPage()">
      <option value="all">All Results</option>
      <option value="success">Success</option>
      <option value="failed">Failed</option>
    </select>
    <select class="srch" id="auditPageSize" style="width:120px" onchange="setAuditPageSize(this.value)">
      <option value="25" selected>25 / page</option>
      <option value="50">50 / page</option>
      <option value="100">100 / page</option>
      <option value="200">200 / page</option>
    </select>
    {% if role == 'admin' %}
    <select class="srch" id="auditUser" style="width:180px" onchange="goFirstPage()">
      <option value="all">All Users</option>
    </select>
    {% endif %}
    <button class="chip active" data-preset="7d" onclick="pickPreset('7d')">7 Days</button>
    <button class="chip" data-preset="1m" onclick="pickPreset('1m')">1 Month</button>
    <button class="chip" data-preset="3m" onclick="pickPreset('3m')">3 Months</button>
    <input class="srch" id="fromDate" style="width:150px" type="date">
    <input class="srch" id="toDate" style="width:150px" type="date">
    <button class="btn" onclick="applyCustomRange()">Custom Range</button>
    <button class="btn" onclick="clearAuditFilters()">Clear Filters</button>
    <div style="flex:1"></div>
    <span id="rcnt" class="page-info"></span>
  </div>

  <div class="panel">
    <div class="table-wrap">
      <table><thead><tr><th>Timestamp</th><th>User</th><th>Type</th><th>Description</th><th>Target</th><th>IP Address</th><th>Result</th></tr></thead>
      <tbody id="tb"></tbody></table>
    </div>
  </div>

  <div class="bbar">
    <span class="page-info" id="pgInfo"></span>
    <div style="flex:1"></div>
    <button class="btn" id="auditPrevBtn" onclick="changePage(-1)">Previous</button>
    <div id="auditPageNums" style="display:flex;align-items:center;gap:6px;flex-wrap:wrap"></div>
    <button class="btn" id="auditNextBtn" onclick="changePage(1)">Next</button>
  </div>

  <div class="footer-note">© Sri Kumaran Childrens Home Educational Council. All rights reserved. Authorized operational use only. Activity on this monitoring system may be logged and reviewed.</div>
</div>

<script>
function escHtml(v){if(v==null)return '';return String(v).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&#x27;');}
const COLORS={login:{bg:'#ebf5fb',c:'#1a5276'},logout:{bg:'#f5f5f5',c:'#666'},maintenance:{bg:'#fff4e8',c:'#9c640c'},bulk:{bg:'#f0ebfe',c:'#5b2c8f'},alert:{bg:'#fdecea',c:'#922b21'},export:{bg:'#eafaf1',c:'#1e8449'},config:{bg:'#fef5e7',c:'#9c640c'},monitor:{bg:'#ebf3ff',c:'#2457a6'}};
let currentPreset='7d';
let currentPage=1;
let totalPages=1;
let pageSize=25;
let searchTimer=null;
function fmtDateTime(v){
  if(!v)return '—';
  const d=new Date(v);
  if(Number.isNaN(d.getTime()))return v;
  const dd=String(d.getDate()).padStart(2,'0');
  const mm=String(d.getMonth()+1).padStart(2,'0');
  const yyyy=d.getFullYear();
  let hh=d.getHours();
  const min=String(d.getMinutes()).padStart(2,'0');
  const ap=hh>=12?'PM':'AM';
  hh=hh%12||12;
  return `${dd}-${mm}-${yyyy} ${String(hh).padStart(2,'0')}:${min} ${ap}`;
}
function syncPresetButtons(active){
  document.querySelectorAll('.chip[data-preset]').forEach(btn=>btn.classList.toggle('active', btn.dataset.preset===active));
}
function pickPreset(preset){
  currentPreset=preset;
  syncPresetButtons(preset);
  currentPage=1;
  load();
}
function applyCustomRange(){
  currentPreset='custom';
  syncPresetButtons('');
  currentPage=1;
  load();
}
function clearAuditFilters(){
  document.getElementById('sq').value='';
  document.getElementById('auditType').value='all';
  document.getElementById('auditResult').value='all';
  document.getElementById('fromDate').value='';
  document.getElementById('toDate').value='';
  const userSel=document.getElementById('auditUser');
  if(userSel) userSel.value='all';
  pickPreset('7d');
}
function setAuditPageSize(value){
  pageSize=Math.max(25, Math.min(200, parseInt(value || '25', 10) || 25));
  currentPage=1;
  load();
}
function changePage(delta){
  const next=currentPage+delta;
  if(next<1 || next>totalPages) return;
  currentPage=next;
  load();
}
function setPage(page){
  if(page<1 || page>totalPages || page===currentPage) return;
  currentPage=page;
  load();
}
function buildPageTokens(current, total){
  if(total<=7){
    const out=[];
    for(let i=1;i<=total;i++) out.push(i);
    return out;
  }
  if(current<=4) return [1,2,3,4,5,'...',total];
  if(current>=total-3) return [1,'...',total-4,total-3,total-2,total-1,total];
  return [1,'...',current-1,current,current+1,'...',total];
}
function renderAuditPageNumbers(){
  const holder=document.getElementById('auditPageNums');
  if(!holder) return;
  const tokens=buildPageTokens(currentPage,totalPages);
  holder.innerHTML=tokens.map(tok=>{
    if(tok==='...') return `<span style="font-size:11px;color:#94a3b8;padding:0 2px">...</span>`;
    const active=tok===currentPage;
    return `<button class="btn ${active?'active':''}" ${active?'disabled':''} onclick="setPage(${tok})" style="${active?'background:#1f6feb;color:#fff;border-color:#1f6feb;':''}">${tok}</button>`;
  }).join('');
}
function goFirstPage(){ currentPage=1; load(); }
function debouncedLoad(){
  window.clearTimeout(searchTimer);
  searchTimer=window.setTimeout(goFirstPage, 250);
}
async function loadAuditUsers(){
  const userSel=document.getElementById('auditUser');
  if(!userSel) return;
  const r=await fetch('/api/audit-users');
  if(!r.ok) return;
  const users=await r.json();
  userSel.innerHTML='<option value="all">All Users</option>'+users.map(u=>`<option value="${String(u||'').replace(/"/g,'&quot;')}">${u}</option>`).join('');
}
async function load(){
  const q=document.getElementById('sq').value;
  const type=document.getElementById('auditType').value;
  const result=document.getElementById('auditResult').value;
  const userSel=document.getElementById('auditUser');
  const auditUser=userSel?userSel.value:'all';
  const p=new URLSearchParams();
  p.set('preset', currentPreset);
  p.set('page', String(currentPage));
  p.set('page_size', String(pageSize));
  if(currentPreset==='custom'){
    p.set('from', document.getElementById('fromDate').value);
    p.set('to', document.getElementById('toDate').value);
  }
  if(type&&type!=='all')p.set('type',type);
  if(result&&result!=='all')p.set('result',result);
  if(auditUser&&auditUser!=='all')p.set('user',auditUser);
  if(q)p.set('q',q);
  const r=await fetch('/api/audit?'+p.toString());
  const data=await r.json();
  const rows=data.items||[];
  totalPages=Math.max(1, Math.ceil((data.total||0)/pageSize));
  currentPage=Math.min(currentPage, totalPages);
  if(data.date_from) document.getElementById('fromDate').value=data.date_from;
  if(data.date_to) document.getElementById('toDate').value=data.date_to;
  document.getElementById('sTot').textContent=data.total || 0;
  document.getElementById('sLg').textContent=rows.filter(x=>x.event_type==='login').length;
  document.getElementById('sMt').textContent=rows.filter(x=>x.event_type==='maintenance').length;
  document.getElementById('sBk').textContent=rows.filter(x=>x.event_type==='bulk').length;
  document.getElementById('sEx').textContent=rows.filter(x=>x.event_type==='export').length;
  document.getElementById('rcnt').textContent=`${data.date_from} to ${data.date_to}`;
  document.getElementById('pgInfo').textContent=`Page ${currentPage} of ${totalPages} • ${rows.length} shown • ${data.total || 0} matched`;
  document.getElementById('auditPrevBtn').disabled=currentPage<=1;
  document.getElementById('auditNextBtn').disabled=currentPage>=totalPages;
  renderAuditPageNumbers();
  document.getElementById('tb').innerHTML=rows.length ? rows.map(row=>{
    const cl=COLORS[row.event_type]||{bg:'#f5f5f5',c:'#666'};
    const ok=row.result==='success';
    return `<tr>
      <td style="font-family:monospace;font-size:11px;color:#70839b">${fmtDateTime(row.ts)}</td>
      <td style="font-weight:600">${escHtml(row.user||'')}</td>
      <td><span class="badge" style="background:${cl.bg};color:${cl.c}">${escHtml(row.event_type||'')}</span></td>
      <td>${escHtml(row.description||'')}</td>
      <td style="font-family:monospace;font-size:11px;color:#70839b">${escHtml(row.target||'')}</td>
      <td style="font-family:monospace;font-size:11px;color:#94a3b8">${escHtml(row.ip_address||'')}</td>
      <td><span class="badge" style="background:${ok?'#eafaf1':'#fdecea'};color:${ok?'#1e8449':'#922b21'}">${escHtml(row.result||'')}</span></td>
    </tr>`;
  }).join('') : '<tr><td colspan="7" style="padding:28px;text-align:center;color:#7a879a">No audit rows found for this range.</td></tr>';
}
loadAuditUsers().then(load);
</script>
</body></html>"""

CAMERA_LOG_HTML = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Camera Logs — Kumarans</title>
<style>
:root{--bg:#f3f6fb;--surface:#ffffff;--line:#dbe3ef;--line-strong:#c8d4e5;--text:#1f2a37;--muted:#6b7a90;--primary:#1f6feb;--primary-soft:#e8f1ff;--shadow:0 10px 30px rgba(15,23,42,.06)}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:system-ui,sans-serif;background:linear-gradient(180deg,#f8fafc 0%,var(--bg) 100%);font-size:13px;color:var(--text);min-height:100vh;display:flex;flex-direction:column}
.topbar{padding:14px 20px;background:rgba(255,255,255,.96);border-bottom:1px solid var(--line);position:sticky;top:0;z-index:100;backdrop-filter:blur(12px)}
.top-row{display:grid;grid-template-columns:1fr auto 1fr;align-items:center;gap:14px}
.brand img{width:342px;max-width:100%;height:auto;object-fit:contain;filter:drop-shadow(0 2px 6px rgba(15,23,42,.10))}
.titlebar{justify-self:center;display:flex;flex-direction:column;align-items:center;gap:8px;text-align:center}
.title-main{font-size:20px;font-weight:700;line-height:1.1;color:var(--text)}
.site-pill{font-size:11px;font-weight:700;color:#2159b3;background:var(--primary-soft);border:1px solid #cfe0ff;border-radius:999px;padding:5px 12px;line-height:1}
.top-actions{display:flex;justify-content:flex-end;gap:8px;flex-wrap:wrap}
.nav-link,.btn,.chip{padding:8px 12px;border-radius:10px;font-size:12px;border:1px solid var(--line);background:#fff;color:#475569;text-decoration:none;font-weight:600;cursor:pointer}
.nav-link:hover,.btn:hover,.chip:hover{background:#f8fbff;border-color:var(--line-strong)}
.chip.active{background:var(--primary);color:#fff;border-color:var(--primary)}
.page{flex:1;display:flex;flex-direction:column}
.stats-row{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;padding:16px 20px 0}
.scard{background:var(--surface);border-radius:14px;padding:14px 16px;border:1px solid var(--line);box-shadow:var(--shadow)}
.slbl{font-size:11px;color:var(--muted);margin-bottom:4px}.sval{font-size:24px;font-weight:700}
.fbar{display:flex;gap:8px;align-items:center;padding:14px 20px;background:transparent;flex-wrap:wrap}
.srch{padding:9px 12px;border:1px solid var(--line);border-radius:10px;font-size:12px;outline:none;background:#fff;color:var(--text)}
.srch:focus{border-color:var(--primary);box-shadow:0 0 0 3px rgba(31,111,235,.10)}
.panel{margin:0 20px;background:var(--surface);border-radius:16px;border:1px solid var(--line);box-shadow:var(--shadow);overflow:hidden}
.table-wrap{overflow:auto}
table{width:100%;border-collapse:collapse;background:#fff;font-size:12px}
thead th{padding:10px 12px;text-align:left;background:#f8fafc;font-weight:700;color:#5d6d83;border-bottom:1px solid #edf2f7;white-space:nowrap}
tbody tr{border-bottom:1px solid #f1f5f9}tbody tr:hover{background:#f8fbff}td{padding:9px 12px}
.badge{display:inline-block;padding:3px 8px;border-radius:999px;font-size:10px;font-weight:700}
.bbar{display:flex;align-items:center;gap:8px;padding:12px 20px 18px;background:transparent;flex-wrap:wrap}
.page-info{font-size:11px;color:#70839b}
.overlay{display:none;position:fixed;inset:0;background:rgba(15,23,42,.45);z-index:200;align-items:flex-start;justify-content:center;padding:40px 12px;overflow-y:auto}
.overlay.show{display:flex}
.modal{background:#fff;border-radius:16px;width:560px;max-width:100%;border:1px solid var(--line);box-shadow:0 24px 60px rgba(15,23,42,.16)}
.mhdr{display:flex;align-items:flex-start;justify-content:space-between;padding:16px 20px;border-bottom:1px solid #edf2f7}
.mbody{padding:16px 20px}
.pvbox{width:100%;height:auto;min-height:0;background:transparent;border:none;display:block;margin-bottom:16px;overflow:visible;position:relative}
.pvbox img{display:block;width:100%;height:auto;object-fit:contain}
.irow{display:flex;font-size:12px;padding:7px 0;border-bottom:1px solid #f1f5f9}
.ilbl{width:120px;color:#70839b;flex-shrink:0}
.footer-note{padding:0 18px 18px;color:#7a879a;font-size:11px;line-height:1.6;text-align:center;margin-top:auto}
@media (max-width: 980px){.top-row{grid-template-columns:1fr;justify-items:start}.titlebar{justify-self:start;align-items:flex-start}.top-actions{justify-content:flex-start}.stats-row{grid-template-columns:repeat(2,1fr)}}
@media (max-width: 640px){.topbar,.stats-row,.fbar,.bbar{padding-left:14px;padding-right:14px}.stats-row{grid-template-columns:1fr}.srch,.btn,.nav-link,.chip{width:100%}.panel{margin:0 14px}}
</style></head><body>
<div class="topbar">
  <div class="top-row">
    <div class="brand"><img src="https://kumarans.org/images/Sri%20Kumaran%20Childrens%20Home.png" alt="Sri Kumaran logo"></div>
    <div class="titlebar">
      <div class="title-main">Camera Logs</div>
      <strong class="site-pill">{{ site_name }}</strong>
    </div>
    <div class="top-actions">
      <a href="/" class="nav-link">Dashboard</a>
      <a href="/nvr-monitor" class="nav-link">NVR Monitor</a>
      <a href="/audit" class="nav-link">Audit Log</a>
      <a href="/logout" class="nav-link">Sign out ({{ user }})</a>
    </div>
  </div>
</div>

<div class="page">
  <div class="stats-row">
    <div class="scard"><div class="slbl">Matched Events</div><div class="sval" id="sTot">—</div></div>
    <div class="scard"><div class="slbl">Offline On Page</div><div class="sval" style="color:#c0392b" id="sOff">—</div></div>
    <div class="scard"><div class="slbl">Recovered On Page</div><div class="sval" style="color:#27ae60" id="sOn">—</div></div>
    <div class="scard"><div class="slbl">Shown On Page</div><div class="sval" id="sShown">—</div></div>
  </div>

  <div class="fbar">
    <input class="srch" type="search" id="lq" placeholder="Search camera, IP, zone, NVR..." oninput="debouncedLoadLogs()">
    <select class="srch" id="eventType" style="width:140px" onchange="goFirstLogPage()">
      <option value="all">All Events</option>
      <option value="offline">Offline</option>
      <option value="online">Online</option>
    </select>
    <select class="srch" id="logZone" style="width:150px" onchange="goFirstLogPage()">
      <option value="">All Zones</option>
      {% for z in zones %}<option value="{{ z }}">{{ z }}</option>{% endfor %}
    </select>
    <select class="srch" id="logNvr" style="width:150px" onchange="goFirstLogPage()">
      <option value="">All NVRs</option>
      {% for n in nvrs %}<option value="{{ n }}">{{ n }}</option>{% endfor %}
    </select>
    <select class="srch" id="logPageSize" style="width:120px" onchange="setLogPageSize(this.value)">
      <option value="25" selected>25 / page</option>
      <option value="50">50 / page</option>
      <option value="100">100 / page</option>
      <option value="200">200 / page</option>
    </select>
    <button class="chip active" data-preset="7d" onclick="pickLogPreset('7d')">7 Days</button>
    <button class="chip" data-preset="1m" onclick="pickLogPreset('1m')">1 Month</button>
    <button class="chip" data-preset="3m" onclick="pickLogPreset('3m')">3 Months</button>
    <input class="srch" id="fromDate" style="width:150px" type="date">
    <input class="srch" id="toDate" style="width:150px" type="date">
    <button class="btn" onclick="applyLogCustomRange()">Custom Range</button>
    <button class="btn" onclick="clearCameraLogFilters()">Clear Filters</button>
    <div style="flex:1"></div>
    <span id="rcnt" class="page-info"></span>
  </div>

  <div class="panel">
    <div class="table-wrap">
      <table><thead><tr><th>Timestamp</th><th>Camera</th><th>IP</th><th>Zone</th><th>NVR</th><th>Event</th><th>Duration</th></tr></thead>
      <tbody id="tb"></tbody></table>
    </div>
  </div>

  <div class="bbar">
    <span class="page-info" id="pgInfo">Tracks camera offline and recovery events.</span>
    <div style="flex:1"></div>
    <button class="btn" id="logPrevBtn" onclick="changeLogPage(-1)">Previous</button>
    <div id="logPageNums" style="display:flex;align-items:center;gap:6px;flex-wrap:wrap"></div>
    <button class="btn" id="logNextBtn" onclick="changeLogPage(1)">Next</button>
    <button class="btn" onclick="window.location='/export/offline/excel'">Offline Excel</button>
  </div>

  <div class="footer-note">© Sri Kumaran Childrens Home Educational Council. All rights reserved. Authorized operational use only. Activity on this monitoring system may be logged and reviewed.</div>
</div>

<div class="overlay" id="logOv" onclick="if(event.target===this)closeLogModal()">
  <div class="modal">
    <div class="mhdr">
      <div>
        <div style="font-size:16px;font-weight:600;display:flex;align-items:center;gap:8px"><span id="logNmDot" style="width:10px;height:10px;border-radius:50%;display:inline-block;flex:0 0 10px;background:#bbb"></span><span id="logNm"></span></div>
        <div style="font-size:12px;color:#70839b;margin-top:3px" id="logSb"></div>
      </div>
      <div style="display:flex;gap:6px">
        <button class="btn" onclick="closeLogModal()">Close</button>
      </div>
    </div>
    <div class="mbody">
      <div class="pvbox" id="logPvb"><div style="text-align:center;color:#94a3b8">Loading preview...</div></div>
      <div id="logRws"></div>
      <div style="border-top:1px solid #edf2f7;padding-top:12px;margin-top:12px">
        <div style="font-size:11px;font-weight:700;color:#94a3b8;margin-bottom:8px">RECENT EVENTS</div>
        <div id="logEv"></div>
      </div>
    </div>
  </div>
</div>
<script>
function escHtml(v){if(v==null)return '';return String(v).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;').replace(/'/g,'&#x27;');}
let currentLogPreset='7d';
let currentLogPage=1;
let totalLogPages=1;
let logPageSize=25;
let logSearchTimer=null;
let logPreviewTimer=null;
function copyTextValue(value){
  if(!value) return;
  if(navigator.clipboard && navigator.clipboard.writeText){
    navigator.clipboard.writeText(value);
  }
}
function resolvePlayerStream(streamUrls){
  if(!streamUrls) return '';
  const host=window.location.hostname;
  if(host==='127.0.0.1'||host==='localhost'){
    return streamUrls.player_local || streamUrls.player || '';
  }
  return streamUrls.player || streamUrls.player_local || '';
}
async function toggleFullscreen(elementId){
  const el=document.getElementById(elementId);
  if(!el) return;
  const img=el.querySelector('img');
  if(document.fullscreenElement){
    await document.exitFullscreen();
    if(img){
      img.style.width='100%';
      img.style.height='auto';
      img.style.maxHeight='65vh';
      img.style.objectFit='contain';
    }
    return;
  }
  if(el.requestFullscreen){
    await el.requestFullscreen();
    if(img){
      img.style.width='100vw';
      img.style.height='100vh';
      img.style.maxHeight='100vh';
      img.style.objectFit='contain';
    }
  }
}
function stopLogPreview(){
  if(logPreviewTimer){
    window.clearInterval(logPreviewTimer);
    logPreviewTimer=null;
  }
}
function startLogPreview(snapshotUrl){
  stopLogPreview();
  if(!snapshotUrl) return;
  const img=document.getElementById('logPreviewImg');
  if(!img) return;
  const refresh=()=>{ img.src=`${snapshotUrl}${snapshotUrl.includes('?')?'&':'?'}t=${Date.now()}`; };
  refresh();
  logPreviewTimer=window.setInterval(refresh, 500);
}
function fmtDateTime(v){
  if(!v)return '—';
  const d=new Date(v);
  if(Number.isNaN(d.getTime()))return v;
  const dd=String(d.getDate()).padStart(2,'0');
  const mm=String(d.getMonth()+1).padStart(2,'0');
  const yyyy=d.getFullYear();
  let hh=d.getHours();
  const min=String(d.getMinutes()).padStart(2,'0');
  const ap=hh>=12?'PM':'AM';
  hh=hh%12||12;
  return `${dd}-${mm}-${yyyy} ${String(hh).padStart(2,'0')}:${min} ${ap}`;
}
function fmtDuration(s){
  if(!s)return '—';
  const total=Math.max(0, parseInt(s,10) || 0);
  const d=Math.floor(total/86400);
  const h=Math.floor((total%86400)/3600);
  const m=Math.floor((total%3600)/60);
  const sec=total%60;
  const parts=[];
  if(d)parts.push(`${d}d`);
  if(h)parts.push(`${h}h`);
  if(m)parts.push(`${m}m`);
  if(sec)parts.push(`${sec}s`);
  return parts.join(' ') || '0s';
}
function statusHtml(c){
  if(c.maintenance) return '<span style="color:#e67e22;font-weight:500">Maintenance</span>';
  if(c.online) return '<span style="color:#27ae60;font-weight:500">Online</span>';
  return '<span style="color:#c0392b;font-weight:500">Offline</span>';
}
function syncLogPresetButtons(active){
  document.querySelectorAll('.chip[data-preset]').forEach(btn=>btn.classList.toggle('active', btn.dataset.preset===active));
}
function pickLogPreset(preset){
  currentLogPreset=preset;
  syncLogPresetButtons(preset);
  currentLogPage=1;
  loadLogs();
}
function applyLogCustomRange(){
  currentLogPreset='custom';
  syncLogPresetButtons('');
  currentLogPage=1;
  loadLogs();
}
function clearCameraLogFilters(){
  document.getElementById('lq').value='';
  document.getElementById('eventType').value='all';
  document.getElementById('logZone').value='';
  document.getElementById('logNvr').value='';
  document.getElementById('fromDate').value='';
  document.getElementById('toDate').value='';
  pickLogPreset('7d');
}
function setLogPageSize(value){
  logPageSize=Math.max(25, Math.min(200, parseInt(value || '25', 10) || 25));
  currentLogPage=1;
  loadLogs();
}
function changeLogPage(delta){
  const next=currentLogPage+delta;
  if(next<1 || next>totalLogPages) return;
  currentLogPage=next;
  loadLogs();
}
function setLogPage(page){
  if(page<1 || page>totalLogPages || page===currentLogPage) return;
  currentLogPage=page;
  loadLogs();
}
function buildPageTokens(current, total){
  if(total<=7){
    const out=[];
    for(let i=1;i<=total;i++) out.push(i);
    return out;
  }
  if(current<=4) return [1,2,3,4,5,'...',total];
  if(current>=total-3) return [1,'...',total-4,total-3,total-2,total-1,total];
  return [1,'...',current-1,current,current+1,'...',total];
}
function renderLogPageNumbers(){
  const holder=document.getElementById('logPageNums');
  if(!holder) return;
  const tokens=buildPageTokens(currentLogPage,totalLogPages);
  holder.innerHTML=tokens.map(tok=>{
    if(tok==='...') return `<span style="font-size:11px;color:#94a3b8;padding:0 2px">...</span>`;
    const active=tok===currentLogPage;
    return `<button class="btn ${active?'active':''}" ${active?'disabled':''} onclick="setLogPage(${tok})" style="${active?'background:#1f6feb;color:#fff;border-color:#1f6feb;':''}">${tok}</button>`;
  }).join('');
}
function goFirstLogPage(){ currentLogPage=1; loadLogs(); }
function debouncedLoadLogs(){
  window.clearTimeout(logSearchTimer);
  logSearchTimer=window.setTimeout(goFirstLogPage, 250);
}
let logCam=null;
async function loadLogs(){
  const p=new URLSearchParams();
  const q=document.getElementById('lq').value;
  const event=document.getElementById('eventType').value;
  const zone=document.getElementById('logZone').value;
  const nvr=document.getElementById('logNvr').value;
  p.set('preset', currentLogPreset);
  p.set('page', String(currentLogPage));
  p.set('page_size', String(logPageSize));
  if(currentLogPreset==='custom'){
    p.set('from', document.getElementById('fromDate').value);
    p.set('to', document.getElementById('toDate').value);
  }
  if(q)p.set('q',q);
  if(event && event!=='all')p.set('event',event);
  if(zone)p.set('zone',zone);
  if(nvr)p.set('nvr',nvr);
  const r=await fetch('/api/camera-logs?'+p);
  const data=await r.json();
  const rows=data.items||[];
  totalLogPages=Math.max(1, Math.ceil((data.total||0)/logPageSize));
  currentLogPage=Math.min(currentLogPage, totalLogPages);
  if(data.date_from) document.getElementById('fromDate').value=data.date_from;
  if(data.date_to) document.getElementById('toDate').value=data.date_to;
  document.getElementById('sTot').textContent=data.total || 0;
  document.getElementById('sOff').textContent=rows.filter(x=>x.event==='offline').length;
  document.getElementById('sOn').textContent=rows.filter(x=>x.event==='online').length;
  document.getElementById('sShown').textContent=rows.length;
  document.getElementById('rcnt').textContent=`${data.date_from} to ${data.date_to}`;
  document.getElementById('pgInfo').textContent=`Page ${currentLogPage} of ${totalLogPages} • ${rows.length} shown • ${data.total || 0} matched`;
  document.getElementById('logPrevBtn').disabled=currentLogPage<=1;
  document.getElementById('logNextBtn').disabled=currentLogPage>=totalLogPages;
  renderLogPageNumbers();
  document.getElementById('tb').innerHTML=rows.length ? rows.map(row=>`<tr>
    <td style="font-family:monospace;font-size:11px;color:#70839b">${fmtDateTime(row.ts)}</td>
    <td><button type="button" class="btn" style="padding:0;border:none;background:none;color:#2c3e50;font-size:12px;font-weight:700" onclick="event.preventDefault();event.stopPropagation();openLogModal('${escHtml(row.ip)}')">${escHtml(row.name)||'—'}</button></td>
    <td style="font-family:monospace;font-size:11px;color:#70839b">${escHtml(row.ip||'')}</td>
    <td>${escHtml(row.zone||'—')}</td>
    <td>${escHtml(row.nvr_name||'—')}</td>
    <td><span class="badge" style="background:${row.event==='offline'?'#fdecea':'#eafaf1'};color:${row.event==='offline'?'#922b21':'#1e8449'}">${escHtml(row.event||'')}</span></td>
    <td>${row.event==='online'?fmtDuration(row.duration_s):'—'}</td>
  </tr>`).join('') : '<tr><td colspan="7" style="padding:28px;text-align:center;color:#7a879a">No camera log rows found for this range.</td></tr>';
}
function renderLogModal(c){
  logCam=c;
  document.getElementById('logNm').textContent=c.name || c.ip;
  document.getElementById('logNmDot').style.background=(c.maintenance?'#e67e22':c.online?'#27ae60':'#e74c3c');
  document.getElementById('logSb').textContent=`${c.ip}  •  ${c.location||''}, ${c.zone||''}  •  ${c.nvr_name||''} Ch.${c.nvr_channel||1}`;
  const pb=document.getElementById('logPvb');
  const browserStream=(c.stream_urls||{}).browser || (c.stream_urls||{}).mjpeg || '';
  const snapshotStream=(c.stream_urls||{}).snapshot || '';
  const playerStream=resolvePlayerStream(c.stream_urls||{});
  const rtspLink=(c.stream_urls||{}).rtsp || '';
  {% if role == 'admin' %}
  const rtspActions = rtspLink ? `<div style="display:flex;gap:8px;justify-content:center;flex-wrap:wrap;margin-top:10px">
      <a class="btn" href="${rtspLink}" style="text-decoration:none" target="_blank" rel="noopener">Open RTSP</a>
      <button type="button" class="btn" onclick="copyTextValue(decodeURIComponent('${encodeURIComponent(rtspLink)}'))">Copy RTSP URL</button>
    </div>` : '';
  {% else %}
  const rtspActions = '';
  {% endif %}
  if(c.online && !c.maintenance && c.stream_urls){
    if(playerStream){
      stopLogPreview();
      pb.innerHTML=`<div style="text-align:center;width:100%">
        <div id="logPreviewFrame" ondblclick="toggleFullscreen('logPreviewFrame')" style="width:100%;max-width:100%;aspect-ratio:4 / 3;border:1px solid #cbd5e1;border-radius:12px;overflow:hidden;cursor:zoom-in;background:#fff">
          <iframe src="${playerStream}" allow="autoplay; fullscreen; picture-in-picture" allowfullscreen style="display:block;width:100%;height:100%;border:0;background:#fff"></iframe>
        </div>
        ${rtspActions}
      </div>`;
    }else{
      pb.innerHTML=`<div style="text-align:center;width:100%">
        <div id="logPreviewFrame" ondblclick="toggleFullscreen('logPreviewFrame')" style="width:100%;border:1px solid #cbd5e1;border-radius:12px;overflow:hidden;cursor:zoom-in;background:#fff">
          <img id="logPreviewImg" style="width:100%;height:auto;max-height:65vh;object-fit:contain;display:block;background:#fff">
        </div>
        ${rtspActions}
      </div>`;
      const img=document.getElementById('logPreviewImg');
      if(browserStream){
        stopLogPreview();
        img.onerror=()=>{
          img.onerror=null;
          if(snapshotStream){
            startLogPreview(snapshotStream);
          }else{
            img.parentElement.innerHTML=`<div style="color:#aaa;font-size:12px;padding:24px;text-align:center;border:1px solid #e5e7eb;border-radius:12px;background:#fff">Live preview unavailable</div>`;
          }
        };
        img.src=browserStream;
      }else if(snapshotStream){
        img.onerror=()=>{
          img.onerror=null;
          img.parentElement.innerHTML=`<div style="color:#aaa;font-size:12px;padding:24px;text-align:center;border:1px solid #e5e7eb;border-radius:12px;background:#fff">Live preview unavailable</div>`;
        };
        startLogPreview(snapshotStream);
      }else{
        img.parentElement.innerHTML=`<div style="color:#aaa;font-size:12px;padding:24px;text-align:center;border:1px solid #e5e7eb;border-radius:12px;background:#fff">Live preview unavailable</div>`;
      }
    }
  } else if(!c.online){
    stopLogPreview();
    pb.innerHTML=`<div style="text-align:center;color:#c0392b"><div style="font-size:32px;margin-bottom:8px">🔴</div><div style="font-weight:600">Camera Offline</div><div style="font-size:11px;color:#aaa;margin-top:6px">Offline since ${fmtDateTime(c.offline_since)}</div></div>`;
  } else {
    stopLogPreview();
    pb.innerHTML=`<div style="text-align:center;color:#e67e22"><div style="font-size:32px;margin-bottom:8px">🟡</div><div style="font-weight:600">Maintenance Mode</div>${rtspActions}</div>`;
  }
  document.getElementById('logRws').innerHTML=[
    ['Zone',escHtml(c.zone||'—')],
    ['Location',escHtml(c.location||'—')],
    ['Brand',escHtml((c.brand||'').charAt(0).toUpperCase()+(c.brand||'').slice(1))],
    ['NVR / Channel',escHtml(c.nvr_name||'')+'  /  Ch.'+escHtml(c.nvr_channel||1)],
    ['Status',statusHtml(c)],
    ['Notes',escHtml(c.notes||'—')],
  ].map(([l,v])=>`<div class="irow"><span class="ilbl">${l}</span><span>${v}</span></div>`).join('');
  {% if role == 'admin' %}
  document.getElementById('logRws').innerHTML += `<div class="irow"><span class="ilbl">RTSP URL</span><span style="font-family:ui-monospace,SFMono-Regular,monospace;font-size:10px;color:var(--primary);word-break:break-all;overflow-wrap:anywhere">${escHtml((c.stream_urls||{}).rtsp||'—')}</span></div>`;
  {% endif %}
  const ev=(c.history||[]).slice(0,5).map(h=>`<div style="font-size:11px;padding:4px 0;border-bottom:1px solid #f5f5f5;color:${h.event==='offline'?'#c0392b':h.event==='online'?'#27ae60':'#888'}">${h.event==='offline'?'🔴':'🟢'} ${h.event.charAt(0).toUpperCase()+h.event.slice(1)} — ${fmtDateTime(h.ts)}${h.event==='online'&&h.duration_s?` <span style="color:#888">(${fmtDuration(h.duration_s)} downtime)</span>`:''}</div>`).join('');
  document.getElementById('logEv').innerHTML=ev||'<div style="color:#aaa;font-size:11px">No events recorded yet</div>';
  document.getElementById('logOv').classList.add('show');
  document.body.style.overflow='hidden';
}
async function openLogModal(ip){
  const r=await fetch('/api/camera/'+ip);
  const c=await r.json();
  renderLogModal(c);
}
async function refreshLogModal(){
  if(!logCam || !document.getElementById('logOv').classList.contains('show')) return;
  const r=await fetch('/api/camera/'+logCam.ip);
  const c=await r.json();
  renderLogModal(c);
}
function closeLogModal(){
  stopLogPreview();
  document.getElementById('logOv').classList.remove('show');
  document.body.style.overflow='';
  logCam=null;
}
setInterval(async()=>{await loadLogs();await refreshLogModal();}, 30000);
loadLogs();
</script>
</body></html>"""

# ── Main ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # With use_reloader=True, Werkzeug spawns a child process that actually serves
    # requests (WERKZEUG_RUN_MAIN='true'). The parent process only watches files.
    # Start the scheduler only in the child (or when reloader is off) to avoid
    # launching duplicate scheduler threads.
    _reloader_parent = os.environ.get("WERKZEUG_RUN_MAIN") != "true"
    if not _reloader_parent:
        db.init_db()
        loaded = db.load_cameras_from_csv()
        log.info("Loaded %d cameras from CSV", loaded)
        monitor.start_scheduler()
        log.info("Starting CamMonitor on http://%s:%d", WEB_HOST, WEB_PORT)
    try:
        app.run(host=WEB_HOST, port=WEB_PORT, debug=False, use_reloader=True, threaded=True)
    finally:
        if not _reloader_parent:
            monitor.stop_scheduler()
